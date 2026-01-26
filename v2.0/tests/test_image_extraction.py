"""
Unit tests for image extraction functionality.
Tests the extract_images_from_sheet function and related image handling.
"""
import pytest
import sys
from pathlib import Path
from unittest.mock import Mock, patch

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

# Import after path is set
import excel_to_md
extract_images_from_sheet = excel_to_md.extract_images_from_sheet
extract_print_area_for_csv = excel_to_md.extract_print_area_for_csv
sanitize_sheet_name = excel_to_md.sanitize_sheet_name


# ============================================================
# Tests for extract_images_from_sheet
# ============================================================

class TestExtractImagesFromSheet:
    """Tests for image extraction from worksheet."""

    def test_no_images_returns_empty_dict(self, tmp_path):
        """Worksheet without images should return empty dict."""
        # Create mock worksheet without images
        ws = Mock()
        ws._images = []

        result = extract_images_from_sheet(
            ws,
            tmp_path,
            "Sheet1",
            "test_file",
            {}
        )

        assert result == {}
        assert isinstance(result, dict)

    def test_worksheet_without_images_attribute(self, tmp_path):
        """Worksheet without _images attribute should return empty dict."""
        ws = Mock(spec=[])  # Mock without _images attribute

        result = extract_images_from_sheet(
            ws,
            tmp_path,
            "Sheet1",
            "test_file",
            {}
        )

        assert result == {}

    def test_creates_subdirectory(self, tmp_path):
        """Should create subdirectory named after markdown file with _images suffix."""
        # Create mock image to trigger directory creation
        mock_img = Mock()
        mock_img._data = Mock(return_value=b'\x89PNG\r\n\x1a\n' + b'fake png data')
        mock_img.format = 'png'

        mock_anchor = Mock()
        mock_from = Mock()
        mock_from.row = 0
        mock_from.col = 0
        mock_anchor._from = mock_from
        mock_img.anchor = mock_anchor

        ws = Mock()
        ws._images = [mock_img]
        md_basename = "my_output"

        extract_images_from_sheet(ws, tmp_path, "Sheet1", md_basename, {})

        expected_dir = tmp_path / f"{md_basename}_images"
        assert expected_dir.exists()
        assert expected_dir.is_dir()

    def test_png_image_extraction(self, tmp_path):
        """Test extraction of PNG image."""
        # Create mock image with PNG data
        mock_img = Mock()
        mock_img._data = Mock(return_value=b'\x89PNG\r\n\x1a\n' + b'fake png data')
        mock_img.format = 'png'

        # Mock anchor with position
        mock_anchor = Mock()
        mock_from = Mock()
        mock_from.row = 0  # 0-based
        mock_from.col = 0  # 0-based
        mock_anchor._from = mock_from
        mock_img.anchor = mock_anchor

        # Create mock worksheet
        ws = Mock()
        ws._images = [mock_img]

        result = extract_images_from_sheet(ws, tmp_path, "TestSheet", "output", {})

        # Should have one mapping
        assert len(result) == 1
        # Cell position should be 1-based
        assert (1, 1) in result
        # Path should be relative with _images suffix
        assert result[(1, 1)] == "output_images/TestSheet_img_1.png"

        # Verify file was created
        img_file = tmp_path / "output_images" / "TestSheet_img_1.png"
        assert img_file.exists()

    def test_jpg_image_detection_from_magic_bytes(self, tmp_path):
        """Test JPEG format detection from magic bytes."""
        # Create mock without format attribute using spec
        mock_img = Mock(spec=['_data', 'anchor'])
        # JPEG magic bytes: FF D8 FF
        mock_img._data = Mock(return_value=b'\xff\xd8\xff\xe0' + b'fake jpeg data')

        mock_anchor = Mock()
        mock_from = Mock()
        mock_from.row = 1  # 0-based -> (2, 2) in 1-based
        mock_from.col = 1
        mock_anchor._from = mock_from
        mock_img.anchor = mock_anchor

        ws = Mock()
        ws._images = [mock_img]

        result = extract_images_from_sheet(ws, tmp_path, "Sheet1", "out", {})

        assert (2, 2) in result
        assert result[(2, 2)] == "out_images/Sheet1_img_1.jpg"

        img_file = tmp_path / "out_images" / "Sheet1_img_1.jpg"
        assert img_file.exists()

    def test_gif_image_detection_from_magic_bytes(self, tmp_path):
        """Test GIF format detection from magic bytes."""
        # Create mock without format attribute using spec
        mock_img = Mock(spec=['_data', 'anchor'])
        # GIF magic bytes: GIF
        mock_img._data = Mock(return_value=b'GIF89a' + b'fake gif data')

        mock_anchor = Mock()
        mock_from = Mock()
        mock_from.row = 0
        mock_from.col = 0
        mock_anchor._from = mock_from
        mock_img.anchor = mock_anchor

        ws = Mock()
        ws._images = [mock_img]

        result = extract_images_from_sheet(ws, tmp_path, "Sheet1", "out", {})

        assert result[(1, 1)] == "out_images/Sheet1_img_1.gif"

    def test_multiple_images(self, tmp_path):
        """Test extraction of multiple images."""
        images = []
        positions = [(0, 0), (2, 3), (5, 1)]

        for idx, (row, col) in enumerate(positions):
            mock_img = Mock()
            mock_img._data = Mock(return_value=b'\x89PNG' + f'data{idx}'.encode())
            mock_img.format = 'png'

            mock_anchor = Mock()
            mock_from = Mock()
            mock_from.row = row
            mock_from.col = col
            mock_anchor._from = mock_from
            mock_img.anchor = mock_anchor

            images.append(mock_img)

        ws = Mock()
        ws._images = images

        result = extract_images_from_sheet(ws, tmp_path, "Multi", "test", {})

        # Should have 3 mappings
        assert len(result) == 3
        # Check positions (converted to 1-based)
        assert (1, 1) in result
        assert (3, 4) in result
        assert (6, 2) in result

    def test_onecellanchor_with_direct_row_col(self, tmp_path):
        """Test OneCellAnchor type with direct row/col attributes."""
        mock_img = Mock()
        mock_img._data = Mock(return_value=b'\x89PNG' + b'data')
        mock_img.format = 'png'

        # Mock anchor without _from but with direct row/col using spec
        mock_anchor = Mock(spec=['row', 'col'])
        mock_anchor.row = 3  # 0-based
        mock_anchor.col = 4  # 0-based
        mock_img.anchor = mock_anchor

        ws = Mock()
        ws._images = [mock_img]

        result = extract_images_from_sheet(ws, tmp_path, "Sheet", "out", {})

        # Should convert to 1-based
        assert (4, 5) in result

    def test_unknown_anchor_type_skipped(self, tmp_path):
        """Images with unknown anchor type should be skipped."""
        mock_img = Mock()
        mock_img._data = Mock(return_value=b'\x89PNG' + b'data')
        mock_img.format = 'png'

        # Mock anchor without _from and without row/col
        mock_anchor = Mock(spec=[])  # Empty spec
        mock_img.anchor = mock_anchor

        ws = Mock()
        ws._images = [mock_img]

        result = extract_images_from_sheet(ws, tmp_path, "Sheet", "out", {})

        # Should return empty dict (image skipped)
        assert len(result) == 0

    def test_sanitized_sheet_name_in_filename(self, tmp_path):
        """Sheet name should be sanitized in filename."""
        mock_img = Mock()
        mock_img._data = Mock(return_value=b'\x89PNG' + b'data')
        mock_img.format = 'png'

        mock_anchor = Mock()
        mock_from = Mock()
        mock_from.row = 0
        mock_from.col = 0
        mock_anchor._from = mock_from
        mock_img.anchor = mock_anchor

        ws = Mock()
        ws._images = [mock_img]

        # Use sheet name with special characters
        result = extract_images_from_sheet(ws, tmp_path, "Sheet/With:Special*Chars", "out", {})

        # Check that file was created with sanitized name
        img_file = tmp_path / "out_images" / "Sheet_With_Special_Chars_img_1.png"
        assert img_file.exists()

    def test_image_extraction_exception_handling(self, tmp_path):
        """Exceptions during image extraction should be handled gracefully."""
        mock_img = Mock()
        # Simulate exception when getting data
        mock_img._data = Mock(side_effect=Exception("Image data error"))

        ws = Mock()
        ws._images = [mock_img]

        # Should not raise exception
        result = extract_images_from_sheet(ws, tmp_path, "Sheet", "out", {})

        # Should return empty dict (failed image skipped)
        assert result == {}

    def test_default_png_extension_fallback(self, tmp_path):
        """Unknown format should default to PNG extension."""
        # Create mock without format attribute using spec
        mock_img = Mock(spec=['_data', 'anchor'])
        mock_img._data = Mock(return_value=b'unknown format data')

        mock_anchor = Mock()
        mock_from = Mock()
        mock_from.row = 0
        mock_from.col = 0
        mock_anchor._from = mock_from
        mock_img.anchor = mock_anchor

        ws = Mock()
        ws._images = [mock_img]

        result = extract_images_from_sheet(ws, tmp_path, "Sheet", "out", {})

        # Should use .png as default
        assert result[(1, 1)] == "out_images/Sheet_img_1.png"


# ============================================================
# Tests for extract_print_area_for_csv with images
# ============================================================

class TestExtractPrintAreaWithImages:
    """Tests for CSV extraction with image links."""

    @pytest.fixture
    def mock_worksheet(self):
        """Create a mock worksheet for testing."""
        ws = Mock()

        # Mock cells
        def mock_cell(row, column):
            cell = Mock()
            cell.value = f"R{row}C{column}"
            return cell

        ws.cell = mock_cell
        return ws

    @pytest.fixture
    def mock_opts(self):
        """Create mock options dictionary."""
        return {
            'strip_whitespace': True,
            'csv_apply_merge_policy': True,
        }

    def test_csv_extraction_with_image_link(self, mock_worksheet, mock_opts):
        """Cell with image should output Markdown image link."""
        area = (1, 1, 2, 2)
        merged_lookup = {}
        cell_to_image = {
            (1, 1): "output/Sheet1_img_1.png"
        }

        # Mock cell_display_value to return empty string
        with patch('excel2md.csv_export.cell_display_value', return_value=""):
            with patch('excel2md.csv_export.a1_from_rc', return_value="A1"):
                result = extract_print_area_for_csv(
                    mock_worksheet,
                    area,
                    mock_opts,
                    merged_lookup,
                    cell_to_image
                )

        # First cell should have image link
        assert result[0][0] == "![Image at A1](output/Sheet1_img_1.png)"

    def test_csv_extraction_with_image_alt_text(self, mock_worksheet, mock_opts):
        """Image link should use cell value as alt text."""
        area = (1, 1, 1, 1)
        merged_lookup = {}
        cell_to_image = {
            (1, 1): "output/image.png"
        }

        # Mock cell_display_value to return meaningful text
        with patch('excel2md.csv_export.cell_display_value', return_value="Company Logo"):
            result = extract_print_area_for_csv(
                mock_worksheet,
                area,
                mock_opts,
                merged_lookup,
                cell_to_image
            )

        # Should use cell value as alt text
        assert result[0][0] == "![Company Logo](output/image.png)"

    def test_csv_extraction_without_images(self, mock_worksheet, mock_opts):
        """CSV extraction should work normally without images."""
        area = (1, 1, 2, 2)
        merged_lookup = {}
        cell_to_image = None  # No images

        with patch('excel2md.csv_export.cell_display_value', return_value="Data"):
            with patch('excel2md.csv_export.normalize_numeric_text', side_effect=lambda x, o: x):
                result = extract_print_area_for_csv(
                    mock_worksheet,
                    area,
                    mock_opts,
                    merged_lookup,
                    cell_to_image
                )

        # Should have 2 rows, 2 columns
        assert len(result) == 2
        assert len(result[0]) == 2

    def test_csv_extraction_mixed_images_and_data(self, mock_worksheet, mock_opts):
        """CSV extraction with some cells having images and others not."""
        area = (1, 1, 2, 2)
        merged_lookup = {}
        cell_to_image = {
            (1, 1): "output/img1.png",
            (2, 2): "output/img2.jpg",
        }

        with patch('excel2md.csv_export.cell_display_value', return_value="Data"):
            with patch('excel2md.csv_export.normalize_numeric_text', side_effect=lambda x, o: x):
                result = extract_print_area_for_csv(
                    mock_worksheet,
                    area,
                    mock_opts,
                    merged_lookup,
                    cell_to_image
                )

        # Cell (1,1) should have image link
        assert "![" in result[0][0] and "](output/img1.png)" in result[0][0]
        # Cell (2,2) should have image link
        assert "![" in result[1][1] and "](output/img2.jpg)" in result[1][1]
        # Other cells should have regular data
        # Cell (1,2) - no image
        assert "![" not in result[0][1]


# ============================================================
# Integration test with real worksheet
# ============================================================

class TestImageExtractionIntegration:
    """Integration tests with real openpyxl worksheet."""

    def test_real_worksheet_without_images(self, tmp_path):
        """Real worksheet without images should work correctly."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        ws['B1'] = 'Data'

        result = extract_images_from_sheet(ws, tmp_path, "Sheet1", "test", {})

        assert result == {}

    def test_path_creation_with_nested_structure(self, tmp_path):
        """Should create nested directory structure correctly."""
        # Create mock image to trigger directory creation
        mock_img = Mock()
        mock_img._data = Mock(return_value=b'\x89PNG\r\n\x1a\n' + b'fake png data')
        mock_img.format = 'png'

        mock_anchor = Mock()
        mock_from = Mock()
        mock_from.row = 0
        mock_from.col = 0
        mock_anchor._from = mock_from
        mock_img.anchor = mock_anchor

        ws = Mock()
        ws._images = [mock_img]

        nested_basename = "nested/output/file"
        extract_images_from_sheet(ws, tmp_path, "Sheet", nested_basename, {})

        # Should create nested directory with _images suffix
        expected_dir = tmp_path / f"{nested_basename}_images"
        assert expected_dir.exists()
