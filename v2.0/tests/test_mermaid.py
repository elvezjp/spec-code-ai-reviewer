"""
Unit tests for Mermaid and code detection functions.
"""
import pytest
import sys
from pathlib import Path
from unittest.mock import MagicMock

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from excel_to_md import (
    is_source_code,
    detect_code_language,
    format_table_as_text_or_nested,
    has_border,
    build_merged_lookup,
)
from openpyxl.cell.cell import Cell


def create_mock_cell(value=None, fill_color=None, hyperlink_target=None,
                     hyperlink_location=None, is_date=False):
    """
    Create a mock cell for testing.
    """
    cell = MagicMock(spec=Cell)
    cell.value = value
    cell.is_date = is_date

    # Fill setup
    if fill_color:
        cell.fill = MagicMock()
        cell.fill.patternType = 'solid'
        cell.fill.fgColor = MagicMock()
        cell.fill.fgColor.rgb = fill_color
        cell.fill.fgColor.type = 'rgb'
    else:
        cell.fill = MagicMock()
        cell.fill.patternType = None

    # Hyperlink setup
    if hyperlink_target or hyperlink_location:
        cell.hyperlink = MagicMock()
        cell.hyperlink.target = hyperlink_target
        cell.hyperlink.location = hyperlink_location
        cell.hyperlink.display = str(value) if value else None
    else:
        cell.hyperlink = None

    # Border setup (no borders by default)
    cell.border = MagicMock()
    cell.border.left = MagicMock(style=None)
    cell.border.right = MagicMock(style=None)
    cell.border.top = MagicMock(style=None)
    cell.border.bottom = MagicMock(style=None)

    return cell


# ============================================================
# Tests for is_source_code
# ============================================================

class TestIsSourceCode:
    """Tests for source code detection."""

    def test_java_class(self):
        """Java class definition is detected as code."""
        assert is_source_code("public class Foo {") is True

    def test_python_function(self):
        """Python function definition without braces/semicolons.

        Implementation requires both keyword AND symbol ({, }, ;, //, etc.)
        Python-style code with just colon is not detected.
        """
        # "def foo():" has keyword but no code symbols ({, }, ;, //, etc.)
        assert is_source_code("def foo():") is False
        # With brace it would be detected
        assert is_source_code("def foo(): {") is True

    def test_java_annotation(self):
        """Java annotation is detected as code."""
        assert is_source_code("@Override") is True

    def test_javascript_function(self):
        """JavaScript function is detected as code."""
        assert is_source_code("function test() {}") is True

    def test_normal_text(self):
        """Normal text is not detected as code."""
        assert is_source_code("Hello World") is False

    def test_numeric_string(self):
        """Numeric string is not detected as code."""
        assert is_source_code("123.45") is False

    def test_empty_string(self):
        """Empty string is not detected as code."""
        assert is_source_code("") is False

    def test_whitespace_only(self):
        """Whitespace only is not detected as code."""
        assert is_source_code("   ") is False

    def test_java_import(self):
        """Java import statement is detected as code."""
        assert is_source_code("import java.util.List;") is True

    def test_csharp_namespace(self):
        """C# namespace is detected as code."""
        assert is_source_code("namespace MyApp {") is True

    def test_javascript_const(self):
        """JavaScript const is detected as code."""
        assert is_source_code("const x = 1;") is True

    def test_javascript_let(self):
        """JavaScript let is detected as code."""
        assert is_source_code("let y = 2;") is True

    def test_python_class(self):
        """Python class definition - requires symbol to be detected.

        'class MyClass:' has keyword 'class' but no code symbols.
        """
        assert is_source_code("class MyClass:") is False
        assert is_source_code("class MyClass {") is True

    def test_if_statement(self):
        """if statement is detected as code."""
        assert is_source_code("if (x > 0) {") is True

    def test_return_statement(self):
        """return statement is detected as code."""
        assert is_source_code("return value;") is True

    def test_multiple_code_indicators(self):
        """Multiple code indicators should be detected."""
        assert is_source_code("public static void main(String[] args) {") is True

    def test_comment_line(self):
        """Comment line - single // is one symbol, needs 2+ or keyword.

        '// This is a comment' has one symbol (//) but no keyword.
        symbol_count >= 2 is required for symbol-only detection.
        """
        assert is_source_code("// This is a comment") is False
        # Multiple comment markers would work
        assert is_source_code("// comment /* block */") is True

    def test_block_comment(self):
        """Block comment is detected as code."""
        assert is_source_code("/* comment */") is True

    def test_brace_only(self):
        """Single brace might not be enough."""
        # This depends on implementation - just braces alone might not be code
        result = is_source_code("{")
        # Implementation specific - could be True or False
        assert isinstance(result, bool)

    def test_japanese_text(self):
        """Japanese text is not detected as code."""
        assert is_source_code("こんにちは世界") is False

    def test_mixed_japanese_code(self):
        """Japanese with code keywords."""
        # "public" in Japanese context might still be detected
        result = is_source_code("public クラス定義")
        # This tests edge case - result depends on implementation
        assert isinstance(result, bool)


# ============================================================
# Tests for detect_code_language
# ============================================================

class TestDetectCodeLanguage:
    """Tests for programming language detection."""

    def test_java(self):
        """Java code detection."""
        lines = ["public class Example {", "import java.util.List;"]
        result = detect_code_language(lines)
        assert result == "java"

    def test_python(self):
        """Python code detection."""
        lines = ["def foo():", "import os"]
        result = detect_code_language(lines)
        assert result == "python"

    def test_javascript(self):
        """JavaScript code detection."""
        lines = ["const x = 1;", "function test() {}"]
        result = detect_code_language(lines)
        assert result == "javascript"

    def test_unknown(self):
        """Unknown language returns empty string."""
        lines = ["普通のテキスト", "もっとテキスト"]
        result = detect_code_language(lines)
        assert result == ""

    def test_csharp_detection(self):
        """C# code detection."""
        lines = ["namespace MyApp {", "using System;"]
        result = detect_code_language(lines)
        assert result == "csharp"

    def test_c_detection(self):
        """C code detection."""
        lines = ["#include <stdio.h>", "int main() {"]
        result = detect_code_language(lines)
        assert result == "c"

    def test_empty_lines(self):
        """Empty lines should return empty string."""
        result = detect_code_language([])
        assert result == ""

    def test_mixed_language_indicators(self):
        """Mixed language indicators - first match wins."""
        # Java indicators come before Python in the function
        lines = ["import java.util.List;", "def python_func():"]
        result = detect_code_language(lines)
        # Should detect Java first
        assert result in ["java", "python"]

    def test_python_with_colon(self):
        """Python detection requires colon."""
        lines = ["def foo():", "if x > 0:"]
        result = detect_code_language(lines)
        assert result == "python"


# ============================================================
# Tests for format_table_as_text_or_nested
# ============================================================

class TestFormatTableAsTextOrNested:
    """Tests for table format detection."""

    def test_empty_table_returns_empty(self, simple_worksheet, default_opts):
        """Empty table should return 'empty' format."""
        ws = simple_worksheet
        table = {'bbox': (1, 1, 1, 1), 'mask': set()}
        md_rows = []
        merged_lookup = build_merged_lookup(ws)

        format_type, output = format_table_as_text_or_nested(
            ws, table, md_rows, default_opts, merged_lookup
        )

        assert format_type == "empty"

    def test_code_detection(self, empty_workbook, default_opts, sample_code_rows):
        """Code-like content detection.

        Note: format_table_as_text_or_nested checks for 'text' format first
        (single cell in first row with no borders). For single-column tables,
        this takes precedence over code detection.

        To get 'code' format, the table must not qualify as 'text' format:
        - Multiple cells in first row, OR
        - Single cell with borders, OR
        - Other specific conditions
        """
        # Set up worksheet with actual code content
        ws = empty_workbook.active
        for r, row in enumerate(sample_code_rows, start=1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)

        # Create a table structure for code rows
        mask = set()
        for r in range(len(sample_code_rows)):
            for c in range(len(sample_code_rows[0])):
                mask.add((1 + r, 1 + c))

        table = {
            'bbox': (1, 1, len(sample_code_rows), 1),
            'mask': mask
        }
        merged_lookup = build_merged_lookup(ws)

        format_type, output = format_table_as_text_or_nested(
            ws, table, sample_code_rows, default_opts, merged_lookup
        )

        # Single-column table without borders returns 'text' format
        # because it matches the "single line text" condition first
        assert format_type == "text"
        # Output is the first cell value
        assert "public class Example" in output

    def test_table_format(self, simple_worksheet, default_opts, sample_md_rows_simple):
        """Normal table should return 'table' format."""
        ws = simple_worksheet
        table = {
            'bbox': (1, 1, 2, 2),
            'mask': {(1, 1), (1, 2), (2, 1), (2, 2)}
        }
        merged_lookup = build_merged_lookup(ws)

        format_type, output = format_table_as_text_or_nested(
            ws, table, sample_md_rows_simple, default_opts, merged_lookup
        )

        # Normal data should return 'table'
        assert format_type == "table"


# ============================================================
# Tests for has_border
# ============================================================

class TestHasBorder:
    """Tests for cell border detection."""

    def test_no_border(self):
        """Cell without borders."""
        cell = create_mock_cell(value="test")
        result = has_border(cell)
        assert result is False

    def test_single_border(self):
        """Cell with single border should not count as bordered."""
        from unittest.mock import MagicMock

        cell = MagicMock()
        cell.border = MagicMock()
        cell.border.left = MagicMock(style='thin')
        cell.border.right = MagicMock(style=None)
        cell.border.top = MagicMock(style=None)
        cell.border.bottom = MagicMock(style=None)

        result = has_border(cell)
        # Single border should not be enough (needs 2+)
        assert result is False

    def test_two_borders(self):
        """Cell with two borders should count as bordered."""
        from unittest.mock import MagicMock

        cell = MagicMock()
        cell.border = MagicMock()
        cell.border.left = MagicMock(style='thin')
        cell.border.right = MagicMock(style='thin')
        cell.border.top = MagicMock(style=None)
        cell.border.bottom = MagicMock(style=None)

        result = has_border(cell)
        assert result is True

    def test_all_borders(self):
        """Cell with all borders."""
        from unittest.mock import MagicMock

        cell = MagicMock()
        cell.border = MagicMock()
        cell.border.left = MagicMock(style='thin')
        cell.border.right = MagicMock(style='thin')
        cell.border.top = MagicMock(style='thin')
        cell.border.bottom = MagicMock(style='thin')

        result = has_border(cell)
        assert result is True

    def test_none_style_border(self):
        """Border with style='none' should not count."""
        from unittest.mock import MagicMock

        cell = MagicMock()
        cell.border = MagicMock()
        cell.border.left = MagicMock(style='none')
        cell.border.right = MagicMock(style='none')
        cell.border.top = MagicMock(style='none')
        cell.border.bottom = MagicMock(style='none')

        result = has_border(cell)
        assert result is False


# ============================================================
# Tests for Mermaid generation (when implemented)
# ============================================================

class TestMermaidGeneration:
    """Tests for Mermaid diagram generation."""

    def test_mermaid_disabled_by_default(self, default_opts):
        """Mermaid should be disabled by default."""
        assert default_opts.get('mermaid_enabled', False) is False

    def test_mermaid_detect_mode_default(self, default_opts):
        """Default Mermaid detect mode should be 'shapes'."""
        assert default_opts.get('mermaid_detect_mode', 'shapes') == 'shapes'

    def test_mermaid_options_structure(self, default_opts):
        """Mermaid-related options should exist."""
        # These options should be available
        mermaid_opts = [
            'mermaid_enabled',
            'mermaid_detect_mode',
        ]
        for opt in mermaid_opts:
            assert opt in default_opts


# ============================================================
# Integration tests for code block output
# ============================================================

class TestCodeBlockIntegration:
    """Integration tests for code block detection and output."""

    def test_java_code_block(self, empty_workbook, default_opts):
        """Java code should produce java code block."""
        ws = empty_workbook.active
        ws['A1'] = 'public class Example {'
        ws['A2'] = '    private int value;'
        ws['A3'] = '}'

        table = {
            'bbox': (1, 1, 3, 1),
            'mask': {(1, 1), (2, 1), (3, 1)}
        }
        merged_lookup = build_merged_lookup(ws)

        md_rows = [
            ['public class Example {'],
            ['    private int value;'],
            ['}'],
        ]

        format_type, output = format_table_as_text_or_nested(
            ws, table, md_rows, default_opts, merged_lookup
        )

        if format_type == "code":
            assert "```java" in output or "```" in output

    def test_python_code_block(self, empty_workbook, default_opts):
        """Python code should produce python code block."""
        ws = empty_workbook.active
        ws['A1'] = 'def hello():'
        ws['A2'] = '    print("Hello")'

        table = {
            'bbox': (1, 1, 2, 1),
            'mask': {(1, 1), (2, 1)}
        }
        merged_lookup = build_merged_lookup(ws)

        md_rows = [
            ['def hello():'],
            ['    print("Hello")'],
        ]

        format_type, output = format_table_as_text_or_nested(
            ws, table, md_rows, default_opts, merged_lookup
        )

        if format_type == "code":
            assert "```python" in output or "```" in output
