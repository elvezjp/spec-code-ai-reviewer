"""Excel2mdToolの変換機能テスト。"""

from io import BytesIO

import pytest
from openpyxl import Workbook

from app.markdown_tools import get_available_tools, get_markdown_tool
from app.markdown_tools.excel2md_tool import Excel2mdTool


class TestExcel2mdToolProperties:
    """Excel2mdToolのプロパティテスト。"""

    def test_name_property(self):
        """nameプロパティが正しい値を返す。"""
        tool = Excel2mdTool()

        assert tool.name == "excel2md"

    def test_display_name_property(self):
        """display_nameプロパティが正しい値を返す。"""
        tool = Excel2mdTool()

        assert tool.display_name == "excel2md (CSV)"


class TestExcel2mdToolRegistry:
    """Excel2mdToolのレジストリ統合テスト。"""

    def test_tool_registered_in_registry(self):
        """ツールがレジストリに登録されている。"""
        tools = get_available_tools()
        tool_names = [t["name"] for t in tools]

        assert "excel2md" in tool_names

    def test_get_tool_by_name(self):
        """get_markdown_tool("excel2md")でツールを取得できる。"""
        tool = get_markdown_tool("excel2md")

        assert isinstance(tool, Excel2mdTool)

    def test_case_insensitive_access(self):
        """大文字小文字を区別しないアクセス。"""
        tool_lower = get_markdown_tool("excel2md")
        tool_upper = get_markdown_tool("EXCEL2MD")
        tool_mixed = get_markdown_tool("Excel2Md")

        assert isinstance(tool_lower, Excel2mdTool)
        assert isinstance(tool_upper, Excel2mdTool)
        assert isinstance(tool_mixed, Excel2mdTool)


class TestExcel2mdToolConvert:
    """Excel2mdTool.convert() メソッドのテスト。"""

    @pytest.fixture
    def tool(self):
        """Excel2mdToolインスタンスを提供。"""
        return Excel2mdTool()

    @pytest.fixture
    def sample_xlsx_content(self) -> bytes:
        """テスト用の最小限のxlsxファイルを生成。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Header1"
        ws["B1"] = "Header2"
        ws["A2"] = "Value1"
        ws["B2"] = "Value2"

        # 印刷領域を設定
        ws.print_area = "A1:B2"

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def test_convert_returns_markdown_string(self, tool, sample_xlsx_content):
        """convert()がMarkdown文字列を返す。"""
        result = tool.convert(sample_xlsx_content, "test.xlsx")

        assert isinstance(result, str)
        assert len(result) > 0

    def test_convert_contains_csv_code_block(self, tool, sample_xlsx_content):
        """変換結果にCSVコードブロックが含まれる。"""
        result = tool.convert(sample_xlsx_content, "test.xlsx")

        assert "```csv" in result

    def test_convert_contains_sheet_data(self, tool, sample_xlsx_content):
        """変換結果にシートのデータが含まれる。"""
        result = tool.convert(sample_xlsx_content, "test.xlsx")

        # CSVマークダウンにはシート内容が含まれるはず
        assert "Header1" in result or "header1" in result.lower()

    def test_convert_contains_overview_section(self, tool, sample_xlsx_content):
        """変換結果に概要セクションが含まれる。"""
        result = tool.convert(sample_xlsx_content, "test.xlsx")

        # spec.mdで定義されている概要セクション
        assert "概要" in result

    def test_convert_no_metadata_section(self, tool, sample_xlsx_content):
        """変換結果に検証用メタデータセクションが含まれない（--no-csv-include-metadata）。"""
        result = tool.convert(sample_xlsx_content, "test.xlsx")

        # --no-csv-include-metadataオプションによりメタデータセクションは含まれない
        assert "検証用メタデータ" not in result

    def test_convert_via_registry(self, sample_xlsx_content):
        """レジストリ経由で取得したツールで変換できる。"""
        tool = get_markdown_tool("excel2md")
        result = tool.convert(sample_xlsx_content, "test.xlsx")

        assert "```csv" in result

    def test_convert_with_multiple_sheets(self, tool):
        """複数シートのExcelを変換できる。"""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "Data1"
        ws1.print_area = "A1:A1"

        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "Data2"
        ws2.print_area = "A1:A1"

        buffer = BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()

        result = tool.convert(content, "multi_sheet.xlsx")

        assert isinstance(result, str)
        assert len(result) > 0
        # 両方のシートが含まれている
        assert "Sheet1" in result or "Data1" in result
        assert "Sheet2" in result or "Data2" in result
