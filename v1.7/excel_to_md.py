# -*- coding: utf-8 -*-
"""
Excel -> Markdown Converter
Spec Compliance: v1.7 (CSV Markdown Mode Extensions)
File: excel_to_md.py

更新点（v1.7 / 2025-12-25）:
- v1.6の全機能を継承（ハイパーリンク平文出力モード、シート分割出力機能）
- CSVマークダウンモードの拡張機能を追加
  - --csv-include-description: 概要セクション（説明文）の出力を制御（デフォルト: ON）
  - CSVマークダウンでも --mermaid-enabled が有効に（mermaid_detect_mode="shapes" の場合のみ）
  - 複数ファイルを変換・結合する際のトークン数削減に対応

更新点（v1.6 / 2025-11-XX）:
- ハイパーリンク平文出力モード（inline_plain）を追加
- シート分割出力機能（--split-by-sheet）を追加

更新点（v1.5 / 2025-11-11）:
- v1.4の全機能を継承（Mermaidフローチャート変換、シェイプ検出）
- §⑫ CSVマークダウン出力機能を追加
"""

from pathlib import Path
import argparse
import sys
import re
import unicodedata
from typing import List, Tuple, Optional, Dict, Set

VERSION = "1.7"  # SPEC version tag

MD_RESERVED_SAFE = r"\\|\||\*|_|~|#|>|\[|\]|\(|\)|\{|\}|\+|\-|\.|!|`"
MD_ESCAPE_RE = re.compile(MD_RESERVED_SAFE)

NUMERIC_PATTERN = re.compile(
    r"""
^\s*
(\()?                                   # optional opening parenthesis
(?P<sign>[+-])?                         # optional sign
(?P<currency>[¥$€£₩])?                  # optional currency
(?P<int>\d{1,3}(?:[,，]\d{3})*|\d+)     # integer part
(?:[\.．](?P<frac>\d+))?                # optional fractional part
(?:[eE](?P<exp>[+-]?\d+))?              # optional exponent
(\))?                                   # optional closing parenthesis
\s*(?P<pct>%)?
\s*$
""",
    re.VERBOSE,
)

WHITESPACE_CHARS = {
    "\t", "\u0020", "\u00A0", "\u1680", "\u180E", "\u2000", "\u2001",
    "\u2002", "\u2003", "\u2004", "\u2005", "\u2006", "\u2007",
    "\u2008", "\u2009", "\u200A", "\u202F", "\u205F", "\u3000",
}

CONTROL_REMOVE_RANGES = [
    (0x0000, 0x0008),
    (0x000B, 0x000C),
    (0x000E, 0x001F),
    (0x007F, 0x007F),
    (0x0080, 0x009F),
    (0x00AD, 0x00AD),
    (0x200B, 0x200D),
    (0x2060, 0x2060),
    (0x2066, 0x2069),
]


# ===== v1.4 additions: Mermaid + dispatcher =====
import re as _re
import unicodedata as _unicodedata
import zipfile as _zipfile
import xml.etree.ElementTree as _ET
from typing import Dict, Tuple, Optional, List, Set, Any

# DrawingML名前空間
_DRAWINGML_NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
}

def _v14_normalize_header_name(s: str) -> str:
    s = s or ""
    s = _unicodedata.normalize("NFKC", s)
    s = s.casefold()
    s = " ".join(s.split())
    return s

def _v14_sanitize_node_id(name: str) -> str:
    s = _unicodedata.normalize("NFKC", name or "")
    # keep only ascii letters, digits, underscore. Replace others with underscore.
    s = "".join(ch if ("a"<=ch<="z" or "A"<=ch<="Z" or "0"<=ch<="9" or ch=="_") else "_" for ch in s)
    if not s:
        s = "_"
    if s[0].isdigit():
        s = "_" + s
    # collapse multiple underscores
    s = _re.sub(r"_+", "_", s)
    return s

def is_code_block(md_rows) -> bool:
    """Extracted code detection using existing is_source_code() logic per §A.x+1.6.

    Uses the same logic as format_table_as_text_or_nested() to maintain compatibility.
    """
    if not md_rows:
        return False

    code_lines = []
    is_code = False

    for row in md_rows:
        if not row:
            if is_code:
                code_lines.append("")
            continue

        # Get the first non-empty cell value
        row_text = ""
        for val in row:
            if val and str(val).strip():
                row_text = str(val).strip()
                break

        if row_text and is_source_code(row_text):
            code_lines.append(row_text)
            is_code = True
        elif is_code and row_text:
            # If we've started a code block, continue collecting lines
            code_lines.append(row_text)
        elif is_code and not row_text:
            # Empty line in code block - preserve it
            code_lines.append("")

    # If we detected code, return True
    return is_code and len(code_lines) > 0

def _v14_resolve_columns_by_name(header_row, mapping):
    norm_map = {k: _v14_normalize_header_name(v) for k, v in mapping.items() if v}
    index = {}
    dups = set()
    for i, name in enumerate(header_row):
        key = _v14_normalize_header_name(str(name or ""))
        for want_k, want in norm_map.items():
            if key == want and want_k not in index:
                index[want_k] = i
            elif key == want and want_k in index:
                dups.add(want_k)
    if dups:
        warn(f"duplicate header names detected for {sorted(list(dups))}: first occurrence will be used")
    if "from" not in index or "to" not in index:
        return None
    return index

def _v14_extract_shapes_to_mermaid(xlsx_path: str, ws, opts) -> Optional[str]:
    """Extract shapes from DrawingML and build Mermaid flowchart per §⑦''.

    Args:
        xlsx_path: Path to Excel file
        ws: Worksheet object
        opts: Options dictionary

    Returns:
        Mermaid code block string or None if no shapes detected
    """
    try:
        # Build cell grid {(row, col): value} (0-based)
        grid: Dict[Tuple[int, int], str] = {}
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for c_idx, val in enumerate(row, start=1):
                if val is not None and str(val).strip():
                    grid[(r_idx-1, c_idx-1)] = str(val).strip()  # 0-based

        # Open Excel as ZIP and find DrawingML files
        z = _zipfile.ZipFile(xlsx_path, 'r')

        # Find the drawing file corresponding to this sheet per §⑦''
        # 1. Get sheet name from worksheet object
        sheet_name = ws.title

        # 2. Find sheet ID from workbook.xml
        sheet_id = None
        try:
            wb_xml = z.read('xl/workbook.xml').decode('utf-8')
            wb_root = _ET.fromstring(wb_xml)
            ns_wb = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                     'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
            sheets = wb_root.findall('.//main:sheet', ns_wb)
            for sheet in sheets:
                if sheet.get('name') == sheet_name:
                    # Get rId from relationship
                    r_id = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                    if r_id:
                        # Find sheet index from workbook.xml.rels
                        wb_rels_xml = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
                        wb_rels_root = _ET.fromstring(wb_rels_xml)
                        ns_rel = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                        for rel in wb_rels_root.findall('.//r:Relationship', ns_rel):
                            if rel.get('Id') == r_id:
                                target = rel.get('Target')
                                # Extract sheet number from target (e.g., "worksheets/sheet1.xml" -> "1")
                                if 'sheet' in target:
                                    # Find the number after "sheet" and before ".xml"
                                    match = re.search(r'sheet(\d+)\.xml', target)
                                    if match:
                                        sheet_id = match.group(1)
                                break
                    break
        except Exception as e:
            warn(f"Failed to find sheet ID for '{sheet_name}': {e}")

        # 3. Find drawing file from sheet relationship file
        drawing_path = None
        if sheet_id:
            sheet_rel_path = f"xl/worksheets/_rels/sheet{sheet_id}.xml.rels"
            if sheet_rel_path in z.namelist():
                try:
                    rels_xml = z.read(sheet_rel_path)
                    rels_root = _ET.fromstring(rels_xml)
                    ns_rel = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                    for rel in rels_root.findall('.//r:Relationship', ns_rel):
                        if rel.get('Type') == 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing':
                            target = rel.get('Target')
                            # Resolve relative path (xl/worksheets/基準)
                            if target.startswith('../'):
                                drawing_path = target.replace('../', 'xl/')
                            else:
                                drawing_path = f"xl/worksheets/{target}"
                            break
                except Exception as e:
                    warn(f"Failed to parse relationship file '{sheet_rel_path}': {e}")

        if not drawing_path or drawing_path not in z.namelist():
            z.close()
            return None  # No drawing file for this sheet

        # 4. Parse the corresponding DrawingML file
        xml_data = z.read(drawing_path)
        root = _ET.fromstring(xml_data)

        nodes: List[Dict[str, Any]] = []
        edges: List[Dict[str, Any]] = []
        tmp_id = 1

        def get_id(cnv):
            if cnv is not None and "id" in cnv.attrib:
                return f"s{cnv.attrib['id']}"
            nonlocal tmp_id
            tmp_id += 1
            return f"local{tmp_id}"

        def get_text_from_txBody(sp):
            """Get text from txBody element per §⑦''.

            Note: In Excel DrawingML, text is stored in xdr:txBody (not a:txBody).
            The structure is: xdr:txBody -> a:p -> a:r -> a:t
            """
            texts = []
            # First try xdr:txBody (correct structure in Excel)
            for txbody in sp.findall(".//xdr:txBody", _DRAWINGML_NS):
                for p in txbody.findall(".//a:p", _DRAWINGML_NS):
                    runs = []
                    for r in p.findall(".//a:r", _DRAWINGML_NS):
                        t = r.find("a:t", _DRAWINGML_NS)
                        if t is not None and t.text:
                            runs.append(t.text)
                    if runs:
                        texts.append("".join(runs))
            # Fallback: try a:txBody (for compatibility)
            if not texts:
                for txbody in sp.findall(".//a:txBody", _DRAWINGML_NS):
                    for p in txbody.findall(".//a:p", _DRAWINGML_NS):
                        runs = []
                        for r in p.findall(".//a:r", _DRAWINGML_NS):
                            t = r.find("a:t", _DRAWINGML_NS)
                            if t is not None and t.text:
                                runs.append(t.text)
                        if runs:
                            texts.append("".join(runs))
            return "\n".join([t.strip() for t in texts if t and t.strip()])

        def cell_bbox(anc):
            def int_or_none(e):
                return int(e.text) if e is not None and e.text and e.text.isdigit() else None
            fr_c = anc.find(".//xdr:from/xdr:col", _DRAWINGML_NS)
            fr_r = anc.find(".//xdr:from/xdr:row", _DRAWINGML_NS)
            to_c = anc.find(".//xdr:to/xdr:col", _DRAWINGML_NS)
            to_r = anc.find(".//xdr:to/xdr:row", _DRAWINGML_NS)
            if all(x is not None for x in (fr_c, fr_r, to_c, to_r)):
                return [int_or_none(fr_r), int_or_none(fr_c), int_or_none(to_r), int_or_none(to_c)]
            return None

        def text_from_bbox(bbox):
            if not bbox or None in bbox:
                return ''
            r1, c1, r2, c2 = bbox
            pad = 1
            texts = []
            for r in range(max(0, r1-pad), r2+pad+1):
                for c in range(max(0, c1-pad), c2+pad+1):
                    v = grid.get((r, c))
                    if v:
                        texts.append(v)
            if texts:
                seen = set()
                uniq = [t for t in texts if t not in seen and not seen.add(t)]
                longest = max(uniq, key=len)
                return longest if len(longest) >= 4 else "\n".join(uniq[:4])
            return ''

        # Scan anchors
        anchors = root.findall(".//xdr:twoCellAnchor", _DRAWINGML_NS) + root.findall(".//xdr:oneCellAnchor", _DRAWINGML_NS)

        for anc in anchors:
            sp = anc.find(".//xdr:sp", _DRAWINGML_NS)
            cxn = anc.find(".//xdr:cxnSp", _DRAWINGML_NS)
            bbox = cell_bbox(anc)

            if sp is not None:
                cNvPr = sp.find(".//xdr:cNvPr", _DRAWINGML_NS)
                _id = get_id(cNvPr)
                name = cNvPr.get("name") if cNvPr is not None else None

                # テキスト取得（優先順位1: txBody、優先順位2: bbox近傍、優先順位3: 名前）
                text = get_text_from_txBody(sp)
                if not text:
                    text = text_from_bbox(bbox)
                if not text:
                    text = name or ""

                # 形状種類の判定
                prstGeom = sp.find(".//a:prstGeom", _DRAWINGML_NS)
                prst = prstGeom.get("prst") if prstGeom is not None else None
                shape_type = "process"  # デフォルト
                if prst:
                    if prst in ("flowChartDecision", "diamond"):
                        shape_type = "decision"
                    elif prst in ("flowChartTerminator", "ellipse", "roundRect"):
                        shape_type = "terminator"
                    elif prst in ("flowChartInputOutput", "trapezoid"):
                        shape_type = "io"
                    elif prst in ("flowChartPreparation", "hexagon"):
                        shape_type = "prep"
                    elif prst == "flowChartManualOperation":
                        shape_type = "manual"
                    elif prst == "flowChartDocument":
                        shape_type = "document"
                    elif prst == "flowChartConnector":
                        shape_type = "connector"

                nodes.append({"id": _id, "name": name, "text": text, "bbox": bbox, "type": shape_type})
            elif cxn is not None:
                st = cxn.find(".//xdr:stCxn", _DRAWINGML_NS)
                ed = cxn.find(".//xdr:endCxn", _DRAWINGML_NS)
                st_id = f"s{st.get('id')}" if st is not None and st.get("id") else None
                ed_id = f"s{ed.get('id')}" if ed is not None and ed.get("id") else None
                if st_id and ed_id:
                    edges.append({"from": st_id, "to": ed_id})

        z.close()

        if not nodes:
            return None

        # Infer edges if insufficient
        valid_edges = [e for e in edges if e.get('from') and e.get('to')]
        if len(valid_edges) < max(2, int(0.3*len(nodes))):
            edges = valid_edges + _v14_infer_edges(nodes, edges)

        # Build Mermaid code
        direction = opts.get("mermaid_direction", "TD")
        node_id_policy = opts.get("mermaid_node_id_policy", "auto")

        def escape_mermaid_text(text: str) -> str:
            """Mermaid表示名の特殊文字をHTMLエンティティに変換 per §⑦''"""
            if not text:
                return ""
            # 主要な特殊文字をHTMLエンティティに変換
            text = text.replace("[", "&#91;")
            text = text.replace("]", "&#93;")
            text = text.replace("{", "&#123;")
            text = text.replace("}", "&#125;")
            text = text.replace("|", "&#124;")
            text = text.replace('"', "&quot;")
            text = text.replace("'", "&#39;")
            return text

        def format_node(nid: str, display: str, shape_type: str) -> str:
            """シェイプ種類に応じたMermaidノード形式を生成 per §⑦''"""
            display_escaped = escape_mermaid_text(display)
            if shape_type == "decision":
                return f'  {nid}{{"{display_escaped}"}}'
            elif shape_type == "terminator":
                return f'  {nid}(["{display_escaped}"])'
            elif shape_type == "io":
                return f'  {nid}[("{display_escaped}")]'
            elif shape_type == "prep":
                return f'  {nid}[{{"{display_escaped}"}}]'
            elif shape_type in ("manual", "document"):
                return f'  {nid}[("{display_escaped}")]'
            else:  # process, connector, その他
                return f'  {nid}["{display_escaped}"]'

        lines = []
        lines.append("```mermaid")
        lines.append(f"flowchart {direction}")

        # Map original IDs to sanitized IDs
        node_map = {}
        for n in nodes:
            display = n["text"] or n["name"] or n["id"]
            shape_type = n.get("type", "process")

            # ノードID生成 per §⑦''
            if node_id_policy == "shape_id":
                # Excel描画IDをそのまま使用（s{id}形式）
                nid = n["id"]  # 既に "s{id}" 形式で格納されている
            elif node_id_policy == "auto":
                # 表示名をサニタイズしてID化
                nid = _v14_sanitize_node_id(str(display))
                # Handle duplicates
                base_nid = nid
                counter = 2
                while nid in node_map.values():
                    nid = f"{base_nid}_{counter}"
                    counter += 1
            else:  # explicit（将来拡張）
                nid = n["id"]

            node_map[n["id"]] = nid
            lines.append(format_node(nid, display, shape_type))

        # Add edges
        for e in edges:
            from_id = node_map.get(e.get("from"))
            to_id = node_map.get(e.get("to"))
            if from_id and to_id:
                # 推論されたエッジかどうかで形式を変更（任意）
                inferred = e.get("inferred", False)
                if inferred:
                    lines.append(f'  {from_id} -.->|inferred| {to_id}')
                else:
                    lines.append(f'  {from_id} --> {to_id}')

        lines.append("```")
        return "\n".join(lines)

    except Exception as e:
        warn(f"Shape extraction failed: {e}")
        return None

def _v14_infer_edges(nodes: List[Dict], exist_edges: List[Dict], max_out: int = 2, v_bias: float = 1.0) -> List[Dict]:
    """Infer edges based on vertical proximity per §⑦''."""
    def center_of(bbox):
        if not bbox or None in bbox:
            return None
        r1, c1, r2, c2 = bbox
        return ((r1 + r2) / 2.0, (c1 + c2) / 2.0)

    node_list = [n for n in nodes if n.get("bbox")]
    centers = {n["id"]: center_of(n.get("bbox")) for n in node_list}
    existing_pairs = {(e.get("from"), e.get("to")) for e in exist_edges if e.get("from") and e.get("to")}
    new_edges = []

    for n in node_list:
        nid = n["id"]
        c = centers.get(nid)
        if not c:
            continue
        dists = []
        for m in node_list:
            mid = m["id"]
            if mid == nid:
                continue
            mc = centers.get(mid)
            if not mc:
                continue
            dr = mc[0] - c[0]
            dc = abs(mc[1] - c[1])
            if dr >= 0:  # prefer downward flow
                dist = dr * v_bias + dc
                dists.append((dist, mid))
        dists.sort(key=lambda x: x[0])
        out = 0
        for _, mid in dists:
            if out >= max_out:
                break
            pair = (nid, mid)
            if pair in existing_pairs:
                continue
            new_edges.append({"from": nid, "to": mid, "bbox": None, "inferred": True})
            existing_pairs.add(pair)
            out += 1

    return new_edges

def is_flow_table(md_rows, opts):
    """Return (True, colmap) if table looks like flow table. colmap may be None for heuristic."""
    detect_mode = opts.get("mermaid_detect_mode","shapes")
    if detect_mode == "none":
        return False, None
    if not md_rows or not md_rows[0]:
        return False, None

    # column_headers mode
    if detect_mode == "column_headers":
        # §7.0: header_detection=none の場合、column_headers は適用不可
        header_detection = opts.get("header_detection", True)
        if not header_detection:  # header_detection=none (False)
            return False, None

        header = md_rows[0]
        # §7.0: 先頭行が空（全空セル）の場合はヘッダ無しとみなし、column_headers 判定は不成立
        def _is_empty_cell(v): return not (v and str(v).strip())
        if all(_is_empty_cell(cell) for cell in header):
            return False, None

        mapping = opts.get("mermaid_columns", {"from":"From","to":"To","label":"Label","group":None,"note":None})
        colmap = _v14_resolve_columns_by_name(header, mapping)
        if colmap:
            return True, colmap
        else:
            return False, None

    # heuristic mode
    # §7.0: コード形式（優先度1）に一致する場合は適用しない（誤検出回避）
    if is_code_block(md_rows):
        return False, None

    # thresholds
    min_rows = int(opts.get("mermaid_heuristic_min_rows", 3))
    arrow_ratio = float(opts.get("mermaid_heuristic_arrow_ratio", 0.3))
    ratio_min = float(opts.get("mermaid_heuristic_len_median_ratio_min", 0.4))
    ratio_max = float(opts.get("mermaid_heuristic_len_median_ratio_max", 2.5))
    data = md_rows[1:] if len(md_rows)>1 else []
    if len(md_rows[0]) < 2:
        return False, None
    # count rows with first two columns non-empty (先頭2列: 列インデックス 0 と 1)
    def _val(x): return (str(x).strip() if x is not None else "")
    nonempty12 = [r for r in data if len(r)>=2 and _val(r[0]) and _val(r[1])]
    if len(nonempty12) < min_rows:
        return False, None
    # arrow presence ratio: データ行全体に対して計算（§A.x+1.3）
    arrow_lines = 0
    pat = _re.compile(r"(->|→|⇒)")
    for r in data:
        if any(pat.search(str(c or "")) for c in r):
            arrow_lines += 1
    if len(data) == 0 or (arrow_lines / len(data)) < arrow_ratio:
        return False, None
    # median length ratio (空列削除前の元の列インデックス 0 と 1 を使用)
    import statistics as _stats
    def _len_nfkc(s): return len(_unicodedata.normalize("NFKC", str(s or "")))
    l0 = [_len_nfkc(r[0]) for r in data if len(r) >= 1]
    l1 = [_len_nfkc(r[1]) for r in data if len(r) >= 2]
    if not l0 or not l1:
        return False, None
    med0 = _stats.median(l0)
    med1 = _stats.median(l1) if any(l1) else 1
    ratio = med0 / max(med1, 1)  # ゼロ割防止
    if ratio < ratio_min or ratio > ratio_max:
        return False, None
    # heuristic assumes columns: 0,1,(2 label),(3 group),(4 note)
    colmap = {"from":0, "to":1, "label":2 if len(md_rows[0])>2 else None, "group":3 if len(md_rows[0])>3 else None, "note":4 if len(md_rows[0])>4 else None}
    return True, colmap

def build_mermaid(md_rows, opts, colmap):
    """Build mermaid flowchart codeblock from md_rows and colmap indices (or name indices)."""
    diagram_type = opts.get("mermaid_diagram_type", "flowchart")
    direction = opts.get("mermaid_direction","TD")
    keep_table = bool(opts.get("mermaid_keep_source_table", True))
    node_id_policy = opts.get("mermaid_node_id_policy", "auto")
    group_behavior = opts.get("mermaid_group_column_behavior", "subgraph")

    # Prepare sets
    nodes = {}
    edges = set()
    groups = {}
    header = md_rows[0]
    data_rows = md_rows[1:] if len(md_rows)>1 else []

    def node_id(name):
        if node_id_policy == "explicit":
            # Future extension: use explicit ID column if available
            # For now, fall back to auto
            pass
        # auto policy: sanitize display name
        nid = _v14_sanitize_node_id(str(name))
        i = 2
        base = nid
        while nid in nodes.values():
            nid = f"{base}_{i}"; i += 1
        return nid

    def get_val(row, idx):
        try:
            if idx is None: return ""
            return row[idx] if idx < len(row) else ""
        except Exception:
            return ""

    # Build nodes/edges
    for row in data_rows:
        f = get_val(row, colmap.get("from")) if isinstance(colmap.get("from"), int) else get_val(row, colmap.get("from"))
        t = get_val(row, colmap.get("to"))
        if not str(f).strip() or not str(t).strip():
            continue
        fid = nodes.get(f) or node_id(f); nodes[f]=fid
        tid = nodes.get(t) or node_id(t); nodes[t]=tid
        label = get_val(row, colmap.get("label"))
        key = (fid, tid, str(label or "").strip())
        if opts.get("mermaid_dedupe_edges", True):
            if key in edges:
                continue
        edges.add(key)
        g = get_val(row, colmap.get("group"))
        if g and group_behavior == "subgraph":
            groups.setdefault(g, set()).update([fid, tid])
        # group_behavior="ignore" の場合は groups に追加しない

    # Emit code
    lines = []
    lines.append("```mermaid")

    # diagram_type handling (currently only flowchart is supported)
    if diagram_type == "flowchart":
        lines.append(f"flowchart {direction}")
    elif diagram_type in ("sequence", "state"):
        # Future extension: sequenceDiagram / stateDiagram support
        # For now, fall back to flowchart
        warn(f"mermaid_diagram_type='{diagram_type}' is not yet implemented, using flowchart")
        lines.append(f"flowchart {direction}")
    else:
        lines.append(f"flowchart {direction}")

    # subgraphs (only if group_behavior="subgraph")
    if group_behavior == "subgraph":
        for gname in sorted(groups.keys()):
            lines.append(f"  subgraph {gname}")
            for nid in sorted(groups[gname]):
                # find display name
                disp = None
                for k,v in nodes.items():
                    if v==nid: disp = k; break
                lines.append(f'    {nid}["{str(disp)}"]')
            lines.append("  end")

    # nodes not in any group
    grouped = set().union(*groups.values()) if groups else set()
    for k, nid in sorted(nodes.items(), key=lambda kv: str(kv[0])):
        if nid in grouped:
            continue
        lines.append(f'  {nid}["{str(k)}"]')

    # edges
    for (a,b,lab) in sorted(edges):
        if lab:
            lines.append(f'  {a} -->|{lab}| {b}')
        else:
            lines.append(f'  {a} --> {b}')
    lines.append("```")
    return "\n".join(lines)

def dispatch_table_output(ws, tbl, md_rows, opts, merged_lookup, xlsx_path=None):
    """Unified dispatcher per v1.4: code -> mermaid -> text -> nested -> table

    Args:
        ws: Worksheet object
        tbl: Table dict
        md_rows: Markdown rows
        opts: Options dictionary
        merged_lookup: Merged cell lookup
        xlsx_path: Path to Excel file (not used, kept for compatibility)
    """
    skip_on_fallback = opts.get("dispatch_skip_code_and_mermaid_on_fallback", True)
    code_failed = False

    # 1) Code (最優先)
    try:
        if is_code_block(md_rows):
            # Extract code lines using same logic as format_table_as_text_or_nested
            code_lines = []
            is_code = False
            for row in md_rows:
                if not row:
                    if is_code:
                        code_lines.append("")
                    continue
                row_text = ""
                for val in row:
                    if val and str(val).strip():
                        row_text = str(val).strip()
                        break
                if row_text and is_source_code(row_text):
                    code_lines.append(row_text)
                    is_code = True
                elif is_code and row_text:
                    code_lines.append(row_text)
                elif is_code and not row_text:
                    code_lines.append("")

            if is_code and code_lines:
                # Detect language and format code block
                language = detect_code_language(code_lines)
                code_block = "```" + (language if language else "") + "\n"
                code_block += "\n".join(code_lines)
                code_block += "\n```"
                return "code", code_block
    except Exception as e:
        warn(f"code detection failed: {e}")
        code_failed = True
        if skip_on_fallback:
            # Skip Mermaid detection and go directly to priority 3
            pass
        else:
            # Continue to Mermaid detection
            pass

    # 2) Mermaid (skip if code failed and skip_on_fallback is True)
    if not (code_failed and skip_on_fallback):
        try:
            if opts.get("mermaid_enabled", False):
                detect_mode = opts.get("mermaid_detect_mode", "shapes")

                # 2a) Table-based modes (column_headers/heuristic)
                # Note: shapes mode is handled at sheet level, not here
                if detect_mode in ("column_headers", "heuristic"):
                    ok, colmap = is_flow_table(md_rows, opts)
                    if ok:
                        mer = build_mermaid(md_rows, opts, colmap)
                        if opts.get("mermaid_keep_source_table", True):
                            # header_detection handling: opts["header_detection"] is boolean (True/False)
                            # per §7.0: if header_detection=none, explicitly pass False
                            hdr = opts.get("header_detection", True)
                            # Use existing table generator with explicit header_detection
                            table_md = make_markdown_table(md_rows, header_detection=hdr, align_detect=opts.get("align_detection", True), align_threshold=opts.get("numbers_right_threshold", 0.8))
                            return "mermaid", mer + "\n" + table_md
                        else:
                            return "mermaid", mer
        except Exception as e:
            warn(f"Mermaid generation failed: {e}")
            # fallthrough to (3) - フォールバックは優先度3（単一行）から開始

    # 3-5) delegate to existing logic (フォールバック時は優先度3から)
    ftype, out = format_table_as_text_or_nested(ws, tbl, md_rows, opts, merged_lookup)
    if ftype == "table":
        hdr = opts.get("header_detection", True)
        out = make_markdown_table(md_rows, header_detection=hdr, align_detect=opts.get("align_detection", True), align_threshold=opts.get("numbers_right_threshold", 0.8))
    elif ftype in ("text","nested","code","empty"):
        pass
    else:
        # fallback to normal table
        hdr = opts.get("header_detection", True)
        out = make_markdown_table(md_rows, header_detection=hdr, align_detect=opts.get("align_detection", True), align_threshold=opts.get("numbers_right_threshold", 0.8))
        ftype = "table"
    return ftype, out
# ===== end v1.4 additions =====
def warn(msg: str):
    print(f"[WARN] {msg}", file=sys.stderr)

def info(msg: str):
    print(f"[INFO] {msg}", file=sys.stderr)

def remove_control_chars(text: str) -> str:
    out = []
    for ch in text:
        code = ord(ch)
        if any(lo <= code <= hi for lo, hi in CONTROL_REMOVE_RANGES):
            continue
        out.append(ch)
    return "".join(out)

def is_whitespace_only(s: str) -> bool:
    return all((c in WHITESPACE_CHARS) for c in s) and len(s) > 0

def md_escape(s: str, level: str = "safe") -> str:
    if not s:
        return s
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")
    s = s.replace("|", r"\|")
    if level == "safe":
        s = MD_ESCAPE_RE.sub(lambda m: "\\" + m.group(0), s)
    elif level == "aggressive":
        s = re.sub(r".", lambda m: "\\" + m.group(0), s)
    return s

def load_workbook_safe(path, read_only=False):
    try:
        from openpyxl import load_workbook
        # read_only=False で fill 等のスタイル参照を確実化
        wb = load_workbook(filename=path, read_only=read_only, data_only=True)
        return wb
    except Exception as e:
        print(f"[ERROR] Failed to open workbook: {e}", file=sys.stderr)
        sys.exit(2)

def a1_from_rc(r: int, c: int) -> str:
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(c)}{r}"

def parse_dimension(ws):
    dim = ws.calculate_dimension()
    from openpyxl.utils import range_boundaries
    try:
        min_col, min_row, max_col, max_row = range_boundaries(dim)
        return (min_row, min_col, max_row, max_col)
    except Exception:
        return (1, 1, ws.max_row, ws.max_column)

def get_print_areas(ws, mode: str):
    """Get print areas from worksheet, with validation per spec §③.

    Per spec §③: validates print areas.
    Per spec §9: outputs INFO log when print area is obtained or fallback range is used.

    Args:
        ws: Worksheet object
        mode: no_print_area_mode setting

    Returns:
        List of print area tuples (min_row, min_col, max_row, max_col) (1-based)
    """
    areas = []
    sheet_max_row = ws.max_row
    sheet_max_col = ws.max_column

    try:
        pa = ws.print_area
        if pa:
            # Per spec §③: Print_Area can be a string (single range) or iterable (multiple ranges)
            if isinstance(pa, str):
                # Single range as string (e.g., "'Sheet1'!$A$1:$Z$100" or "$A$1:$Z$100")
                rngs = [pa]
            else:
                # Multiple ranges as iterable (list/tuple)
                try:
                    rngs = list(pa)
                except (TypeError, AttributeError):
                    # If it's not iterable, treat as single range
                    rngs = [pa]
            for r in rngs:
                try:
                    from openpyxl.utils import range_boundaries
                    # Per spec §③: remove sheet name from range string if present
                    # range_boundaries expects format like "$A$1:$Z$100", not "'Sheet1'!$A$1:$Z$100"
                    range_str = str(r)
                    if '!' in range_str:
                        # Extract range part after '!' (e.g., "'Sheet1'!$A$1:$Z$100" -> "$A$1:$Z$100")
                        range_str = range_str.split('!', 1)[1]
                    (min_col, min_row, max_col, max_row) = range_boundaries(range_str)
                    # Per spec §③: validate print area
                    # Check if it's a valid rectangle
                    if min_row > max_row or min_col > max_col:
                        warn(f"Invalid print area range (min > max): {r} in sheet '{ws.title}'")
                        continue
                    # Check if row/column numbers are positive integers
                    if min_row < 1 or min_col < 1 or max_row < 1 or max_col < 1:
                        warn(f"Invalid print area range (negative or zero): {r} in sheet '{ws.title}'")
                        continue
                    # Per spec §③: limit to sheet's maximum range if exceeded
                    if max_row > sheet_max_row:
                        warn(f"Print area max_row ({max_row}) exceeds sheet max_row ({sheet_max_row}), limiting to {sheet_max_row} in sheet '{ws.title}'")
                        max_row = sheet_max_row
                    if max_col > sheet_max_col:
                        warn(f"Print area max_col ({max_col}) exceeds sheet max_col ({sheet_max_col}), limiting to {sheet_max_col} in sheet '{ws.title}'")
                        max_col = sheet_max_col
                    if min_row > sheet_max_row or min_col > sheet_max_col:
                        warn(f"Print area min_row/min_col exceeds sheet maximum, skipping range {r} in sheet '{ws.title}'")
                        continue
                    # Adjust min_row/min_col if they exceed sheet maximum (shouldn't happen, but safety check)
                    min_row = min(min_row, sheet_max_row)
                    min_col = min(min_col, sheet_max_col)
                    areas.append((min_row, min_col, max_row, max_col))
                except Exception as e:
                    warn(f"Failed to parse print area range '{r}' in sheet '{ws.title}': {e}")
                    continue
    except Exception as e:
        warn(f"Failed to get print area from sheet '{ws.title}': {e}")

    if areas:
        # Per spec §9: INFO（印刷領域関連）
        area_str = ", ".join([f"({r0},{c0},{r1},{c1})" for r0, c0, r1, c1 in areas])
        info(f"Print area for sheet '{ws.title}': [{area_str}]")
        return areas

    # Per spec §③: no print area set, use no_print_area_mode
    if mode == "skip_sheet":
        warn(f"No print area set for sheet '{ws.title}', skipping sheet (no_print_area_mode=skip_sheet)")
        return []
    if mode == "entire_sheet_range":
        fallback_area = (1, 1, sheet_max_row, sheet_max_col)
        # Per spec §9: INFO（印刷領域関連）
        info(f"No print area set for sheet '{ws.title}', using entire sheet range: {fallback_area}")
        return [fallback_area]
    # used_range mode (default)
    mr, mc, MR, MC = parse_dimension(ws)
    fallback_area = (mr, mc, MR, MC)
    # Per spec §9: INFO（印刷領域関連）
    info(f"No print area set for sheet '{ws.title}', using used range: {fallback_area}")
    return [fallback_area]

def no_fill(cell, readonly_fill_policy: str = "assume_no_fill") -> bool:
    """Check if cell has no fill or white fill (white fill is treated as no fill).

    Returns True if:
    - Fill is not set (None or patternType is None/none)
    - Fill is white (#FFFFFF or indexed=64) - treated as no fill per user requirement
    Returns False if:
    - Fill has any color other than white
    """
    fill = getattr(cell, "fill", None)
    if fill is None:
        # Style may be unavailable in read_only. Use policy.
        return True if readonly_fill_policy == "assume_no_fill" else False
    patternType = getattr(fill, "patternType", None)
    if patternType in (None, "none"):
        return True
    if patternType == "solid":
        try:
            fg = fill.fgColor
            bg = fill.bgColor

            # Check foreground color
            fg_rgb = None
            try:
                fg_rgb = getattr(fg, "rgb", None)
                # Check if fg_rgb is actually a valid RGB value (not an error message)
                if fg_rgb and isinstance(fg_rgb, str) and len(fg_rgb) >= 6:
                    fg_rgb_str = fg_rgb[-6:].upper()
                    if fg_rgb_str == "FFFFFF":
                        # White foreground - check if background is also white or no fill
                        bg_rgb = getattr(bg, "rgb", None)
                        if bg_rgb and isinstance(bg_rgb, str) and len(bg_rgb) >= 6:
                            try:
                                bg_rgb_str = bg_rgb[-6:].upper()
                                if bg_rgb_str == "FFFFFF":
                                    return True
                            except:
                                pass
                        # Check indexed background
                        try:
                            bg_type = getattr(bg, 'type', None)
                            if bg_type == 'indexed' and hasattr(bg, 'indexed'):
                                bg_indexed = bg.indexed
                                if bg_indexed == 64:
                                    return True
                        except:
                            pass
                        # If foreground is white but background is not, it's still considered white
                        return True
                    else:
                        # Any other foreground color means the cell has fill (not empty)
                        return False
            except:
                pass

            # Check indexed foreground colors
            # Only check fg.indexed if fg.type is 'indexed'
            fg_indexed = None
            try:
                fg_type = getattr(fg, 'type', None)
                if fg_type == 'indexed' and hasattr(fg, 'indexed'):
                    fg_indexed = fg.indexed
            except:
                pass

            if fg_indexed == 64:
                # White foreground - check background
                try:
                    bg_type = getattr(bg, 'type', None)
                    if bg_type == 'indexed' and hasattr(bg, 'indexed'):
                        bg_indexed = bg.indexed
                        if bg_indexed == 64 or bg_indexed is None:
                            return True
                    else:
                        return True
                except:
                    return True
            # Note: Don't return False here if fg_indexed is not None and not 64
            # We need to check bgColor first, as bgColor being white should override fgColor

            # Check background color (if foreground is theme, not set, or white)
            # This is important for cases where fgColor is theme type but bgColor is white
            bg_rgb = getattr(bg, "rgb", None)
            if bg_rgb:
                try:
                    bg_rgb_str = str(bg_rgb)[-6:].upper() if len(str(bg_rgb)) >= 6 else str(bg_rgb).upper()
                    if bg_rgb_str == "FFFFFF":
                        return True
                except:
                    pass

            # Check indexed background colors (this is the key check for white background)
            # Only check bg.indexed if bg.type is 'indexed'
            try:
                bg_type = getattr(bg, 'type', None)
                if bg_type == 'indexed' and hasattr(bg, 'indexed'):
                    bg_indexed = bg.indexed
                    if bg_indexed == 64:
                        # White background - treat as no fill
                        return True
                    elif bg_indexed is not None:
                        # Any other indexed background color means the cell has fill
                        return False
            except:
                pass

        except Exception:
            return False
    return False

def excel_is_date(cell) -> bool:
    try:
        return cell.is_date
    except Exception:
        return False

def format_value(cell, opts) -> str:
    """Display value with basic formatting per §5.6"""
    v = cell.value
    if v is None:
        return ""
    # Dates
    if opts.get("detect_dates", True) and excel_is_date(cell):
        import datetime as _dt
        if isinstance(v, (_dt.datetime, _dt.date)):
            fmt = opts.get("date_format_override") or opts.get("date_default_format", "YYYY-MM-DD")
            # Simple tokens: map to strftime
            fmt_map = {
                "YYYY": "%Y", "MM": "%m", "DD": "%d",
                "HH": "%H", "mm": "%M", "ss": "%S",
            }
            pyfmt = fmt
            for k, f in fmt_map.items():
                pyfmt = pyfmt.replace(k, f)
            return v.strftime(pyfmt)
    # Strings
    if isinstance(v, str):
        return v
    # Numbers & others
    s = str(v)
    return s

def cell_display_value(cell, opts) -> str:
    text = format_value(cell, opts)
    if isinstance(text, str):
        text = unicodedata.normalize("NFC", text)
        text = remove_control_chars(text)
        if opts["strip_whitespace"]:
            text = text.strip()
    return "" if text is None else str(text)

def cell_is_empty(cell, opts) -> bool:
    text = cell_display_value(cell, opts)
    if text == "" or is_whitespace_only(text):
        if no_fill(cell, opts.get("readonly_fill_policy","assume_no_fill")):
            return True
    return False

def collect_hidden(ws):
    hidden_rows = set(i for i, d in ws.row_dimensions.items() if getattr(d, "hidden", False))
    hidden_cols_idx = set()
    for key, d in ws.column_dimensions.items():
        if getattr(d, "hidden", False):
            from openpyxl.utils import column_index_from_string
            try:
                hidden_cols_idx.add(column_index_from_string(key))
            except Exception:
                pass
    return hidden_rows, hidden_cols_idx

def build_nonempty_grid(ws, area, hidden_policy="ignore", opts=None) -> Tuple[List[List[int]], int, int, int, int]:
    r0, c0, r1, c1 = area
    rows = r1 - r0 + 1
    cols = c1 - c0 + 1
    grid = [[0]*cols for _ in range(rows)]
    merged_blocks = list(ws.merged_cells.ranges) if getattr(ws, "merged_cells", None) else []
    merged_coords = []
    for rng in merged_blocks:
        min_row, min_col, max_row, max_col = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        # Per spec §6.4: if merged cell doesn't intersect with print area, skip it
        if max_row < r0 or min_row > r1 or max_col < c0 or min_col > c1:
            continue
        # Per spec §6.4: if top-left cell of merged cell is outside print area, exclude it
        if min_row < r0 or min_col < c0:
            continue  # Top-left cell is outside print area
        # Only include the part of merged cell that is within print area
        merged_coords.append((max(r0, min_row), max(c0, min_col), min(r1, max_row), min(c1, max_col)))

    hidden_rows, hidden_cols_idx = collect_hidden(ws)

    for rr in range(rows):
        for cc in range(cols):
            R = r0 + rr
            C = c0 + cc
            if hidden_policy == "exclude" and (R in hidden_rows or C in hidden_cols_idx):
                continue
            cell = ws.cell(row=R, column=C)
            empty = cell_is_empty(cell, opts)
            grid[rr][cc] = 0 if empty else 1

    # merged influence
    # Per spec §6.4: only mark cells within print area
    for mr0, mc0, mr1, mc1 in merged_coords:
        found_nonempty = False
        # Check if any cell in merged range (within print area) is non-empty
        for R in range(mr0, mr1+1):
            for C in range(mc0, mc1+1):
                # Per spec §6.4: only check cells within print area
                if R < r0 or R > r1 or C < c0 or C > c1:
                    continue  # Cell is outside print area
                cell = ws.cell(row=R, column=C)
                if not cell_is_empty(cell, opts):
                    found_nonempty = True
                    break
            if found_nonempty:
                break
        if found_nonempty:
            # Per spec §6.4: only mark cells within print area
            for R in range(mr0, mr1+1):
                for C in range(mc0, mc1+1):
                    # Per spec §6.4: only mark cells within print area
                    if R < r0 or R > r1 or C < c0 or C > c1:
                        continue  # Cell is outside print area
                    # Convert to grid coordinates
                    grid_r = R - r0
                    grid_c = C - c0
                    if 0 <= grid_r < rows and 0 <= grid_c < cols:
                        grid[grid_r][grid_c] = 1
    return grid, r0, c0, r1, c1

def enumerate_histogram_rectangles(grid: List[List[int]]) -> List[Tuple[int,int,int,int]]:
    if not grid or not grid[0]:
        return []
    R = len(grid)
    C = len(grid[0])
    H = [0]*C
    rects = []
    for r in range(R):
        for c in range(C):
            H[c] = H[c] + 1 if grid[r][c] == 1 else 0
        stack = []  # (start_index, height)
        for c in range(C + 1):
            h = H[c] if c < C else 0
            last = c
            while stack and stack[-1][1] > h:
                i, hh = stack.pop()
                width = c - i
                top = r - hh + 1
                left = i
                bottom = r
                right = c - 1
                if hh > 0 and width > 0:
                    rects.append((top, left, bottom, right))
                last = i
            if not stack or stack[-1][1] < h:
                stack.append((last, h))
    return rects

def carve_rectangles(grid: List[List[int]]) -> List[Tuple[int,int,int,int]]:
    def any_ones(g):
        for row in g:
            if any(v == 1 for v in row):
                return True
        return False

    out = []
    g = [row[:] for row in grid]
    while any_ones(g):
        candidates = enumerate_histogram_rectangles(g)
        if not candidates:
            break
        candidates.sort(key=lambda r: ((r[2]-r[0]+1)*(r[3]-r[1]+1)), reverse=True)
        top,left,bottom,right = candidates[0]
        out.append((top,left,bottom,right))
        for r in range(top, bottom+1):
            for c in range(left, right+1):
                g[r][c] = 0
    out.sort(key=lambda r: (r[0], r[1], (r[2]-r[0]+1)*(r[3]-r[1]+1)))
    return out

def bfs_components(grid: List[List[int]]) -> List[Set[Tuple[int,int]]]:
    R = len(grid); C = len(grid[0]) if R else 0
    seen = [[False]*C for _ in range(R)]
    comps = []
    from collections import deque
    for r in range(R):
        for c in range(C):
            if grid[r][c] == 1 and not seen[r][c]:
                q = deque()
                q.append((r,c)); seen[r][c] = True
                comp = set()
                comp.add((r,c))
                while q:
                    rr,cc = q.popleft()
                    for dr,dc in ((1,0),(-1,0),(0,1),(0,-1)):
                        nr,nc = rr+dr, cc+dc
                        if 0<=nr<R and 0<=nc<C and grid[nr][nc]==1 and not seen[nr][nc]:
                            seen[nr][nc]=True
                            q.append((nr,nc))
                            comp.add((nr,nc))
                comps.append(comp)
    return comps

def rectangles_for_component(comp: Set[Tuple[int,int]], grid_shape: Tuple[int,int]) -> List[Tuple[int,int,int,int]]:
    """Build a mask for the component and carve rectangles inside it."""
    R,C = grid_shape
    # Build component-only grid
    g = [[0]*C for _ in range(R)]
    for (r,c) in comp:
        g[r][c] = 1
    return carve_rectangles(g)

def union_rects(rects: List[Tuple[int,int,int,int]]) -> List[Tuple[int,int,int,int]]:
    """Line-sweep union of rectangles (row-based), returns disjoint rectangles covering union."""
    if not rects:
        return []
    events = {}
    min_r = min(r0 for r0,_,r1,_ in rects)
    max_r = max(r1 for r0,_,r1,_ in rects)
    for r0,c0,r1,c1 in rects:
        events.setdefault(r0, []).append(("add",(c0,c1)))
        events.setdefault(r1+1, []).append(("rem",(c0,c1)))
    active: List[Tuple[int,int]] = []
    def merge_intervals(iv):
        if not iv: return []
        iv.sort()
        out = [list(iv[0])]
        for s,e in iv[1:]:
            if s <= out[-1][1]+0:
                out[-1][1] = max(out[-1][1], e)
            else:
                out.append([s,e])
        return [(a,b) for a,b in out]

    out_rects = []
    prev_row = None
    prev_spans = []
    # sweep rows
    for row in range(min_r, max_r+2):  # include sentinel
        for ev in events.get(row, []):
            kind,(s,e) = ev
            if kind=="add":
                active.append((s,e))
            else:
                try:
                    active.remove((s,e))
                except ValueError:
                    pass
        spans = merge_intervals(active)
        if prev_row is None:
            prev_row = row
            prev_spans = [(s,e) for s,e in spans]
            continue
        # If spans changed, close previous run
        if spans != prev_spans:
            # Close rectangles from prev_row to row-1
            for s,e in prev_spans:
                out_rects.append((prev_row, s, row-1, e))
            prev_row = row
            prev_spans = [(s,e) for s,e in spans]
    return out_rects

def grid_to_tables(ws, area, hidden_policy="ignore", opts=None):
    """Return list of logical tables; each table: dict with 'rects' (list of sheet-coord rects) and 'bbox'"""
    if opts is None:
        opts = {}
    grid, r0, c0, r1, c1 = build_nonempty_grid(ws, area, hidden_policy=hidden_policy, opts=opts)

    # Detect empty rows and columns to split tables
    # Note: Check actual cells, not grid (which may be marked by merged cells)
    R = len(grid)
    C = len(grid[0]) if R else 0

    # Find completely empty rows by checking actual cells
    # For row emptiness detection, we check:
    # 1. Cell value (text) - must be empty or whitespace only
    # 2. Cell fill - white fill is treated as no fill (same as blank)
    #    Any other color means the cell is not empty
    # We need to consider merged cells: only check the top-left cell of merged ranges.
    # Per spec §6.4: only merged cells within print area are processed
    merged_lookup = build_merged_lookup(ws, area)

    empty_rows = set()
    for r in range(R):
        row_num = r0 + r
        is_empty = True
        for c in range(C):
            col_num = c0 + c
            # Check if this cell is part of a merged range
            tl = merged_lookup.get((row_num, col_num))
            if tl:
                # Only check the top-left cell of merged ranges
                if (row_num, col_num) != tl:
                    continue  # Skip non-top-left cells in merged ranges
                cell = ws.cell(row=tl[0], column=tl[1])
            else:
                cell = ws.cell(row=row_num, column=col_num)
            # Check cell value and fill
            text = cell_display_value(cell, opts)
            has_text = text and text.strip()
            # If cell has fill (not white), it's not empty
            has_fill = not no_fill(cell, opts.get("readonly_fill_policy", "assume_no_fill"))
            if has_text or has_fill:
                is_empty = False
                break
        if is_empty:
            empty_rows.add(r)

    # Find completely empty columns by checking actual cells
    # For column emptiness detection, we check:
    # 1. Cell value (text) - must be empty or whitespace only
    # 2. Cell fill - white fill is treated as no fill (same as blank)
    #    Any other color means the cell is not empty
    # We need to consider merged cells: only check the top-left cell of merged ranges.
    empty_cols = set()
    for c in range(C):
        col_num = c0 + c
        is_empty = True
        for r in range(R):
            row_num = r0 + r
            # Check if this cell is part of a merged range
            tl = merged_lookup.get((row_num, col_num))
            if tl:
                # Only check the top-left cell of merged ranges
                if (row_num, col_num) != tl:
                    continue  # Skip non-top-left cells in merged ranges
                cell = ws.cell(row=tl[0], column=tl[1])
            else:
                cell = ws.cell(row=row_num, column=col_num)
            # Check cell value and fill
            text = cell_display_value(cell, opts)
            has_text = text and text.strip()
            # If cell has fill (not white), it's not empty
            has_fill = not no_fill(cell, opts.get("readonly_fill_policy", "assume_no_fill"))
            if has_text or has_fill:
                is_empty = False
                break
        if is_empty:
            empty_cols.add(c)

    # Split grid by empty rows/columns: only connect cells horizontally or vertically
    # but not across empty rows/columns
    def is_connected(r1, c1, r2, c2):
        """Check if two adjacent cells can be in the same table (not separated by empty row/col)"""
        # For adjacent cells (dr=±1 or dc=±1), check if there's an empty row/col between them
        # Same row (horizontal neighbors): check if any column between is empty
        if r1 == r2:
            min_c, max_c = min(c1, c2), max(c1, c2)
            # For adjacent cells, max_c - min_c should be 1, so check if min_c or max_c is in empty_cols
            return not (min_c in empty_cols or max_c in empty_cols)
        # Same column (vertical neighbors): check if any row between is empty
        if c1 == c2:
            min_r, max_r = min(r1, r2), max(r1, r2)
            # For adjacent cells, max_r - min_r should be 1, so check if min_r or max_r is in empty_rows
            return not (min_r in empty_rows or max_r in empty_rows)
        # Different row and column: check if they can be connected via a path
        # Two cells in different rows and columns can be in the same table if:
        # 1. They are in adjacent rows (dr=±1) and their columns are not separated by empty columns
        # 2. They are in adjacent columns (dc=±1) and their rows are not separated by empty rows
        dr = abs(r2 - r1)
        dc = abs(c2 - c1)
        if dr == 1 and dc == 1:
            # Diagonal neighbors: check if both horizontal and vertical paths are not blocked
            # Check horizontal path: (r1, c1) -> (r1, c2) -> (r2, c2)
            h_path_ok = not (c1 in empty_cols or c2 in empty_cols) and r1 not in empty_rows
            # Check vertical path: (r1, c1) -> (r2, c1) -> (r2, c2)
            v_path_ok = not (r1 in empty_rows or r2 in empty_rows) and c1 not in empty_cols
            # If either path is open, they can be connected
            return h_path_ok or v_path_ok
        # Not adjacent: not directly connected
        return False

    # Modified BFS that respects empty row/column boundaries
    seen = [[False]*C for _ in range(R)]
    comps = []
    from collections import deque
    for r in range(R):
        for c in range(C):
            if grid[r][c] == 1 and not seen[r][c]:
                q = deque()
                q.append((r,c)); seen[r][c] = True
                comp = set()
                comp.add((r,c))
                while q:
                    rr,cc = q.popleft()
                    # Check neighbors (horizontal/vertical and diagonal, respecting boundaries)
                    for dr,dc in ((1,0),(-1,0),(0,1),(0,-1),(1,1),(1,-1),(-1,1),(-1,-1)):
                        nr,nc = rr+dr, cc+dc
                        if 0<=nr<R and 0<=nc<C and grid[nr][nc]==1 and not seen[nr][nc]:
                            # Only connect if not separated by empty row/column
                            if is_connected(rr, cc, nr, nc):
                                seen[nr][nc]=True
                                q.append((nr,nc))
                                comp.add((nr,nc))
                comps.append(comp)

    tables = []
    # Per spec §⑥: ensure bbox and mask are within print area
    # Note: r0, c0, r1, c1 from build_nonempty_grid are the same as area
    # comp contains relative coordinates (0-based) within the grid, which is already within print area
    for comp in comps:
        rects_local = rectangles_for_component(comp, (len(grid), len(grid[0])))
        # convert to sheet coords (r0+r, c0+c are already within print area since grid is built from area)
        rects_sheet = [(r0+t, c0+l, r0+b, c0+r) for (t,l,b,r) in rects_local]
        # bounding box
        min_r = min(r for r,c in comp)
        max_r = max(r for r,c in comp)
        min_c = min(c for r,c in comp)
        max_c = max(c for r,c in comp)
        # Per spec §⑥: bbox and mask are already within print area since grid is built from area
        # But we ensure it explicitly for safety
        bbox = (r0+min_r, c0+min_c, r0+max_r, c0+max_c)
        # Per spec §⑥: mask contains only cells within print area
        # Since comp is relative to grid (which is within print area), r0+r and c0+c are within print area
        # But we filter explicitly to ensure no cells outside print area
        mask = {(r0+r, c0+c) for (r,c) in comp if r0 <= r0+r <= r1 and c0 <= c0+c <= c1}
        # Per spec §⑥: ensure bbox is within print area (should already be, but ensure explicitly)
        bbox_min_row = max(bbox[0], r0)
        bbox_min_col = max(bbox[1], c0)
        bbox_max_row = min(bbox[2], r1)
        bbox_max_col = min(bbox[3], c1)
        bbox = (bbox_min_row, bbox_min_col, bbox_max_row, bbox_max_col)
        tables.append({"rects": rects_sheet, "bbox": bbox, "mask": mask})
    # sort by (top,left)
    tables.sort(key=lambda t: (t["bbox"][0], t["bbox"][1]))
    return tables

def numeric_like(s: str) -> bool:
    m = NUMERIC_PATTERN.match(s)
    if not m:
        return False
    open_paren = m.group(1)
    # closing paren is the penultimate group in groups()
    groups = m.groups()
    close_paren = groups[-2] if len(groups)>=2 else None
    if open_paren and not close_paren:
        return False
    if close_paren and not open_paren:
        return False
    return True

def normalize_numeric_text(s: str, opts) -> str:
    """Apply percent/currency/thousand settings to a numeric-like string for output control."""
    if not numeric_like(s):
        return s
    raw = s
    # Currency strip (start or surrounded by spaces)
    if opts.get("currency_symbol","keep") == "strip":
        raw = re.sub(r"^\s*[¥$€£₩]\s*", "", raw)
    # Remove grouping if requested (only digits separators)
    if opts.get("numeric_thousand_sep","keep") == "remove":
        raw = raw.replace(",", "").replace("，","")
    # Percent handling
    if opts.get("percent_format","keep") == "numeric":
        has_pct = raw.strip().endswith("%")
        raw = raw.replace("%", "")
        if has_pct and opts.get("percent_divide_100"):
            try:
                # convert to float and divide
                raw_num = float(raw)
                raw = str(raw_num/100.0)
            except Exception:
                pass
    return raw

def hyperlink_info(cell):
    try:
        hl = cell.hyperlink
    except Exception:
        return None
    if not hl:
        return None
    target = getattr(hl, "target", None)
    location = getattr(hl, "location", None)
    display = getattr(hl, "display", None)
    return {"target": target, "location": location, "display": display}

def is_valid_url(target: str) -> bool:
    return bool(re.match(r"^(https?://|mailto:|file://|/|\.{1,2}/)", target or ""))

def build_merged_lookup(ws, area=None):
    """Map each cell (r,c) in a merged range to its top-left (r0,c0).

    Per spec §6.4: only include merged cells within print area.
    If area is provided, only cells within the print area are included.

    Args:
        ws: Worksheet object
        area: Optional print area tuple (r0, c0, r1, c1). If provided, only merged cells
              that intersect with the print area are included, and only cells within
              the print area are added to the lookup.
    """
    lookup = {}
    for rng in getattr(ws, "merged_cells", []):
        try:
            r0,c0,r1,c1 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        except Exception:
            continue

        # Per spec §6.4: if top-left cell is outside print area, exclude the merged cell
        if area is not None:
            area_r0, area_c0, area_r1, area_c1 = area
            # Check if merged cell intersects with print area
            if r1 < area_r0 or r0 > area_r1 or c1 < area_c0 or c0 > area_c1:
                continue  # Merged cell is completely outside print area
            # Per spec §6.4: if top-left cell is outside print area, exclude it
            if r0 < area_r0 or c0 < area_c0:
                continue  # Top-left cell is outside print area

        for R in range(r0, r1+1):
            for C in range(c0, c1+1):
                # Per spec §6.4: only include cells within print area
                if area is not None:
                    area_r0, area_c0, area_r1, area_c1 = area
                    if R < area_r0 or R > area_r1 or C < area_c0 or C > area_c1:
                        continue  # This cell is outside print area
                lookup[(R,C)] = (r0,c0)
    return lookup

def detect_table_title(ws, table, merged_lookup, opts, print_area=None):
    """Detect if table has a title from a large merged cell at the top-left.

    Per spec §7.1: only check cells within print area.

    Args:
        ws: Worksheet object
        table: Table dict with 'bbox' and 'mask'
        merged_lookup: Merged cell lookup (only includes cells within print area)
        opts: Options dictionary
        print_area: Optional print area tuple (r0, c0, r1, c1) to ensure cells outside are excluded
    """
    min_row, min_col, max_row, max_col = table["bbox"]
    mask: Set[Tuple[int,int]] = table["mask"]

    # Per spec §7.1: only check cells within print area
    if print_area is not None:
        area_r0, area_c0, area_r1, area_c1 = print_area
        min_row = max(min_row, area_r0)
        min_col = max(min_col, area_c0)
        max_row = min(max_row, area_r1)
        max_col = min(max_col, area_c1)

    # Check first 3 rows for a large merged cell that spans 3+ columns
    for r in range(min_row, min(min_row + 3, max_row + 1)):
        for c in range(min_col, min(min_col + 10, max_col + 1)):  # Check first 10 columns, but within print area
            if (r, c) not in mask:
                continue
            # Per spec §7.1: only check cells within print area
            if print_area is not None:
                area_r0, area_c0, area_r1, area_c1 = print_area
                if r < area_r0 or r > area_r1 or c < area_c0 or c > area_c1:
                    continue
            tl = merged_lookup.get((r, c))
            if tl and tl == (r, c):  # This is a top-left of a merged cell
                # Get the merged range
                merged_cells = getattr(ws, "merged_cells", None)
                if merged_cells:
                    for rng in merged_cells.ranges:
                        if rng.min_row == r and rng.min_col == c:
                            # Per spec §7.1: ensure top-left cell is within print area
                            if print_area is not None:
                                area_r0, area_c0, area_r1, area_c1 = print_area
                                if r < area_r0 or c < area_c0:
                                    continue  # Top-left cell is outside print area
                                # Adjust span_cols to be within print area
                                span_cols = min(rng.max_col, area_c1) - c + 1
                            else:
                                span_cols = rng.max_col - rng.min_col + 1
                            # Check if it spans 3+ columns and is in the first few rows
                            if span_cols >= 3 and r <= min_row + 2:
                                cell = ws.cell(row=r, column=c)
                                text = cell_display_value(cell, opts)
                                if text and text.strip():
                                    # This could be a title - return it and the column range to exclude
                                    # Per spec §7.1: only exclude columns within print area
                                    if print_area is not None:
                                        area_r0, area_c0, area_r1, area_c1 = print_area
                                        exclude_cols = set(range(c, min(rng.max_col, area_c1) + 1))
                                    else:
                                        exclude_cols = set(range(c, rng.max_col + 1))
                                    return text.strip(), exclude_cols
    return None, set()

def extract_table(ws, table, opts, footnotes, footnote_index_start, merged_lookup, print_area=None):
    """Extract Markdown rows for a logical table (possibly multiple rects).

    Per spec §⑦: only process cells within print area.

    Args:
        ws: Worksheet object
        table: Table dict with 'bbox' and 'mask'
        opts: Options dictionary
        footnotes: Footnotes list
        footnote_index_start: Starting footnote index
        merged_lookup: Merged cell lookup (only includes cells within print area)
        print_area: Optional print area tuple (r0, c0, r1, c1) to ensure cells outside are excluded
    """
    min_row, min_col, max_row, max_col = table["bbox"]
    mask: Set[Tuple[int,int]] = table["mask"]

    # Per spec §⑦: ensure all cells are within print area
    if print_area is not None:
        area_r0, area_c0, area_r1, area_c1 = print_area
        # Filter mask to only include cells within print area
        mask = {(r, c) for (r, c) in mask if area_r0 <= r <= area_r1 and area_c0 <= c <= area_c1}
        # Adjust bbox to be within print area
        min_row = max(min_row, area_r0)
        min_col = max(min_col, area_c0)
        max_row = min(max_row, area_r1)
        max_col = min(max_col, area_c1)

    # Detect table title from large merged cell
    # Create updated table dict with filtered mask and adjusted bbox
    updated_table = {"bbox": (min_row, min_col, max_row, max_col), "mask": mask}
    table_title, title_cols = detect_table_title(ws, updated_table, merged_lookup, opts, print_area)

    # Find columns that actually have data (non-empty cells in mask), excluding title columns
    # First, collect all columns that are in mask and not in title_cols
    candidate_cols = set()
    for R in range(min_row, max_row+1):
        for C in range(min_col, max_col+1):
            if (R, C) in mask and C not in title_cols:
                candidate_cols.add(C)

    # Then, check which columns actually have non-empty cell values
    # Note: We need to apply the same processing as in extract_table to get the final cell values
    # This includes: merged cell handling, numeric formatting, hyperlink processing, and markdown escaping
    used_cols = set()
    for C in candidate_cols:
        has_data = False
        for R in range(min_row, max_row+1):
            if (R, C) not in mask:
                continue
            cell = ws.cell(row=R, column=C)
            # Check merged cell handling (same as in extract_table)
            tl = merged_lookup.get((R, C))
            if tl:
                if (R, C) == tl:
                    # Top-left cell: check its value
                    cell = ws.cell(row=tl[0], column=tl[1])
                    text = cell_display_value(cell, opts)
                else:
                    # Other cells in merged range: empty if top_left_only
                    if opts.get("merge_policy") == "top_left_only":
                        text = ""
                    else:
                        cell = ws.cell(row=tl[0], column=tl[1])
                        text = cell_display_value(cell, opts)
            else:
                text = cell_display_value(cell, opts)

            # Apply the same processing as in extract_table
            # numeric formatting overrides
            text = normalize_numeric_text(text, opts)

            # hyperlink processing (but we only need to check if there's data, not format it)
            hl = hyperlink_info(cell)
            if hl:
                disp = text if text else (hl.get("display") or "")
                if disp and disp.strip():
                    text = disp  # Use display text for checking

            # Check if text is non-empty after processing
            if text and text.strip():
                has_data = True
                break
        if has_data:
            used_cols.add(C)

    # Sort columns
    used_cols = sorted(used_cols)
    if not used_cols:
        return [], [], False, table_title

    md_rows = []
    note_refs = []
    max_cells = opts["max_cells_per_table"]
    count_cells = 0

    for R in range(min_row, max_row+1):
        # Per spec §⑦: ensure row is within print area
        if print_area is not None:
            area_r0, area_c0, area_r1, area_c1 = print_area
            if R < area_r0 or R > area_r1:
                continue  # Row is outside print area
        # Check if this row is completely empty (all cells in used_cols are empty)
        row_is_empty = True
        for C in used_cols:
            if (R, C) not in mask:
                continue
            # Per spec §⑦: ensure cell is within print area
            if print_area is not None:
                area_r0, area_c0, area_r1, area_c1 = print_area
                if R < area_r0 or R > area_r1 or C < area_c0 or C > area_c1:
                    continue  # Cell is outside print area
            cell = ws.cell(row=R, column=C)
            tl = merged_lookup.get((R, C))
            if tl:
                if (R, C) == tl:
                    cell = ws.cell(row=tl[0], column=tl[1])
                    text = cell_display_value(cell, opts)
                    if text and text.strip():
                        row_is_empty = False
                        break
                else:
                    if opts.get("merge_policy") != "top_left_only":
                        cell = ws.cell(row=tl[0], column=tl[1])
                        text = cell_display_value(cell, opts)
                        if text and text.strip():
                            row_is_empty = False
                            break
            else:
                text = cell_display_value(cell, opts)
                if text and text.strip():
                    row_is_empty = False
                    break

        # Skip completely empty rows
        if row_is_empty:
            continue

        row_vals = []
        for C in used_cols:
            count_cells += 1
            if count_cells > max_cells:
                return md_rows, note_refs, True
            if (R,C) not in mask:
                row_vals.append("")  # outside component -> empty
                continue
            # Per spec §⑦: ensure cell is within print area before processing
            if print_area is not None:
                area_r0, area_c0, area_r1, area_c1 = print_area
                if R < area_r0 or R > area_r1 or C < area_c0 or C > area_c1:
                    row_vals.append("")  # Cell is outside print area -> empty
                    continue
            cell = ws.cell(row=R, column=C)
            # merged handling
            tl = merged_lookup.get((R,C))
            if tl:
                # Check if this is the top-left cell of the merged range
                if (R, C) == tl:
                    # Top-left cell: use its value
                    cell = ws.cell(row=tl[0], column=tl[1])
                    text = cell_display_value(cell, opts)
                else:
                    # Other cells in merged range: empty if top_left_only, otherwise use top-left value
                    if opts.get("merge_policy") == "top_left_only":
                        text = ""
                    else:
                        # expand/repeat: use top-left value
                        cell = ws.cell(row=tl[0], column=tl[1])
                        text = cell_display_value(cell, opts)
            else:
                text = cell_display_value(cell, opts)
            # numeric formatting overrides
            text = normalize_numeric_text(text, opts)

            hl = hyperlink_info(cell)
            if hl:
                disp = text if text else (hl.get("display") or "")
                if hl.get("target"):
                    link = hl["target"]
                    if not is_valid_url(link):
                        warn(f"Invalid URL detected at {a1_from_rc(R,C)}: {link}")
                    if opts["hyperlink_mode"] in ("inline", "both"):
                        text = f"[{disp}]({link})"
                    elif opts["hyperlink_mode"] == "inline_plain":
                        text = f"{disp} ({link})"
                    if opts["hyperlink_mode"] in ("footnote", "both"):
                        n = footnote_index_start + len(footnotes)
                        note_refs.append((n, link))
                        text = f"{text}[^{n}]"
                elif hl.get("location"):
                    loc = hl["location"]
                    if opts["hyperlink_mode"] in ("inline", "both"):
                        text = f"[{disp}]({loc})"
                    elif opts["hyperlink_mode"] == "inline_plain":
                        text = f"{disp} (→{loc})"
                    elif opts["hyperlink_mode"] in ("footnote", "both"):
                        n = footnote_index_start + len(footnotes)
                        note_refs.append((n, loc))
                        text = f"{disp}[^{n}]"

            text = md_escape(text, opts["markdown_escape_level"])
            row_vals.append(text)
        md_rows.append(row_vals)
    return md_rows, note_refs, False, table_title

def has_border(cell):
    """Check if cell has any border lines.

    For table format detection, we consider a cell to have borders only if it has
    borders on multiple sides (not just one side), as single-side borders are often
    used for visual separation rather than table structure.
    """
    try:
        border = cell.border
        borders = [
            border.left.style is not None and border.left.style != 'none',
            border.right.style is not None and border.right.style != 'none',
            border.top.style is not None and border.top.style != 'none',
            border.bottom.style is not None and border.bottom.style != 'none'
        ]
        # Consider it a border only if 2 or more sides have borders
        return sum(borders) >= 2
    except:
        return False

def is_source_code(text: str) -> bool:
    """Check if text appears to be source code based on §7.2.4 specification."""
    if not text or not text.strip():
        return False

    text_lower = text.lower()
    text_stripped = text.strip()

    # Check for code keywords
    code_keywords = [
        'public', 'private', 'protected', 'class', 'interface', 'import', 'package',
        'static', 'final', 'void', 'return', 'if', 'else', 'for', 'while', 'switch',
        'case', 'try', 'catch', 'throw', 'throws', 'extends', 'implements',
        'def', 'function', 'var', 'let', 'const', 'async', 'await',
        'namespace', 'using', 'namespace', 'struct', 'enum'
    ]

    # Check for code symbols
    code_symbols = ['{', '}', ';', '//', '/*', '*/']

    # Check for annotations (Java/C# style) - @ followed by identifier
    # Pattern: @ followed by alphanumeric characters (e.g., @Annotation1, @Override)
    has_annotation = False
    if '@' in text:
        # Check if @ is followed by identifier-like pattern
        import re
        annotation_pattern = r'@[a-zA-Z_][a-zA-Z0-9_]*'
        if re.search(annotation_pattern, text):
            has_annotation = True

    # Check for keywords
    has_keyword = any(kw in text_lower for kw in code_keywords)

    # Check for symbols
    has_symbol = any(sym in text for sym in code_symbols)

    # If it starts with @ and looks like annotation, likely code
    if text_stripped.startswith('@') and len(text_stripped) > 1 and text_stripped[1].isalnum():
        return True

    # If it has annotation marker and looks like code
    if has_annotation and (has_keyword or has_symbol):
        return True

    # If it has both keyword and symbol, likely code
    if has_keyword and has_symbol:
        return True

    # If it has multiple symbols, likely code
    symbol_count = sum(1 for sym in code_symbols if sym in text)
    if symbol_count >= 2:
        return True

    # If it has keyword alone (for single-line code snippets)
    if has_keyword and (has_symbol or has_annotation):
        return True

    return False

def detect_code_language(lines: list) -> str:
    """Detect programming language from code lines."""
    if not lines:
        return ""

    # Combine all lines for analysis
    combined = " ".join(lines).lower()

    # Java indicators
    if any(kw in combined for kw in ['public class', 'private class', 'import java', '@override', '@annotation']):
        return "java"

    # Python indicators
    if any(kw in combined for kw in ['def ', 'import ', 'from ', 'if __name__', 'class ']) and ':' in combined:
        return "python"

    # JavaScript indicators
    if any(kw in combined for kw in ['function ', 'const ', 'let ', 'var ', '=>', 'export ', 'import ']):
        return "javascript"

    # C# indicators
    if any(kw in combined for kw in ['namespace ', 'using ', 'public class', '[attribute']):
        return "csharp"

    # C/C++ indicators
    if any(kw in combined for kw in ['#include', 'int main', 'printf', 'cout']):
        return "c"

    return ""

def format_table_as_text_or_nested(ws, table, md_rows, opts, merged_lookup):
    """Format table as text/nested format based on §7.2 specification.

    Returns:
    - "text": Single line text format (1 cell in first row, no borders)
    - "nested": Nested format (indented text)
    - "code": Source code format (code block)
    - "table": Normal table format
    - "empty": Empty row (just newline)
    """
    if not md_rows:
        return "empty", ""

    min_row, min_col, max_row, max_col = table["bbox"]
    mask = table["mask"]

    # Get used columns by checking which columns have data
    # First, collect all columns that are in mask
    candidate_cols = set()
    for R in range(min_row, max_row + 1):
        for C in range(min_col, max_col + 1):
            if (R, C) in mask:
                candidate_cols.add(C)

    # Then, check which columns actually have non-empty cell values
    used_cols = set()
    for C in candidate_cols:
        has_data = False
        for R in range(min_row, max_row + 1):
            if (R, C) not in mask:
                continue
            cell = ws.cell(row=R, column=C)
            tl = merged_lookup.get((R, C))
            if tl:
                if (R, C) == tl:
                    cell = ws.cell(row=tl[0], column=tl[1])
                    text = cell_display_value(cell, opts)
                else:
                    if opts.get("merge_policy") == "top_left_only":
                        text = ""
                    else:
                        cell = ws.cell(row=tl[0], column=tl[1])
                        text = cell_display_value(cell, opts)
            else:
                text = cell_display_value(cell, opts)
            text = normalize_numeric_text(text, opts)
            if text and text.strip():
                has_data = True
                break
        if has_data:
            used_cols.add(C)

    used_cols = sorted(used_cols)
    if not used_cols:
        return "empty", ""

    # Check first row: single cell text format (§7.2.1)
    first_row = md_rows[0] if md_rows else []
    non_empty_in_first = [i for i, val in enumerate(first_row) if val and val.strip()]

    if len(non_empty_in_first) == 1:
        # Find the actual column index in used_cols
        first_row_col_idx = non_empty_in_first[0]
        if first_row_col_idx < len(used_cols):
            first_row_actual_col = used_cols[first_row_col_idx]

            # Check if the cell with value has no borders
            cell = ws.cell(row=min_row, column=first_row_actual_col)
            if not has_border(cell):
                # Check if other cells in first row are empty and have no borders
                all_other_empty_no_border = True
                for C in used_cols:
                    if C == first_row_actual_col:
                        continue
                    if (min_row, C) not in mask:
                        continue
                    cell = ws.cell(row=min_row, column=C)
                    text = cell_display_value(cell, opts)
                    text = normalize_numeric_text(text, opts)
                    if text and text.strip():
                        all_other_empty_no_border = False
                        break
                    if has_border(cell):
                        all_other_empty_no_border = False
                        break

                if all_other_empty_no_border:
                    # Single line text format
                    text_value = first_row[first_row_col_idx]
                    return "text", text_value

    # Check for source code format (§7.2.4)
    code_lines = []
    is_code_block = False
    code_row_count = 0

    for row_idx, row in enumerate(md_rows):
        if not row:
            if is_code_block:
                # Empty line in code block - preserve it
                code_lines.append("")
            continue

        # Get the first non-empty cell value
        row_text = ""
        for val in row:
            if val and val.strip():
                row_text = val.strip()
                break

        if row_text and is_source_code(row_text):
            code_lines.append(row_text)
            code_row_count += 1
            is_code_block = True
        elif is_code_block and row_text:
            # If we've started a code block, continue collecting lines
            # even if they don't match code patterns (might be continuation)
            code_lines.append(row_text)
        elif is_code_block and not row_text:
            # Empty line in code block - preserve it
            code_lines.append("")

    # If we detected code, format as code block
    if is_code_block and code_lines:
        language = detect_code_language(code_lines)
        code_block = "```" + (language if language else "") + "\n"
        code_block += "\n".join(code_lines)
        code_block += "\n```"
        return "code", code_block

    # Check for nested format (§7.2.3)
    # If first column is empty and second column has value
    nested_lines = []
    use_nested = True
    for row_idx, row in enumerate(md_rows):
        if not row:
            nested_lines.append("")  # Empty line (§7.2.2)
            continue

        # Count non-empty cells in this row
        non_empty = [i for i, val in enumerate(row) if val and val.strip()]

        if len(non_empty) == 0:
            nested_lines.append("")  # Empty line (§7.2.2)
        elif len(non_empty) == 1:
            if non_empty[0] == 0:
                # First column only - normal text (no indent)
                nested_lines.append(row[0])
            else:
                # First column is empty, second+ column has value - nested format
                first_non_empty_idx = non_empty[0]
                indent = "  " * first_non_empty_idx
                nested_lines.append(indent + row[first_non_empty_idx])
        else:
            # Multiple columns with values - check if it's still nested format
            # If all non-empty cells are in sequence starting from index > 0, it's nested
            if non_empty[0] > 0 and all(non_empty[i] == non_empty[0] + i for i in range(len(non_empty))):
                # Sequential columns starting from index > 0 - use first non-empty as nested
                first_non_empty_idx = non_empty[0]
                indent = "  " * first_non_empty_idx
                nested_lines.append(indent + row[first_non_empty_idx])
            else:
                # Multiple columns with values in different positions - use table format
                use_nested = False
                break

    # If all rows are nested format, return nested
    if use_nested and nested_lines:
        return "nested", "\n".join(nested_lines)

    return "table", None

def choose_header_row_heuristic(md_rows):
    """Pick header row using simple heuristics:
    - First row with >=50% non-empty AND less numeric_ratio than next row.
    - Fallback: first non-empty row.
    """
    if not md_rows:
        return None
    def numeric_ratio(row):
        vals = [x for x in row if (x or "").strip()]
        if not vals: return 0.0
        nums = sum(1 for v in vals if numeric_like(v))
        return nums/len(vals)
    first_nonempty = None
    for i, row in enumerate(md_rows[:3]):  # peek first up to 3 rows
        nonempty = sum(1 for v in row if (v or "").strip())
        if first_nonempty is None and nonempty>0:
            first_nonempty = i
        if nonempty >= max(1, len(row)//2):
            r_this = numeric_ratio(row)
            r_next = numeric_ratio(md_rows[i+1]) if i+1 < len(md_rows) else 1.0
            if r_this < r_next:  # header tends to be less numeric than data
                return i
    return first_nonempty if first_nonempty is not None else None

def detect_right_align(col_vals, threshold=0.8):
    """Detect if column should be right-aligned based on numeric ratio."""
    if not col_vals:
        return False
    non_empty = [v for v in col_vals if (v or "").strip()]
    if not non_empty:
        return False
    numeric_count = sum(1 for v in non_empty if numeric_like(str(v)))
    return (numeric_count / len(non_empty)) >= threshold

def make_markdown_table(md_rows, header_detection=True, align_detect=True, align_threshold=0.8):
    if not md_rows:
        return ""

    # Remove completely empty columns
    if md_rows:
        cols = len(md_rows[0])
        non_empty_cols = []
        for c in range(cols):
            # Check if column has any non-empty cell
            has_data = any((row[c] if c < len(row) else "").strip() for row in md_rows)
            if has_data:
                non_empty_cols.append(c)

        # Filter rows to only include non-empty columns
        if non_empty_cols:
            md_rows = [[row[c] if c < len(row) else "" for c in non_empty_cols] for row in md_rows]
        else:
            return ""  # All columns are empty

    header = None
    data = md_rows
    # TODO: heuristic header detection (future spec)
    if header_detection and any((cell or "").strip() for cell in md_rows[0]):
        header = md_rows[0]
        data = md_rows[1:]
    cols = len(header) if header else len(md_rows[0])
    aligns = []
    if align_detect:
        for c in range(cols):
            col_vals = [row[c] for row in data if c < len(row)]
            right = detect_right_align(col_vals, threshold=align_threshold)
            aligns.append("---:" if right else "---")
    else:
        aligns = ["---"] * cols
    lines = []
    if header:
        lines.append("| " + " | ".join(header) + " |")
        lines.append("| " + " | ".join(aligns) + " |")
    for row in data if header else md_rows:
        if len(row) < cols:
            row = row + [""] * (cols - len(row))
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)

# ===== CSV output functions (v1.5) =====
def sanitize_sheet_name(sheet_name: str) -> str:
    """Sanitize sheet name for use in filenames per spec §3.2.2."""
    # Replace filesystem-unsafe characters with underscore
    unsafe_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    sanitized = sheet_name
    for ch in unsafe_chars:
        sanitized = sanitized.replace(ch, '_')
    return sanitized

# ===== CSV Markdown Output Functions (v1.5 spec §3.2.2) =====

def coords_to_excel_range(min_row, min_col, max_row, max_col):
    """Convert cell coordinates to Excel range string (e.g., A1:D10).

    Args:
        min_row: Minimum row (1-based)
        min_col: Minimum column (1-based)
        max_row: Maximum row (1-based)
        max_col: Maximum column (1-based)

    Returns:
        Excel range string (e.g., "A1:D10")
    """
    from openpyxl.utils import get_column_letter
    start_cell = f"{get_column_letter(min_col)}{min_row}"
    end_cell = f"{get_column_letter(max_col)}{max_row}"
    return f"{start_cell}:{end_cell}"

def format_timestamp():
    """Format current timestamp as YYYY-MM-DD HH:MM:SS.

    Returns:
        Formatted timestamp string
    """
    from datetime import datetime
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def write_csv_markdown(wb, csv_data_dict, excel_file_basename, opts, output_dir):
    """Write CSV markdown file containing sheets' CSV data per spec §3.2.2.

    Args:
        wb: Workbook object
        csv_data_dict: Dictionary of {sheet_name: rows} where rows is list of lists
                       Can contain single sheet (for split mode) or all sheets (for normal mode)
        excel_file_basename: Base name for output file (without extension)
                            For split mode: includes sheet name (e.g., "file_Sheet1")
                            For normal mode: just file name (e.g., "file")
        opts: Options dictionary
        output_dir: Output directory path

    Returns:
        Path to created CSV markdown file, or None if failed
    """
    import csv as csv_module
    from io import StringIO
    import json

    try:
        if not csv_data_dict:
            return None

        # Prepare output file path
        output_dir_path = Path(output_dir)
        output_dir_path.mkdir(parents=True, exist_ok=True)
        csv_md_path = output_dir_path / f"{excel_file_basename}_csv.md"

        # Get current timestamp
        timestamp = format_timestamp()

        # Determine if this is single-sheet or multi-sheet output
        is_single_sheet = len(csv_data_dict) == 1
        sheet_names = list(csv_data_dict.keys())

        # Check if description section should be included (v1.7)
        include_description = opts.get("csv_include_description", True)

        # Write markdown file
        with open(csv_md_path, 'w', encoding='utf-8') as f:
            # Write title section per spec §3.2.2
            if is_single_sheet:
                f.write(f"# {sheet_names[0]}\n\n")
                if include_description:
                    f.write(f"- 仕様バージョン: {VERSION}\n")
                    f.write(f"- 元ファイル: {wb.properties.title or Path(excel_file_basename).name}.xlsx\n\n")
            else:
                f.write(f"# CSV出力: {excel_file_basename}.xlsx\n\n")

            # Overview section (level 2) - only if csv_include_description=true (v1.7)
            if include_description:
                f.write("## 概要\n\n")

                # File Information (level 3) - only for multi-sheet
                if not is_single_sheet:
                    f.write("### ファイル情報\n\n")
                    f.write(f"- 元のExcelファイル名: {excel_file_basename}.xlsx\n")
                    f.write(f"- シート数: {len(csv_data_dict)}\n")
                    f.write(f"- 生成日時: {timestamp}\n\n")

                    # About This File (level 3)
                    f.write("### このファイルについて\n\n")
                    f.write("このCSVマークダウンファイルは、AIがExcelの内容を理解できるよう、各シートの印刷領域をCSV形式で出力したファイルです。")
                    f.write("各シートはマークダウン見出しで区切られ、CSVコードブロックで内容が記載されています。")
                else:
                    # Single-sheet: simplified description
                    f.write("このCSVマークダウンファイルは、AIがExcelの内容を理解できるよう、シートの印刷領域をCSV形式で出力したファイルです。")
                # Add metadata description only if metadata is enabled (v1.7)
                if opts.get("csv_include_metadata", True):
                    f.write("ファイル末尾に検証用メタデータセクションがあり、Excel原本との整合性を確認できます。")
                f.write("\n\n")

                # CSV Generation Method (level 3)
                f.write("### CSV生成方法\n\n")
                sheet_text = "シート" if is_single_sheet else "各シート"
                f.write(f"- **出力対象範囲**: {sheet_text}の印刷領域のみを出力（印刷領域外のセルは含まない）\n")
                f.write(f"- **印刷領域の統合**: 複数の印刷領域がある場合は外接矩形として統合\n")
                f.write(f"- **テーブル分割なし**: Markdown出力と異なり、空行・空列でテーブル分割せず印刷領域全体を1つのCSVとして出力\n")
                f.write(f"- **結合セルの処理**: `{opts.get('merge_policy', 'top_left_only')}` 設定に従って処理\n")
                f.write(f"- **数式の扱い**: `{opts.get('value_mode', 'display')}` 設定に従って表示値・数式・両方のいずれかを出力\n")
                f.write(f"- **値の正規化**: `{opts.get('csv_normalize_values', True)}` 設定に従って正規化処理を適用\n\n")

                # CSV Format Specification (level 3)
                f.write("### CSV形式の仕様\n\n")
                f.write(f"- **区切り文字**: `{opts.get('csv_delimiter', ',')}`\n")
                f.write(f"- **引用符の使用**: `{opts.get('csv_quoting', 'minimal')}` 設定に従い、RFC 4180準拠で処理\n")
                f.write(f"- **セル内改行**: 半角スペースに変換される（1レコード=1行を保証）\n")
                f.write(f"- **セル内特殊文字**: 区切り文字や引用符は引用符でエスケープされる\n")
                f.write(f"- **エンコーディング**: `{opts.get('csv_encoding', 'utf-8')}`\n")
                f.write(f"- **空セルの表現**: 空文字列として出力（空行・空列も含めて出力される）\n")
                # Hyperlink mode description
                hyperlink_mode = opts.get("hyperlink_mode", "inline_plain")
                if hyperlink_mode == "text_only":
                    f.write(f"- **ハイパーリンク**: 表示テキストのみを出力（リンク先URLは含まない）\n\n")
                elif hyperlink_mode == "inline":
                    f.write(f"- **ハイパーリンク**: Markdown形式で出力（例: `[表示テキスト](URL)` または `[表示テキスト](シート名+セル番号)`）\n\n")
                else:  # inline_plain or footnote or both (after fallback)
                    f.write(f"- **ハイパーリンク**: 平文形式で出力（例: `表示テキスト (URL)` または `表示テキスト (→シート名+セル番号)`）\n\n")

                # About Mermaid (level 3) - only when mermaid is enabled with shapes mode (v1.7)
                mermaid_enabled = opts.get("mermaid_enabled", False)
                mermaid_detect_mode = opts.get("mermaid_detect_mode", "shapes")
                if mermaid_enabled and mermaid_detect_mode == "shapes":
                    f.write("### Mermaidフローチャートについて\n\n")
                    f.write("- ExcelのShape（図形）を検出し、Mermaid記法のフローチャートとして出力しています\n")
                    f.write("- 各シートのCSVブロックの前に、検出されたShapeがMermaidコードブロックで記載されます\n")
                    f.write("- Shape間の接続（コネクタ）も矢印として表現されます\n\n")

                # About Verification Metadata (level 3) - only when metadata is enabled (v1.7)
                include_metadata = opts.get("csv_include_metadata", True)
                if include_metadata:
                    f.write("### 検証用メタデータについて\n\n")
                    f.write(f"- このファイルの末尾に「検証用メタデータ」セクションが付記されています\n")
                    f.write(f"- メタデータには各シートのExcel原本情報（範囲、行数、列数）とCSV出力結果の比較が含まれます\n")
                    f.write(f"- 検証ステータス（OK/FAILED）により、CSVとExcel原本の整合性を確認できます\n")
                    f.write(f"- 詳細情報はファイル末尾の「検証用メタデータ」セクションを参照してください\n\n")

                # Separator between overview and CSV sections
                f.write("---\n\n")

            # Write each sheet's CSV data in code blocks
            for sheet_name, sheet_data in csv_data_dict.items():
                # For single-sheet mode, don't add sheet name header (already in title)
                if not is_single_sheet:
                    f.write(f"## {sheet_name}\n\n")

                # v1.7: Write Mermaid block before CSV if mermaid_enabled=true and mermaid exists
                if isinstance(sheet_data, dict) and sheet_data.get("mermaid"):
                    f.write(sheet_data["mermaid"])
                    f.write("\n\n")

                # Extract rows from new data structure
                rows = sheet_data["rows"] if isinstance(sheet_data, dict) else sheet_data

                if rows:
                    # Generate CSV string using csv.writer for RFC 4180 compliance
                    output = StringIO()
                    writer = csv_module.writer(output, delimiter=',', quoting=csv_module.QUOTE_MINIMAL, lineterminator='\n')
                    writer.writerows(rows)
                    csv_content = output.getvalue()

                    # Write CSV code block
                    f.write("```csv\n")
                    f.write(csv_content)
                    f.write("```\n\n")
                else:
                    f.write("```csv\n```\n\n")

        info(f"CSV markdown file created: {csv_md_path}")

        # Append validation metadata if enabled (using verify_csv_markdown module)
        if opts.get("csv_include_metadata", True):
            try:
                # Import verification module to append metadata
                import sys
                from pathlib import Path as PathLib

                # Add v1.5 directory to path if not already there
                verify_module_path = PathLib(__file__).parent
                if str(verify_module_path) not in sys.path:
                    sys.path.insert(0, str(verify_module_path))

                from verify_csv_markdown import append_verification_metadata_from_data

                # Call verification module to append metadata
                append_verification_metadata_from_data(
                    str(csv_md_path),
                    f"{excel_file_basename}.xlsx",
                    csv_data_dict,
                    opts
                )
            except Exception as e:
                warn(f"Failed to append verification metadata: {e}")

        return str(csv_md_path)

    except Exception as e:
        warn(f"Failed to write CSV markdown file: {e}")
        return None

def extract_print_area_for_csv(ws, area, opts, merged_lookup):
    """Extract all cell values from print area for CSV output per spec §3.2.2.

    Args:
        ws: Worksheet object
        area: Print area tuple (min_row, min_col, max_row, max_col) (1-based)
        opts: Options dictionary
        merged_lookup: Merged cell lookup (only includes cells within print area)

    Returns:
        List of lists (rows of cell values)
    """
    min_row, min_col, max_row, max_col = area
    rows = []

    for R in range(min_row, max_row + 1):
        row_vals = []
        for C in range(min_col, max_col + 1):
            cell = ws.cell(row=R, column=C)

            # Handle merged cells per spec §3.2.2
            tl = merged_lookup.get((R, C))
            if tl:
                if (R, C) == tl:
                    # Top-left cell: use its value
                    cell = ws.cell(row=tl[0], column=tl[1])
                    text = cell_display_value(cell, opts)
                else:
                    # Other cells in merged range
                    if opts.get("csv_apply_merge_policy", True):
                        # Apply merge_policy
                        if opts.get("merge_policy") == "top_left_only":
                            text = ""
                        else:
                            # expand/repeat: use top-left value
                            cell = ws.cell(row=tl[0], column=tl[1])
                            text = cell_display_value(cell, opts)
                    else:
                        # csv_apply_merge_policy=false: only top-left has value
                        text = ""
            else:
                text = cell_display_value(cell, opts)

            # Apply normalization if requested per spec §3.2.2
            if opts.get("csv_normalize_values", True):
                text = normalize_numeric_text(text, opts)

            # Handle hyperlinks per spec §3.2.2 (updated in v1.6)
            hl = hyperlink_info(cell)
            if hl:
                disp = text if text else (hl.get("display") or "")
                hyperlink_mode = opts.get("hyperlink_mode", "inline_plain")

                # Handle footnote/both modes -> fallback to inline_plain with warning
                if hyperlink_mode in ("footnote", "both"):
                    if not hasattr(extract_print_area_for_csv, '_csv_hyperlink_warning_shown'):
                        warn("CSV markdown does not support footnote mode. Falling back to inline_plain for hyperlinks.")
                        extract_print_area_for_csv._csv_hyperlink_warning_shown = True
                    hyperlink_mode = "inline_plain"

                if hl.get("target"):
                    link = hl["target"]
                    if hyperlink_mode == "inline":
                        text = f"[{disp}]({link})"
                    elif hyperlink_mode == "inline_plain":
                        text = f"{disp} ({link})"
                    elif hyperlink_mode == "text_only":
                        text = disp
                    else:
                        # Default to text_only if unknown mode
                        text = disp
                elif hl.get("location"):
                    loc = hl["location"]
                    if hyperlink_mode == "inline":
                        text = f"[{disp}]({loc})"
                    elif hyperlink_mode == "inline_plain":
                        text = f"{disp} (→{loc})"
                    elif hyperlink_mode == "text_only":
                        text = disp
                    else:
                        text = disp

            # Replace newlines with spaces for CSV markdown (per spec §3.2.2)
            # This ensures 1 record = 1 physical line for AI readability
            if text and isinstance(text, str):
                text = text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')

            row_vals.append(text)
        rows.append(row_vals)

    return rows

# Note: Old CSV file output functions (write_csv_per_sheet, write_csv_merged_book) have been removed.
# v1.5 now only outputs CSV markdown format via write_csv_markdown() function.
# ===== end CSV output functions =====

def run(input_path: str, output_path: Optional[str], args):
    wb = load_workbook_safe(input_path, read_only=args.read_only)
    sheets = wb.sheetnames

    split_by_sheet = getattr(args, "split_by_sheet", False)

    # For split_by_sheet mode, use a dict to store md_lines for each sheet
    if split_by_sheet:
        sheet_md_dict = {}  # {sheet_name: [md_lines]}
        sheet_footnotes_dict = {}  # {sheet_name: [(idx, text)]}
    else:
        md_lines = []
        md_lines.append(f"# 変換結果: {Path(input_path).name}")
        md_lines.append("")
        md_lines.append(f"- 仕様バージョン: {VERSION}")
        md_lines.append(f"- シート数: {len(sheets)}")
        md_lines.append(f"- シート一覧: {', '.join(sheets)}")
        md_lines.append("\n---\n")

    footnotes: List[Tuple[int,str]] = []
    global_footnote_start = 1
    sheet_counter = 0

    opts = {
        "no_print_area_mode": args.no_print_area_mode,
        "value_mode": args.value_mode,
        "merge_policy": args.merge_policy,
        "hyperlink_mode": args.hyperlink_mode,
        "header_detection": (args.header_detection == "first_row"),
        "hidden_policy": args.hidden_policy,
        "strip_whitespace": args.strip_whitespace,
        "escape_pipes": args.escape_pipes,
        "date_format_override": args.date_format_override,
        "date_default_format": args.date_default_format,
        "numeric_thousand_sep": args.numeric_thousand_sep,
        "percent_format": args.percent_format,
        "currency_symbol": args.currency_symbol,
        "percent_divide_100": args.percent_divide_100,
        "readonly_fill_policy": getattr(args, "readonly_fill_policy", "assume_no_fill"),
        "align_detection": (args.align_detection == "numbers_right"),
        "numbers_right_threshold": args.numbers_right_threshold,
        "max_sheet_count": args.max_sheet_count,
        "max_cells_per_table": args.max_cells_per_table,
        "sort_tables": args.sort_tables,
        "footnote_scope": args.footnote_scope,
        "locale": args.locale,
        "markdown_escape_level": args.markdown_escape_level,
        "mermaid_enabled": args.mermaid_enabled,
        "mermaid_detect_mode": args.mermaid_detect_mode,
        "mermaid_diagram_type": getattr(args, "mermaid_diagram_type", "flowchart"),
        "mermaid_direction": args.mermaid_direction,
        "mermaid_keep_source_table": getattr(args, "mermaid_keep_source_table", True),
        "mermaid_dedupe_edges": getattr(args, "mermaid_dedupe_edges", True),
        "mermaid_node_id_policy": getattr(args, "mermaid_node_id_policy", "auto"),
        "mermaid_group_column_behavior": getattr(args, "mermaid_group_column_behavior", "subgraph"),
        "mermaid_columns": (lambda s: {
            "from": (s.split(",")[0].strip() if len(s.split(","))>0 else "From"),
            "to": (s.split(",")[1].strip() if len(s.split(","))>1 else "To"),
            "label": (s.split(",")[2].strip() if len(s.split(","))>2 else "Label"),
            "group": (s.split(",")[3].strip() if len(s.split(","))>3 else None),
            "note": (s.split(",")[4].strip() if len(s.split(","))>4 else None),
        })(args.mermaid_columns),
        "mermaid_heuristic_min_rows": args.mermaid_heuristic_min_rows,
        "mermaid_heuristic_arrow_ratio": args.mermaid_heuristic_arrow_ratio,
        "mermaid_heuristic_len_median_ratio_min": args.mermaid_heuristic_len_median_ratio_min,
        "mermaid_heuristic_len_median_ratio_max": args.mermaid_heuristic_len_median_ratio_max,
        "dispatch_skip_code_and_mermaid_on_fallback": getattr(args, "dispatch_skip_code_and_mermaid_on_fallback", True),

        "detect_dates": True,
        "prefer_excel_display": args.prefer_excel_display,

        # CSV markdown output options (v1.5 spec §⑫, extended in v1.7)
        "csv_output_dir": getattr(args, "csv_output_dir", None),
        "csv_apply_merge_policy": getattr(args, "csv_apply_merge_policy", True),
        "csv_normalize_values": getattr(args, "csv_normalize_values", True),
        "csv_markdown_enabled": getattr(args, "csv_markdown_enabled", True),
        "csv_include_metadata": getattr(args, "csv_include_metadata", True),
        "csv_include_description": getattr(args, "csv_include_description", True),  # v1.7
    }

    # Prepare for CSV markdown output (v1.5 spec §⑫)
    csv_output_dir = opts.get("csv_output_dir") or str(Path(input_path).parent)
    csv_basename = Path(input_path).stem
    csv_markdown_data = {}  # For CSV markdown output (dict of sheet_name: rows)

    for sname in sheets:
        sheet_counter += 1
        ws = wb[sname]
        if getattr(getattr(ws, "protection", None), "sheet", False):
            info(f"Sheet '{sname}' is protected (read-only); proceeding with read-only extraction.")

        # Initialize current_md_lines for this sheet
        if split_by_sheet:
            current_md_lines = []
            current_md_lines.append(f"# {sname}")
            current_md_lines.append("")
            current_md_lines.append(f"- 仕様バージョン: {VERSION}")
            current_md_lines.append(f"- 元ファイル: {Path(input_path).name}")
            current_md_lines.append("\n---\n")
            sheet_md_dict[sname] = current_md_lines
            sheet_footnotes_dict[sname] = []
            if opts["footnote_scope"] == "book":
                # For split_by_sheet mode, treat each sheet as independent (sheet scope)
                footnotes = []
                global_footnote_start = 1
        else:
            current_md_lines = md_lines

        if opts["max_sheet_count"] and sheet_counter > opts["max_sheet_count"]:
            current_md_lines.append(f"## {sname}\n（シート数上限によりスキップ）\n\n---\n")
            continue

        # §4.⑩: 各シート処理開始時に、Markdown見出しとしてシート名を出力
        if not split_by_sheet:
            current_md_lines.append(f"## {sname}\n")

        # §⑦'': シェイプ（図形）からのMermaid検出（シート単位で実行）
        shapes_mermaid = None
        if opts.get("mermaid_enabled", False) and opts.get("mermaid_detect_mode") == "shapes":
            shapes_mermaid = _v14_extract_shapes_to_mermaid(input_path, ws, opts)
            if shapes_mermaid:
                current_md_lines.append(shapes_mermaid + "\n")
                current_md_lines.append("\n---\n")
                # シェイプ検出が成功してもテーブル処理は続行される（仕様書§3.2、§⑦''、§⑩参照）

        # §4.③: シートの印刷領域を取得
        areas = get_print_areas(ws, opts["no_print_area_mode"])
        if not areas:
            current_md_lines.append("（テーブルなし）\n\n---\n")
            continue

        # §5.3: 幾何学的和集合 → 矩形分解
        unioned = union_rects(areas)

        if opts["footnote_scope"] == "sheet":
            footnotes = []
            global_footnote_start = 1

        table_id = 0

        for union_area in unioned:
            # Per spec §6.4: build merged_lookup for each print area separately
            # to ensure only merged cells within the print area are processed
            merged_lookup = build_merged_lookup(ws, union_area)
            tables = grid_to_tables(ws, union_area, hidden_policy=opts["hidden_policy"], opts=opts)
            if not tables:
                continue

            # Process tables
            for tbl in tables:
                table_id += 1

                # Per spec §⑦: pass print_area to ensure only cells within print area are processed
                md_rows, note_refs, truncated, table_title = extract_table(ws, tbl, opts, footnotes, global_footnote_start, merged_lookup, print_area=union_area)

                # テーブル見出しの出力
                if table_title:
                    # §7.1: テーブルタイトルを検出した場合、それをテーブル見出しとして使用
                    current_md_lines.append(f"### {table_title}")
                else:
                    # 付録B.1: テーブル見出しには座標範囲を表示しない
                    current_md_lines.append(f"### Table {table_id}")
                for (n, txt) in note_refs:
                    footnotes.append((n, txt))

                if not md_rows:
                    current_md_lines.append("（テーブルなし）\n")
                    continue

                # §7.2: テーブル形式の判定と出力形式の選択
                format_type, formatted_output = dispatch_table_output(ws, tbl, md_rows, opts, merged_lookup, xlsx_path=input_path)

                if format_type == "text":
                    # §7.2.1: 単一行テキスト形式
                    current_md_lines.append(formatted_output + "\n")
                elif format_type == "nested":
                    # §7.2.3: ネスト形式
                    current_md_lines.append(formatted_output + "\n")
                elif format_type == "code":
                    # §7.2.4: ソースコード形式
                    current_md_lines.append(formatted_output + "\n")
                elif format_type == "mermaid":
                    # §7.6: Mermaid形式（formatted_outputにはMermaidコードと必要に応じて元テーブルが含まれる）
                    current_md_lines.append(formatted_output + "\n")
                elif format_type == "empty":
                    # §7.2.2: 空行の処理
                    current_md_lines.append("\n")
                else:
                    # §7.2.5: 通常のテーブル形式
                    # Note: dispatch_table_output() already handles table formatting,
                    # but this branch handles cases where format_type is "table" from dispatch
                    hdr = opts.get("header_detection", True)
                    table_md = make_markdown_table(
                        md_rows,
                        header_detection=hdr,
                        align_detect=opts["align_detection"],
                        align_threshold=opts["numbers_right_threshold"],
                    )
                    current_md_lines.append(table_md + "\n")
                    if truncated:
                        current_md_lines.append("_※ このテーブルは max_cells_per_table 制限により途中で打ち切られました。_\n")

            current_md_lines.append("\n---\n")

        # CSV markdown data collection per spec §⑫ (v1.5)
        # v1.5 only outputs CSV markdown format, not raw CSV files
        if opts.get("csv_markdown_enabled", True):
            # Collect CSV data for markdown output
            for union_area in unioned:
                merged_lookup = build_merged_lookup(ws, union_area)
                try:
                    csv_rows = extract_print_area_for_csv(ws, union_area, opts, merged_lookup)
                    if csv_rows:
                        # Store both CSV rows and Excel range info
                        excel_range = coords_to_excel_range(*union_area)
                        csv_markdown_data[sname] = {
                            "rows": csv_rows,
                            "range": excel_range,
                            "area": union_area,
                            "mermaid": None  # Will be populated if mermaid_enabled=true and mermaid_detect_mode="shapes"
                        }
                except Exception as e:
                    warn(f"CSV data extraction failed for sheet '{sname}': {e}")
                    # Continue processing (per spec §9.z)

            # v1.7: Extract Mermaid for CSV markdown if mermaid_enabled=true
            # Only mermaid_detect_mode="shapes" is supported for CSV markdown (per spec §3.2.2)
            if opts.get("mermaid_enabled", False) and sname in csv_markdown_data:
                detect_mode = opts.get("mermaid_detect_mode", "shapes")
                if detect_mode == "shapes":
                    try:
                        csv_mermaid = _v14_extract_shapes_to_mermaid(input_path, ws, opts)
                        if csv_mermaid:
                            csv_markdown_data[sname]["mermaid"] = csv_mermaid
                    except Exception as e:
                        warn(f"Mermaid extraction for CSV markdown failed for sheet '{sname}': {e}")
                elif detect_mode in ("column_headers", "heuristic"):
                    # column_headers and heuristic modes are not supported for CSV markdown
                    # because CSV markdown does not split tables
                    warn(f"mermaid_detect_mode='{detect_mode}' is not supported for CSV markdown output (only 'shapes' is supported). Mermaid output will be skipped for sheet '{sname}'.")

        # For split_by_sheet mode, save footnotes for this sheet and add to file
        if split_by_sheet:
            sheet_footnotes_dict[sname] = list(footnotes)
            if footnotes and opts["hyperlink_mode"] in ("footnote", "both"):
                footnotes_sorted = sorted(set(footnotes), key=lambda x: x[0])
                current_md_lines.append("\n")
                for idx, txt in footnotes_sorted:
                    current_md_lines.append(f"[^{idx}]: {txt}")

    # For non-split mode, add footnotes at the end
    if not split_by_sheet:
        if footnotes and opts["hyperlink_mode"] in ("footnote", "both"):
            footnotes_sorted = sorted(set(footnotes), key=lambda x: x[0])
            md_lines.append("\n")
            for idx, txt in footnotes_sorted:
                md_lines.append(f"[^{idx}]: {txt}")

    # Write output file(s) - exclusive mode per v1.6 spec §3.2
    # csv_markdown_enabled=true: CSV markdown only
    # csv_markdown_enabled=false: Regular markdown only
    if opts.get("csv_markdown_enabled", True):
        # CSV markdown output mode
        if csv_markdown_data:
            try:
                if split_by_sheet:
                    # Split by sheet: write each sheet to separate CSV markdown file
                    output_dir = Path(output_path).parent if output_path else Path(input_path).parent
                    output_basename = Path(output_path).stem if output_path else Path(input_path).stem
                    output_files = []

                    for sname in sheets:
                        if sname not in csv_markdown_data:
                            continue
                        # Sanitize sheet name for filename
                        safe_sheet_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in sname)
                        # Write single-sheet CSV markdown
                        single_sheet_data = {sname: csv_markdown_data[sname]}
                        csv_file = write_csv_markdown(
                            wb, single_sheet_data,
                            f"{output_basename}_{safe_sheet_name}",
                            opts, csv_output_dir or str(output_dir)
                        )
                        if csv_file:
                            output_files.append(csv_file)

                    return "\n".join([f"シートごとに分割してCSVマークダウン出力しました:"] + output_files)
                else:
                    # Normal mode: write all sheets to single CSV markdown file
                    csv_file = write_csv_markdown(wb, csv_markdown_data, csv_basename, opts, csv_output_dir)
                    return csv_file or "CSV markdown output completed"
            except Exception as e:
                warn(f"CSV markdown output failed: {e}")
                return None
        else:
            warn("No CSV data to output")
            return None

    # Regular markdown output mode (csv_markdown_enabled=false)
    if split_by_sheet:
        # Split by sheet mode: write each sheet to a separate file
        output_dir = Path(output_path).parent if output_path else Path(input_path).parent
        output_basename = Path(output_path).stem if output_path else Path(input_path).stem
        output_files = []

        for sname in sheets:
            if sname not in sheet_md_dict:
                continue
            # Sanitize sheet name for filename
            safe_sheet_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in sname)
            sheet_output_path = output_dir / f"{output_basename}_{safe_sheet_name}.md"
            Path(sheet_output_path).write_text("\n".join(sheet_md_dict[sname]), encoding="utf-8")
            output_files.append(str(sheet_output_path))

        return "\n".join([f"シートごとに分割して出力しました:"] + output_files)
    else:
        # Normal mode: write all sheets to a single file
        if not output_path:
            output_path = str(Path(input_path).with_suffix(".md"))
        Path(output_path).write_text("\n".join(md_lines), encoding="utf-8")
        return output_path

def build_argparser():
    p = argparse.ArgumentParser(description=f"Excel -> Markdown converter (Spec v{VERSION})")
    p.add_argument("input", help="Path to .xlsx/.xlsm file")
    p.add_argument("-o", "--output", help="Path to output .md file")
    # Spec defaults
    p.add_argument("--no-print-area-mode", choices=["used_range","entire_sheet_range","skip_sheet"], default="used_range")
    p.add_argument("--value-mode", choices=["display","formula","both"], default="display")
    p.add_argument("--merge-policy", choices=["expand","repeat","warn","top_left_only"], default="top_left_only")
    p.add_argument("--hyperlink-mode", choices=["inline","inline_plain","footnote","both","text_only"], default="footnote")
    p.add_argument("--header-detection", choices=["none","first_row","heuristic"], default="first_row")
    p.add_argument("--mermaid-enabled", action="store_true", default=False)
    p.add_argument("--mermaid-detect-mode", choices=["none","column_headers","heuristic","shapes"], default="shapes")
    p.add_argument("--mermaid-diagram-type", choices=["flowchart","sequence","state"], default="flowchart")
    p.add_argument("--mermaid-direction", choices=["TD","LR","BT","RL"], default="TD")
    p.add_argument("--mermaid-keep-source-table", action="store_true", default=True)
    p.add_argument("--no-mermaid-keep-source-table", dest="mermaid_keep_source_table", action="store_false")
    p.add_argument("--mermaid-dedupe-edges", action="store_true", default=True)
    p.add_argument("--no-mermaid-dedupe-edges", dest="mermaid_dedupe_edges", action="store_false")
    p.add_argument("--mermaid-node-id-policy", choices=["auto","shape_id","explicit"], default="auto")
    p.add_argument("--mermaid-group-column-behavior", choices=["subgraph","ignore"], default="subgraph")
    p.add_argument("--mermaid-columns", default="From,To,Label,Group,Note")
    p.add_argument("--mermaid-heuristic-min-rows", type=int, default=3)
    p.add_argument("--mermaid-heuristic-arrow-ratio", type=float, default=0.3)
    p.add_argument("--mermaid-heuristic-len-median-ratio-min", type=float, default=0.4)
    p.add_argument("--mermaid-heuristic-len-median-ratio-max", type=float, default=2.5)
    p.add_argument("--dispatch-skip-code-and-mermaid-on-fallback", action="store_true", default=True)
    p.add_argument("--no-dispatch-skip-code-and-mermaid-on-fallback", dest="dispatch_skip_code_and_mermaid_on_fallback", action="store_false")

    p.add_argument("--hidden-policy", choices=["ignore","include","exclude"], default="ignore")
    p.add_argument("--strip-whitespace", action="store_true", default=True)
    p.add_argument("--no-strip-whitespace", dest="strip_whitespace", action="store_false")
    p.add_argument("--escape-pipes", action="store_true", default=True)
    p.add_argument("--no-escape-pipes", dest="escape_pipes", action="store_false")
    p.add_argument("--date-format-override", default=None)
    p.add_argument("--date-default-format", default="YYYY-MM-DD")
    p.add_argument("--numeric-thousand-sep", choices=["keep","remove"], default="keep")
    p.add_argument("--percent-format", choices=["keep","numeric"], default="keep")
    p.add_argument("--percent-divide-100", action="store_true", help="When percent-format=numeric, divide by 100 (e.g., 12%% -> 0.12)")
    p.add_argument("--currency-symbol", choices=["keep","strip"], default="keep")
    p.add_argument("--align-detection", choices=["none","numbers_right"], default="numbers_right")
    p.add_argument("--numbers-right-threshold", type=float, default=0.8)
    p.add_argument("--max-sheet-count", type=int, default=0)
    p.add_argument("--max-cells-per-table", type=int, default=200000)
    p.add_argument("--sort-tables", choices=["document_order"], default="document_order")
    p.add_argument("--footnote-scope", choices=["book","sheet"], default="book")
    p.add_argument("--locale", default="ja-JP")
    p.add_argument("--markdown-escape-level", choices=["safe","minimal","aggressive"], default="safe")
    p.add_argument("--read-only", action="store_true", help="Use openpyxl read_only=True (styles may be limited)")
    p.add_argument("--prefer-excel-display", action="store_true", default=True, help="Prefer Excel displayed value over formatted output when possible")

    # CSV output options (v1.5)
    # CSV Markdown output options (v1.5 spec §⑫)
    # Note: v1.5 only outputs CSV markdown format, not raw CSV files
    p.add_argument("--csv-output-dir", default=None, help="CSV markdown output directory (default: same as input file)")
    p.add_argument("--csv-apply-merge-policy", action="store_true", default=True, help="Apply merge_policy to CSV data extraction")
    p.add_argument("--no-csv-apply-merge-policy", dest="csv_apply_merge_policy", action="store_false")
    p.add_argument("--csv-normalize-values", action="store_true", default=True, help="Apply numeric normalization to CSV values")
    p.add_argument("--no-csv-normalize-values", dest="csv_normalize_values", action="store_false")
    p.add_argument("--csv-markdown-enabled", action="store_true", default=True, help="Enable CSV markdown output (default: True)")
    p.add_argument("--no-csv-markdown-enabled", dest="csv_markdown_enabled", action="store_false")
    p.add_argument("--csv-include-metadata", action="store_true", default=True, help="Include validation metadata in CSV markdown output (default: True)")
    p.add_argument("--no-csv-include-metadata", dest="csv_include_metadata", action="store_false")
    p.add_argument("--csv-include-description", action="store_true", default=True, help="Include description section in CSV markdown output (default: True)")
    p.add_argument("--no-csv-include-description", dest="csv_include_description", action="store_false")

    # Sheet output options
    p.add_argument("--split-by-sheet", action="store_true", default=False, help="Split output by sheet (generate separate file for each sheet)")

    return p

def main(argv=None):
    parser = build_argparser()
    args = parser.parse_args(argv)
    out = run(args.input, args.output, args)
    print(out)

if __name__ == "__main__":
    main()
