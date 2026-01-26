# -*- coding: utf-8 -*-
"""Table detection utilities.

仕様書参照: §4.2 テーブル検出フロー、§5.4 セル空性の判定
"""

from typing import List, Tuple, Set, Dict

from .cell_utils import cell_is_empty, cell_display_value, no_fill
from .output import warn

def collect_hidden(ws):
    hidden_rows = set(i for i, d in ws.row_dimensions.items() if getattr(d, "hidden", False))
    hidden_cols_idx = set()
    for key, d in ws.column_dimensions.items():
        if getattr(d, "hidden", False):
            from openpyxl.utils import column_index_from_string
            try:
                hidden_cols_idx.add(column_index_from_string(key))
            except Exception:
                pass
    return hidden_rows, hidden_cols_idx

def build_nonempty_grid(ws, area, hidden_policy="ignore", opts=None) -> Tuple[List[List[int]], int, int, int, int]:
    r0, c0, r1, c1 = area
    rows = r1 - r0 + 1
    cols = c1 - c0 + 1
    grid = [[0]*cols for _ in range(rows)]
    merged_blocks = list(ws.merged_cells.ranges) if getattr(ws, "merged_cells", None) else []
    merged_coords = []
    for rng in merged_blocks:
        min_row, min_col, max_row, max_col = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        if max_row < r0 or min_row > r1 or max_col < c0 or min_col > c1:
            continue
        if min_row < r0 or min_col < c0:
            continue  # Top-left cell is outside print area
        # Only include the part of merged cell that is within print area
        merged_coords.append((max(r0, min_row), max(c0, min_col), min(r1, max_row), min(c1, max_col)))

    hidden_rows, hidden_cols_idx = collect_hidden(ws)

    for rr in range(rows):
        for cc in range(cols):
            R = r0 + rr
            C = c0 + cc
            if hidden_policy == "exclude" and (R in hidden_rows or C in hidden_cols_idx):
                continue
            cell = ws.cell(row=R, column=C)
            empty = cell_is_empty(cell, opts)
            grid[rr][cc] = 0 if empty else 1

    # merged influence
    for mr0, mc0, mr1, mc1 in merged_coords:
        found_nonempty = False
        for R in range(mr0, mr1+1):
            for C in range(mc0, mc1+1):
                if R < r0 or R > r1 or C < c0 or C > c1:
                    continue
                cell = ws.cell(row=R, column=C)
                if not cell_is_empty(cell, opts):
                    found_nonempty = True
                    break
            if found_nonempty:
                break
        if found_nonempty:
            for R in range(mr0, mr1+1):
                for C in range(mc0, mc1+1):
                    if R < r0 or R > r1 or C < c0 or C > c1:
                        continue
                    # Convert to grid coordinates
                    grid_r = R - r0
                    grid_c = C - c0
                    if 0 <= grid_r < rows and 0 <= grid_c < cols:
                        grid[grid_r][grid_c] = 1
    return grid, r0, c0, r1, c1

def enumerate_histogram_rectangles(grid: List[List[int]]) -> List[Tuple[int,int,int,int]]:
    if not grid or not grid[0]:
        return []
    R = len(grid)
    C = len(grid[0])
    H = [0]*C
    rects = []
    for r in range(R):
        for c in range(C):
            H[c] = H[c] + 1 if grid[r][c] == 1 else 0
        stack = []  # (start_index, height)
        for c in range(C + 1):
            h = H[c] if c < C else 0
            last = c
            while stack and stack[-1][1] > h:
                i, hh = stack.pop()
                width = c - i
                top = r - hh + 1
                left = i
                bottom = r
                right = c - 1
                if hh > 0 and width > 0:
                    rects.append((top, left, bottom, right))
                last = i
            if not stack or stack[-1][1] < h:
                stack.append((last, h))
    return rects

def carve_rectangles(grid: List[List[int]]) -> List[Tuple[int,int,int,int]]:
    def any_ones(g):
        for row in g:
            if any(v == 1 for v in row):
                return True
        return False

    out = []
    g = [row[:] for row in grid]
    while any_ones(g):
        candidates = enumerate_histogram_rectangles(g)
        if not candidates:
            break
        candidates.sort(key=lambda r: ((r[2]-r[0]+1)*(r[3]-r[1]+1)), reverse=True)
        top,left,bottom,right = candidates[0]
        out.append((top,left,bottom,right))
        for r in range(top, bottom+1):
            for c in range(left, right+1):
                g[r][c] = 0
    out.sort(key=lambda r: (r[0], r[1], (r[2]-r[0]+1)*(r[3]-r[1]+1)))
    return out

def bfs_components(grid: List[List[int]]) -> List[Set[Tuple[int,int]]]:
    R = len(grid); C = len(grid[0]) if R else 0
    seen = [[False]*C for _ in range(R)]
    comps = []
    from collections import deque
    for r in range(R):
        for c in range(C):
            if grid[r][c] == 1 and not seen[r][c]:
                q = deque()
                q.append((r,c)); seen[r][c] = True
                comp = set()
                comp.add((r,c))
                while q:
                    rr,cc = q.popleft()
                    for dr,dc in ((1,0),(-1,0),(0,1),(0,-1)):
                        nr,nc = rr+dr, cc+dc
                        if 0<=nr<R and 0<=nc<C and grid[nr][nc]==1 and not seen[nr][nc]:
                            seen[nr][nc]=True
                            q.append((nr,nc))
                            comp.add((nr,nc))
                comps.append(comp)
    return comps

def rectangles_for_component(comp: Set[Tuple[int,int]], grid_shape: Tuple[int,int]) -> List[Tuple[int,int,int,int]]:
    """Build a mask for the component and carve rectangles inside it."""
    R,C = grid_shape
    # Build component-only grid
    g = [[0]*C for _ in range(R)]
    for (r,c) in comp:
        g[r][c] = 1
    return carve_rectangles(g)

def union_rects(rects: List[Tuple[int,int,int,int]]) -> List[Tuple[int,int,int,int]]:
    """Line-sweep union of rectangles (row-based), returns disjoint rectangles covering union."""
    if not rects:
        return []
    events = {}
    min_r = min(r0 for r0,_,r1,_ in rects)
    max_r = max(r1 for r0,_,r1,_ in rects)
    for r0,c0,r1,c1 in rects:
        events.setdefault(r0, []).append(("add",(c0,c1)))
        events.setdefault(r1+1, []).append(("rem",(c0,c1)))
    active: List[Tuple[int,int]] = []
    def merge_intervals(iv):
        if not iv: return []
        iv.sort()
        out = [list(iv[0])]
        for s,e in iv[1:]:
            if s <= out[-1][1]+0:
                out[-1][1] = max(out[-1][1], e)
            else:
                out.append([s,e])
        return [(a,b) for a,b in out]

    out_rects = []
    prev_row = None
    prev_spans = []
    # sweep rows
    for row in range(min_r, max_r+2):  # include sentinel
        for ev in events.get(row, []):
            kind,(s,e) = ev
            if kind=="add":
                active.append((s,e))
            else:
                try:
                    active.remove((s,e))
                except ValueError:
                    pass
        spans = merge_intervals(active)
        if prev_row is None:
            prev_row = row
            prev_spans = [(s,e) for s,e in spans]
            continue
        # If spans changed, close previous run
        if spans != prev_spans:
            # Close rectangles from prev_row to row-1
            for s,e in prev_spans:
                out_rects.append((prev_row, s, row-1, e))
            prev_row = row
            prev_spans = [(s,e) for s,e in spans]
    return out_rects

def grid_to_tables(ws, area, hidden_policy="ignore", opts=None):
    """Return list of logical tables; each table: dict with 'rects' (list of sheet-coord rects) and 'bbox'"""
    if opts is None:
        opts = {}
    grid, r0, c0, r1, c1 = build_nonempty_grid(ws, area, hidden_policy=hidden_policy, opts=opts)

    # Detect empty rows and columns to split tables
    # Note: Check actual cells, not grid (which may be marked by merged cells)
    R = len(grid)
    C = len(grid[0]) if R else 0

    # Find completely empty rows by checking actual cells
    # For row emptiness detection, we check:
    # 1. Cell value (text) - must be empty or whitespace only
    # 2. Cell fill - white fill is treated as no fill (same as blank)
    #    Any other color means the cell is not empty
    merged_lookup = build_merged_lookup(ws, area)

    empty_rows = set()
    for r in range(R):
        row_num = r0 + r
        is_empty = True
        for c in range(C):
            col_num = c0 + c
            # Check if this cell is part of a merged range
            tl = merged_lookup.get((row_num, col_num))
            if tl:
                # Only check the top-left cell of merged ranges
                if (row_num, col_num) != tl:
                    continue  # Skip non-top-left cells in merged ranges
                cell = ws.cell(row=tl[0], column=tl[1])
            else:
                cell = ws.cell(row=row_num, column=col_num)
            # Check cell value and fill
            text = cell_display_value(cell, opts)
            has_text = text and text.strip()
            # If cell has fill (not white), it's not empty
            has_fill = not no_fill(cell, opts.get("readonly_fill_policy", "assume_no_fill"))
            if has_text or has_fill:
                is_empty = False
                break
        if is_empty:
            empty_rows.add(r)

    # Find completely empty columns by checking actual cells
    # For column emptiness detection, we check:
    # 1. Cell value (text) - must be empty or whitespace only
    # 2. Cell fill - white fill is treated as no fill (same as blank)
    #    Any other color means the cell is not empty
    # We need to consider merged cells: only check the top-left cell of merged ranges.
    empty_cols = set()
    for c in range(C):
        col_num = c0 + c
        is_empty = True
        for r in range(R):
            row_num = r0 + r
            # Check if this cell is part of a merged range
            tl = merged_lookup.get((row_num, col_num))
            if tl:
                # Only check the top-left cell of merged ranges
                if (row_num, col_num) != tl:
                    continue  # Skip non-top-left cells in merged ranges
                cell = ws.cell(row=tl[0], column=tl[1])
            else:
                cell = ws.cell(row=row_num, column=col_num)
            # Check cell value and fill
            text = cell_display_value(cell, opts)
            has_text = text and text.strip()
            # If cell has fill (not white), it's not empty
            has_fill = not no_fill(cell, opts.get("readonly_fill_policy", "assume_no_fill"))
            if has_text or has_fill:
                is_empty = False
                break
        if is_empty:
            empty_cols.add(c)

    # Split grid by empty rows/columns: only connect cells horizontally or vertically
    # but not across empty rows/columns
    def is_connected(r1, c1, r2, c2):
        """Check if two adjacent cells can be in the same table (not separated by empty row/col)"""
        # For adjacent cells (dr=±1 or dc=±1), check if there's an empty row/col between them
        # Same row (horizontal neighbors): check if any column between is empty
        if r1 == r2:
            min_c, max_c = min(c1, c2), max(c1, c2)
            # For adjacent cells, max_c - min_c should be 1, so check if min_c or max_c is in empty_cols
            return not (min_c in empty_cols or max_c in empty_cols)
        # Same column (vertical neighbors): check if any row between is empty
        if c1 == c2:
            min_r, max_r = min(r1, r2), max(r1, r2)
            # For adjacent cells, max_r - min_r should be 1, so check if min_r or max_r is in empty_rows
            return not (min_r in empty_rows or max_r in empty_rows)
        # Different row and column: check if they can be connected via a path
        # Two cells in different rows and columns can be in the same table if:
        # 1. They are in adjacent rows (dr=±1) and their columns are not separated by empty columns
        # 2. They are in adjacent columns (dc=±1) and their rows are not separated by empty rows
        dr = abs(r2 - r1)
        dc = abs(c2 - c1)
        if dr == 1 and dc == 1:
            # Diagonal neighbors: check if both horizontal and vertical paths are not blocked
            # Check horizontal path: (r1, c1) -> (r1, c2) -> (r2, c2)
            h_path_ok = not (c1 in empty_cols or c2 in empty_cols) and r1 not in empty_rows
            # Check vertical path: (r1, c1) -> (r2, c1) -> (r2, c2)
            v_path_ok = not (r1 in empty_rows or r2 in empty_rows) and c1 not in empty_cols
            # If either path is open, they can be connected
            return h_path_ok or v_path_ok
        # Not adjacent: not directly connected
        return False

    seen = [[False]*C for _ in range(R)]
    comps = []
    from collections import deque
    for r in range(R):
        for c in range(C):
            if grid[r][c] == 1 and not seen[r][c]:
                q = deque()
                q.append((r,c)); seen[r][c] = True
                comp = set()
                comp.add((r,c))
                while q:
                    rr,cc = q.popleft()
                    for dr,dc in ((1,0),(-1,0),(0,1),(0,-1),(1,1),(1,-1),(-1,1),(-1,-1)):
                        nr,nc = rr+dr, cc+dc
                        if 0<=nr<R and 0<=nc<C and grid[nr][nc]==1 and not seen[nr][nc]:
                            # Only connect if not separated by empty row/column
                            if is_connected(rr, cc, nr, nc):
                                seen[nr][nc]=True
                                q.append((nr,nc))
                                comp.add((nr,nc))
                comps.append(comp)

    tables = []
    for comp in comps:
        rects_local = rectangles_for_component(comp, (len(grid), len(grid[0])))
        rects_sheet = [(r0+t, c0+l, r0+b, c0+r) for (t,l,b,r) in rects_local]
        min_r = min(r for r,c in comp)
        max_r = max(r for r,c in comp)
        min_c = min(c for r,c in comp)
        max_c = max(c for r,c in comp)
        bbox = (r0+min_r, c0+min_c, r0+max_r, c0+max_c)
        mask = {(r0+r, c0+c) for (r,c) in comp if r0 <= r0+r <= r1 and c0 <= c0+c <= c1}
        bbox_min_row = max(bbox[0], r0)
        bbox_min_col = max(bbox[1], c0)
        bbox_max_row = min(bbox[2], r1)
        bbox_max_col = min(bbox[3], c1)
        bbox = (bbox_min_row, bbox_min_col, bbox_max_row, bbox_max_col)
        tables.append({"rects": rects_sheet, "bbox": bbox, "mask": mask})
    # sort by (top,left)
    tables.sort(key=lambda t: (t["bbox"][0], t["bbox"][1]))
    return tables

def build_merged_lookup(ws, area=None):
    """Map each cell (r,c) in a merged range to its top-left (r0,c0).

    If area is provided, only cells within the print area are included.

    Args:
        ws: Worksheet object
        area: Optional print area tuple (r0, c0, r1, c1). If provided, only merged cells
              that intersect with the print area are included, and only cells within
              the print area are added to the lookup.
    """
    lookup = {}
    for rng in getattr(ws, "merged_cells", []):
        try:
            r0,c0,r1,c1 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        except Exception:
            continue

        if area is not None:
            area_r0, area_c0, area_r1, area_c1 = area
            if r1 < area_r0 or r0 > area_r1 or c1 < area_c0 or c0 > area_c1:
                continue
            if r0 < area_r0 or c0 < area_c0:
                continue

        for R in range(r0, r1+1):
            for C in range(c0, c1+1):
                if area is not None:
                    area_r0, area_c0, area_r1, area_c1 = area
                    if R < area_r0 or R > area_r1 or C < area_c0 or C > area_c1:
                        continue  # This cell is outside print area
                lookup[(R,C)] = (r0,c0)
    return lookup
