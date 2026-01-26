# -*- coding: utf-8 -*-
"""Workbook loading and coordinate utilities.

仕様書参照: §3.1 入力、§4.1 全体処理フロー
"""

from .output import warn, info


def load_workbook_safe(path, read_only=False):
    try:
        from openpyxl import load_workbook
        # read_only=False で fill 等のスタイル参照を確実化
        wb = load_workbook(filename=path, read_only=read_only, data_only=True)
        return wb
    except Exception as e:
        import sys
        print(f"[ERROR] Failed to open workbook: {e}", file=sys.stderr)
        sys.exit(2)


def a1_from_rc(r: int, c: int) -> str:
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(c)}{r}"


def parse_dimension(ws):
    dim = ws.calculate_dimension()
    from openpyxl.utils import range_boundaries
    try:
        min_col, min_row, max_col, max_row = range_boundaries(dim)
        return (min_row, min_col, max_row, max_col)
    except Exception:
        return (1, 1, ws.max_row, ws.max_column)


def get_print_areas(ws, mode: str):
    """Get print areas from worksheet, with validation."""
    areas = []
    sheet_max_row = ws.max_row
    sheet_max_col = ws.max_column

    try:
        pa = ws.print_area
        if pa:
            if isinstance(pa, str):
                rngs = [pa]
            else:
                try:
                    rngs = list(pa)
                except (TypeError, AttributeError):
                    rngs = [pa]
            for r in rngs:
                try:
                    from openpyxl.utils import range_boundaries
                    range_str = str(r)
                    if '!' in range_str:
                        range_str = range_str.split('!', 1)[1]
                    (min_col, min_row, max_col, max_row) = range_boundaries(range_str)
                    if min_row > max_row or min_col > max_col:
                        warn(f"Invalid print area range (min > max): {r} in sheet '{ws.title}'")
                        continue
                    if min_row < 1 or min_col < 1 or max_row < 1 or max_col < 1:
                        warn(f"Invalid print area range (negative or zero): {r} in sheet '{ws.title}'")
                        continue
                    if max_row > sheet_max_row:
                        warn(f"Print area max_row ({max_row}) exceeds sheet max_row ({sheet_max_row}), limiting to {sheet_max_row} in sheet '{ws.title}'")
                        max_row = sheet_max_row
                    if max_col > sheet_max_col:
                        warn(f"Print area max_col ({max_col}) exceeds sheet max_col ({sheet_max_col}), limiting to {sheet_max_col} in sheet '{ws.title}'")
                        max_col = sheet_max_col
                    if min_row > sheet_max_row or min_col > sheet_max_col:
                        warn(f"Print area min_row/min_col exceeds sheet maximum, skipping range {r} in sheet '{ws.title}'")
                        continue
                    min_row = min(min_row, sheet_max_row)
                    min_col = min(min_col, sheet_max_col)
                    areas.append((min_row, min_col, max_row, max_col))
                except Exception as e:
                    warn(f"Failed to parse print area range '{r}' in sheet '{ws.title}': {e}")
                    continue
    except Exception as e:
        warn(f"Failed to get print area from sheet '{ws.title}': {e}")

    if areas:
        area_str = ", ".join([f"({r0},{c0},{r1},{c1})" for r0, c0, r1, c1 in areas])
        info(f"Print area for sheet '{ws.title}': [{area_str}]")
        return areas

    if mode == "skip_sheet":
        warn(f"No print area set for sheet '{ws.title}', skipping sheet (no_print_area_mode=skip_sheet)")
        return []
    if mode == "entire_sheet_range":
        fallback_area = (1, 1, sheet_max_row, sheet_max_col)
        info(f"No print area set for sheet '{ws.title}', using entire sheet range: {fallback_area}")
        return [fallback_area]
    mr, mc, MR, MC = parse_dimension(ws)
    fallback_area = (mr, mc, MR, MC)
    info(f"No print area set for sheet '{ws.title}', using used range: {fallback_area}")
    return [fallback_area]
