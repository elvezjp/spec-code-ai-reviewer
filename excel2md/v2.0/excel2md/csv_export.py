# -*- coding: utf-8 -*-
"""CSV markdown export utilities.

仕様書参照: §3.2.1 CSV Markdown
"""

from pathlib import Path
from typing import Dict, Tuple, Optional

from .cell_utils import cell_display_value, normalize_numeric_text, hyperlink_info
from .output import warn, info
from .workbook_loader import a1_from_rc
from . import __version__ as VERSION

def coords_to_excel_range(min_row, min_col, max_row, max_col):
    """Convert cell coordinates to Excel range string (e.g., A1:D10).

    Args:
        min_row: Minimum row (1-based)
        min_col: Minimum column (1-based)
        max_row: Maximum row (1-based)
        max_col: Maximum column (1-based)

    Returns:
        Excel range string (e.g., "A1:D10")
    """
    from openpyxl.utils import get_column_letter
    start_cell = f"{get_column_letter(min_col)}{min_row}"
    end_cell = f"{get_column_letter(max_col)}{max_row}"
    return f"{start_cell}:{end_cell}"

def format_timestamp():
    """Format current timestamp as YYYY-MM-DD HH:MM:SS.

    Returns:
        Formatted timestamp string
    """
    from datetime import datetime
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def write_csv_markdown(wb, csv_data_dict, excel_file_basename, opts, output_dir):
    """Write CSV markdown file containing sheets' CSV data.

    Args:
        wb: Workbook object
        csv_data_dict: Dictionary of {sheet_name: rows} where rows is list of lists
                       Can contain single sheet (for split mode) or all sheets (for normal mode)
        excel_file_basename: Base name for output file (without extension)
                            For split mode: includes sheet name (e.g., "file_Sheet1")
                            For normal mode: just file name (e.g., "file")
        opts: Options dictionary
        output_dir: Output directory path

    Returns:
        Path to created CSV markdown file, or None if failed
    """
    import csv as csv_module
    from io import StringIO
    import json

    try:
        if not csv_data_dict:
            return None

        # Prepare output file path
        output_dir_path = Path(output_dir)
        output_dir_path.mkdir(parents=True, exist_ok=True)
        csv_md_path = output_dir_path / f"{excel_file_basename}_csv.md"

        # Get current timestamp
        timestamp = format_timestamp()

        # Determine if this is single-sheet or multi-sheet output
        is_single_sheet = len(csv_data_dict) == 1
        sheet_names = list(csv_data_dict.keys())

        include_description = opts.get("csv_include_description", True)

        with open(csv_md_path, 'w', encoding='utf-8') as f:
            if is_single_sheet:
                f.write(f"# {sheet_names[0]}\n\n")
                if include_description:
                    f.write(f"- 仕様バージョン: {VERSION}\n")
                    f.write(f"- 元ファイル: {wb.properties.title or Path(excel_file_basename).name}.xlsx\n\n")
            else:
                f.write(f"# CSV出力: {excel_file_basename}.xlsx\n\n")

            # Overview section
            if include_description:
                f.write("## 概要\n\n")

                # File Information (level 3) - only for multi-sheet
                if not is_single_sheet:
                    f.write("### ファイル情報\n\n")
                    f.write(f"- 元のExcelファイル名: {excel_file_basename}.xlsx\n")
                    f.write(f"- シート数: {len(csv_data_dict)}\n")
                    f.write(f"- 生成日時: {timestamp}\n\n")

                    # About This File (level 3)
                    f.write("### このファイルについて\n\n")
                    f.write("このCSVマークダウンファイルは、AIがExcelの内容を理解できるよう、各シートの印刷領域をCSV形式で出力したファイルです。")
                    f.write("各シートはマークダウン見出しで区切られ、CSVコードブロックで内容が記載されています。")
                else:
                    # Single-sheet: simplified description
                    f.write("このCSVマークダウンファイルは、AIがExcelの内容を理解できるよう、シートの印刷領域をCSV形式で出力したファイルです。")
                # Add metadata description only if metadata is enabled
                if opts.get("csv_include_metadata", True):
                    f.write("ファイル末尾に検証用メタデータセクションがあり、Excel原本との整合性を確認できます。")
                f.write("\n\n")

                # CSV Generation Method (level 3)
                f.write("### CSV生成方法\n\n")
                sheet_text = "シート" if is_single_sheet else "各シート"
                f.write(f"- **出力対象範囲**: {sheet_text}の印刷領域のみを出力（印刷領域外のセルは含まない）\n")
                f.write(f"- **印刷領域の統合**: 複数の印刷領域がある場合は外接矩形として統合\n")
                f.write(f"- **テーブル分割なし**: Markdown出力と異なり、空行・空列でテーブル分割せず印刷領域全体を1つのCSVとして出力\n")
                f.write(f"- **結合セルの処理**: `{opts.get('merge_policy', 'top_left_only')}` 設定に従って処理\n")
                f.write(f"- **数式の扱い**: `{opts.get('value_mode', 'display')}` 設定に従って表示値・数式・両方のいずれかを出力\n")
                f.write(f"- **値の正規化**: `{opts.get('csv_normalize_values', True)}` 設定に従って正規化処理を適用\n\n")

                # CSV Format Specification (level 3)
                f.write("### CSV形式の仕様\n\n")
                f.write(f"- **区切り文字**: `{opts.get('csv_delimiter', ',')}`\n")
                f.write(f"- **引用符の使用**: `{opts.get('csv_quoting', 'minimal')}` 設定に従い、RFC 4180準拠で処理\n")
                f.write(f"- **セル内改行**: 半角スペースに変換される（1レコード=1行を保証）\n")
                f.write(f"- **セル内特殊文字**: 区切り文字や引用符は引用符でエスケープされる\n")
                f.write(f"- **エンコーディング**: `{opts.get('csv_encoding', 'utf-8')}`\n")
                f.write(f"- **空セルの表現**: 空文字列として出力（空行・空列も含めて出力される）\n")
                # Hyperlink mode description
                hyperlink_mode = opts.get("hyperlink_mode", "inline_plain")
                if hyperlink_mode == "text_only":
                    f.write(f"- **ハイパーリンク**: 表示テキストのみを出力（リンク先URLは含まない）\n\n")
                elif hyperlink_mode == "inline":
                    f.write(f"- **ハイパーリンク**: Markdown形式で出力（例: `[表示テキスト](URL)` または `[表示テキスト](シート名+セル番号)`）\n\n")
                else:  # inline_plain or footnote or both (after fallback)
                    f.write(f"- **ハイパーリンク**: 平文形式で出力（例: `表示テキスト (URL)` または `表示テキスト (→シート名+セル番号)`）\n\n")

                # About Mermaid (level 3) - only when mermaid is enabled with shapes mode
                mermaid_enabled = opts.get("mermaid_enabled", False)
                mermaid_detect_mode = opts.get("mermaid_detect_mode", "shapes")
                if mermaid_enabled and mermaid_detect_mode == "shapes":
                    f.write("### Mermaidフローチャートについて\n\n")
                    f.write("- ExcelのShape（図形）を検出し、Mermaid記法のフローチャートとして出力しています\n")
                    f.write("- 各シートのCSVブロックの前に、検出されたShapeがMermaidコードブロックで記載されます\n")
                    f.write("- Shape間の接続（コネクタ）も矢印として表現されます\n\n")

                # About Verification Metadata (level 3)
                include_metadata = opts.get("csv_include_metadata", True)
                if include_metadata:
                    f.write("### 検証用メタデータについて\n\n")
                    f.write(f"- このファイルの末尾に「検証用メタデータ」セクションが付記されています\n")
                    f.write(f"- メタデータには各シートのExcel原本情報（範囲、行数、列数）とCSV出力結果の比較が含まれます\n")
                    f.write(f"- 検証ステータス（OK/FAILED）により、CSVとExcel原本の整合性を確認できます\n")
                    f.write(f"- 詳細情報はファイル末尾の「検証用メタデータ」セクションを参照してください\n\n")

                # Separator between overview and CSV sections
                f.write("---\n\n")

            # Write each sheet's CSV data in code blocks
            for sheet_name, sheet_data in csv_data_dict.items():
                # For single-sheet mode, don't add sheet name header (already in title)
                if not is_single_sheet:
                    f.write(f"## {sheet_name}\n\n")

                # Write Mermaid block before CSV if mermaid_enabled=true and mermaid exists
                if isinstance(sheet_data, dict) and sheet_data.get("mermaid"):
                    f.write(sheet_data["mermaid"])
                    f.write("\n\n")

                # Extract rows from new data structure
                rows = sheet_data["rows"] if isinstance(sheet_data, dict) else sheet_data

                if rows:
                    # Generate CSV string using csv.writer for RFC 4180 compliance
                    output = StringIO()
                    writer = csv_module.writer(output, delimiter=',', quoting=csv_module.QUOTE_MINIMAL, lineterminator='\n')
                    writer.writerows(rows)
                    csv_content = output.getvalue()

                    # Write CSV code block
                    f.write("```csv\n")
                    f.write(csv_content)
                    f.write("```\n\n")
                else:
                    f.write("```csv\n```\n\n")

        info(f"CSV markdown file created: {csv_md_path}")

        # Append validation metadata if enabled (using verify_csv_markdown module)
        if opts.get("csv_include_metadata", True):
            try:
                # Import verification module to append metadata
                import sys
                from pathlib import Path as PathLib

                # Add module directory to path if not already there
                verify_module_path = PathLib(__file__).parent
                if str(verify_module_path) not in sys.path:
                    sys.path.insert(0, str(verify_module_path))

                from verify_csv_markdown import append_verification_metadata_from_data

                # Call verification module to append metadata
                append_verification_metadata_from_data(
                    str(csv_md_path),
                    f"{excel_file_basename}.xlsx",
                    csv_data_dict,
                    opts
                )
            except Exception as e:
                warn(f"Failed to append verification metadata: {e}")

        return str(csv_md_path)

    except Exception as e:
        warn(f"Failed to write CSV markdown file: {e}")
        return None

def extract_print_area_for_csv(ws, area, opts, merged_lookup, cell_to_image=None):
    """Extract all cell values from print area for CSV output.

    This function processes cells within the specified print area and extracts their
    values for CSV markdown output. If a cell contains an image (indicated by
    cell_to_image mapping), a Markdown image link is generated instead of the cell value.

    Args:
        ws: Worksheet object
        area: Print area tuple (min_row, min_col, max_row, max_col) (1-based)
        opts: Options dictionary for value formatting
        merged_lookup: Merged cell lookup (only includes cells within print area)
        cell_to_image: Optional dict mapping (row, col) to image paths.
                       If provided, cells with images will output Markdown image links.

    Returns:
        List of lists (rows of cell values), where each inner list represents
        a row of cells in the print area.
    """
    min_row, min_col, max_row, max_col = area
    if cell_to_image is None:
        cell_to_image = {}
    rows = []

    for R in range(min_row, max_row + 1):
        row_vals = []
        for C in range(min_col, max_col + 1):
            # Check if this cell contains an image
            # If so, generate Markdown image link instead of cell value
            if (R, C) in cell_to_image:
                img_path = cell_to_image[(R, C)]

                # Get cell value for alt text (accessibility)
                cell = ws.cell(row=R, column=C)
                alt_text = cell_display_value(cell, opts).strip()

                # Fallback to cell reference if no meaningful alt text
                if not alt_text:
                    alt_text = f"Image at {a1_from_rc(R, C)}"

                # Create Markdown image link: ![alt text](path)
                # Encode only specific characters that break Markdown links
                encoded_path = img_path.replace('%', '%25').replace(' ', '%20').replace('(', '%28').replace(')', '%29').replace('#', '%23')
                md_image_link = f"![{alt_text}]({encoded_path})"
                row_vals.append(md_image_link)
                continue

            cell = ws.cell(row=R, column=C)

            # Handle merged cells
            tl = merged_lookup.get((R, C))
            if tl:
                if (R, C) == tl:
                    # Top-left cell: use its value
                    cell = ws.cell(row=tl[0], column=tl[1])
                    text = cell_display_value(cell, opts)
                else:
                    # Other cells in merged range
                    if opts.get("csv_apply_merge_policy", True):
                        # Apply merge_policy
                        if opts.get("merge_policy") == "top_left_only":
                            text = ""
                        else:
                            # expand/repeat: use top-left value
                            cell = ws.cell(row=tl[0], column=tl[1])
                            text = cell_display_value(cell, opts)
                    else:
                        # csv_apply_merge_policy=false: only top-left has value
                        text = ""
            else:
                text = cell_display_value(cell, opts)

            # Apply normalization if requested
            if opts.get("csv_normalize_values", True):
                text = normalize_numeric_text(text, opts)

            # Handle hyperlinks
            hl = hyperlink_info(cell)
            if hl:
                disp = text if text else (hl.get("display") or "")
                hyperlink_mode = opts.get("hyperlink_mode", "inline_plain")

                # Handle footnote/both modes -> fallback to inline_plain with warning
                if hyperlink_mode in ("footnote", "both"):
                    if not hasattr(extract_print_area_for_csv, '_csv_hyperlink_warning_shown'):
                        warn("CSV markdown does not support footnote mode. Falling back to inline_plain for hyperlinks.")
                        extract_print_area_for_csv._csv_hyperlink_warning_shown = True
                    hyperlink_mode = "inline_plain"

                if hl.get("target"):
                    link = hl["target"]
                    if hyperlink_mode == "inline":
                        text = f"[{disp}]({link})"
                    elif hyperlink_mode == "inline_plain":
                        text = f"{disp} ({link})"
                    elif hyperlink_mode == "text_only":
                        text = disp
                    else:
                        # Default to text_only if unknown mode
                        text = disp
                elif hl.get("location"):
                    loc = hl["location"]
                    if hyperlink_mode == "inline":
                        text = f"[{disp}]({loc})"
                    elif hyperlink_mode == "inline_plain":
                        text = f"{disp} (→{loc})"
                    elif hyperlink_mode == "text_only":
                        text = disp
                    else:
                        text = disp

            # Replace newlines with spaces for CSV markdown
            # This ensures 1 record = 1 physical line for AI readability
            if text and isinstance(text, str):
                text = text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')

            row_vals.append(text)
        rows.append(row_vals)

    return rows

# ===== end CSV output functions =====
