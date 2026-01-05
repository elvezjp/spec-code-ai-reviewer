#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV Markdown Verification Script
Spec Compliance: v1.6 (CSV Markdown Verification)
File: verify_csv_markdown.py

検証スクリプト - CSVマークダウンファイルの検証
- スタンドアロンモード: CLI経由で検証実行
- プログラマティックモード: Python API経由で検証実行
"""

import argparse
import json
import csv as csv_module
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

try:
    import yaml
except ImportError:
    # YAML is optional - if not available, just skip config file support
    yaml = None


def parse_excel_metadata(excel_path: str, config: Optional[Dict] = None) -> Dict:
    """Parse Excel file and extract metadata for validation.

    Args:
        excel_path: Path to Excel file
        config: Configuration dictionary (optional)

    Returns:
        Dictionary with Excel metadata structure:
        {
            "sheet_count": int,
            "sheets": [
                {"name": str, "rows": int, "cols": int},
                ...
            ]
        }
    """
    if config is None:
        config = {}

    # Default configuration
    no_print_area_mode = config.get("no_print_area_mode", "used_range")

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
        sheets_meta = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Get print area
            print_area = None
            if hasattr(ws, 'print_area') and ws.print_area:
                # Parse print area (format: "A1:D10" or "Sheet1!A1:D10")
                pa_str = ws.print_area
                if "!" in pa_str:
                    pa_str = pa_str.split("!")[-1]

                # Parse range
                match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', pa_str)
                if match:
                    def col_letter_to_num(col_str):
                        num = 0
                        for c in col_str:
                            num = num * 26 + (ord(c) - ord('A') + 1)
                        return num

                    min_col = col_letter_to_num(match.group(1))
                    min_row = int(match.group(2))
                    max_col = col_letter_to_num(match.group(3))
                    max_row = int(match.group(4))
                    print_area = (min_row, min_col, max_row, max_col)

            # If no print area, use fallback mode
            if not print_area:
                if no_print_area_mode == "skip_sheet":
                    continue
                elif no_print_area_mode == "used_range":
                    # Use used range (min/max row/col with data)
                    min_row = ws.min_row
                    max_row = ws.max_row
                    min_col = ws.min_column
                    max_col = ws.max_column
                    print_area = (min_row, min_col, max_row, max_col)
                elif no_print_area_mode == "entire_sheet_range":
                    # Use entire sheet range
                    print_area = (1, 1, ws.max_row, ws.max_column)

            if print_area:
                min_row, min_col, max_row, max_col = print_area
                rows = max_row - min_row + 1
                cols = max_col - min_col + 1

                # Convert coordinates to Excel range string (e.g., "A1:D10")
                from openpyxl.utils import get_column_letter
                start_cell = f"{get_column_letter(min_col)}{min_row}"
                end_cell = f"{get_column_letter(max_col)}{max_row}"
                excel_range = f"{start_cell}:{end_cell}"

                sheets_meta.append({
                    "name": sheet_name,
                    "rows": rows,
                    "cols": cols,
                    "range": excel_range
                })

        wb.close()

        return {
            "sheet_count": len(sheets_meta),
            "sheets": sheets_meta
        }

    except Exception as e:
        print(f"Error parsing Excel file: {e}", file=sys.stderr)
        raise


def parse_csv_markdown(csv_md_path: str) -> Dict:
    """Parse CSV markdown file and extract CSV data blocks.

    Args:
        csv_md_path: Path to CSV markdown file

    Returns:
        Dictionary with CSV metadata structure:
        {
            "sheet_count": int,
            "sheets": [
                {"name": str, "rows": int, "cols": int},
                ...
            ]
        }
    """
    try:
        with open(csv_md_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # Extract sheet sections (## SheetName)
        # Pattern: ## SheetName followed by ```csv ... ```
        sheet_pattern = re.compile(
            r'##\s+(.+?)\n\s*```csv\n(.*?)\n```',
            re.DOTALL
        )

        sheets_meta = []
        for match in sheet_pattern.finditer(content):
            sheet_name = match.group(1).strip()
            csv_content = match.group(2)

            # Parse CSV using csv.reader to handle RFC 4180 escaping
            csv_reader = csv_module.reader(csv_content.splitlines())
            rows = list(csv_reader)

            # Count rows and columns
            row_count = len(rows)
            col_count = max(len(row) for row in rows) if rows else 0

            sheets_meta.append({
                "name": sheet_name,
                "rows": row_count,
                "cols": col_count
            })

        return {
            "sheet_count": len(sheets_meta),
            "sheets": sheets_meta
        }

    except Exception as e:
        print(f"Error parsing CSV markdown file: {e}", file=sys.stderr)
        raise


def validate_metadata(excel_meta: Dict, csv_meta: Dict, excel_filename: str = None) -> Dict:
    """Validate CSV markdown against Excel metadata.

    Args:
        excel_meta: Excel metadata dictionary
        csv_meta: CSV metadata dictionary
        excel_filename: Original Excel filename (optional)

    Returns:
        Validation result dictionary with structure:
        {
            "excel": {...},
            "csv_output": {...},
            "validation": {
                "status": "OK" | "FAILED",
                "errors": [...]
            }
        }
    """
    status = "OK"
    errors = []

    # Check sheet count
    if excel_meta["sheet_count"] != csv_meta["sheet_count"]:
        status = "FAILED"
        errors.append(f"Excel has {excel_meta['sheet_count']} sheets but CSV has {csv_meta['sheet_count']} sheets")

    # Check each sheet
    excel_sheets = {s["name"]: s for s in excel_meta.get("sheets", [])}
    csv_sheets = {s["name"]: s for s in csv_meta.get("sheets", [])}

    # Check if all Excel sheets exist in CSV
    for sheet_name in excel_sheets.keys():
        if sheet_name not in csv_sheets:
            status = "FAILED"
            errors.append(f"Sheet '{sheet_name}' exists in Excel but not in CSV")

    # Check dimensions for matching sheets
    for sheet_name in excel_sheets.keys():
        if sheet_name in csv_sheets:
            e_sheet = excel_sheets[sheet_name]
            c_sheet = csv_sheets[sheet_name]
            if e_sheet["rows"] != c_sheet["rows"]:
                status = "FAILED"
                errors.append(f"Sheet '{sheet_name}': Excel has {e_sheet['rows']} rows but CSV has {c_sheet['rows']} rows")
            if e_sheet["cols"] != c_sheet["cols"]:
                status = "FAILED"
                errors.append(f"Sheet '{sheet_name}': Excel has {e_sheet['cols']} cols but CSV has {c_sheet['cols']} cols")

    # Add filename to excel_meta if provided
    excel_meta_with_filename = excel_meta.copy()
    if excel_filename:
        excel_meta_with_filename["filename"] = excel_filename

    return {
        "excel": excel_meta_with_filename,
        "csv_output": csv_meta,
        "validation": {
            "status": status,
            "errors": errors
        }
    }


def update_metadata_section(csv_md_path: str, validation_result: Dict) -> None:
    """Update validation metadata section in CSV markdown file (markdown format per spec §3.2.2).

    Args:
        csv_md_path: Path to CSV markdown file
        validation_result: Validation result dictionary
    """
    from datetime import datetime

    try:
        with open(csv_md_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # Remove existing metadata section if present
        # Pattern: --- followed by ## 検証用メタデータ ... until next --- or end of file
        metadata_pattern = re.compile(
            r'---\s*\n\s*##\s+検証用メタデータ.*?(?=\n---|$)',
            re.DOTALL
        )
        content = metadata_pattern.sub('', content).rstrip()

        # Build new markdown-formatted metadata section
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        excel_filename = validation_result["excel"].get("filename", "N/A")
        csv_filename = Path(csv_md_path).name

        metadata_lines = []
        metadata_lines.append("\n\n---\n")
        metadata_lines.append("\n## 検証用メタデータ\n\n")

        # Basic information
        metadata_lines.append(f"- **生成日時**: {timestamp}\n")
        metadata_lines.append(f"- **元Excelファイル**: {excel_filename}\n")
        metadata_lines.append(f"- **CSVマークダウンファイル**: {csv_filename}\n")
        metadata_lines.append(f"- **CSVモード**: csv_markdown\n")
        metadata_lines.append(f"- **検証フェーズ**: 1 (Basic - Row/Column Count)\n\n")

        # Per-sheet validation results
        total_cells = 0
        excel_sheets = validation_result["excel"]["sheets"]
        csv_sheets = validation_result["csv_output"]["sheets"]

        for excel_sheet in excel_sheets:
            sheet_name = excel_sheet["name"]
            excel_range = excel_sheet.get("range", "N/A")
            excel_rows = excel_sheet["rows"]
            excel_cols = excel_sheet["cols"]
            cell_count = excel_rows * excel_cols
            total_cells += cell_count

            # Find matching CSV sheet
            csv_sheet = next((s for s in csv_sheets if s["name"] == sheet_name), None)
            csv_rows = csv_sheet["rows"] if csv_sheet else 0
            csv_cols = csv_sheet["cols"] if csv_sheet else 0

            # Determine sheet status
            sheet_status = "OK"
            sheet_errors = []
            if csv_rows != excel_rows:
                sheet_status = "FAILED"
                sheet_errors.append(f"Excel has {excel_rows} rows but CSV has {csv_rows} rows")
            if csv_cols != excel_cols:
                sheet_status = "FAILED"
                sheet_errors.append(f"Excel has {excel_cols} cols but CSV has {csv_cols} cols")

            metadata_lines.append(f"### シート: {sheet_name}\n\n")
            metadata_lines.append(f"- **Excel範囲**: {excel_range}\n")
            metadata_lines.append(f"- **Excel行数**: {excel_rows}\n")
            metadata_lines.append(f"- **Excel列数**: {excel_cols}\n")
            metadata_lines.append(f"- **CSV行数**: {csv_rows}\n")
            metadata_lines.append(f"- **CSV列数**: {csv_cols}\n")
            metadata_lines.append(f"- **セル総数**: {cell_count}\n")
            metadata_lines.append(f"- **ステータス**: {sheet_status}\n")

            # Add errors if any
            for error in sheet_errors:
                metadata_lines.append(f"- **エラー**: {error}\n")

            metadata_lines.append("\n")

        # Overall validation results
        overall_status = validation_result["validation"]["status"]

        metadata_lines.append("### 全体の検証結果\n\n")
        metadata_lines.append(f"- **総セル数**: {total_cells}\n")
        metadata_lines.append(f"- **検証ステータス**: {overall_status}\n")

        # Add error list if failed
        if overall_status == "FAILED" and validation_result["validation"].get("errors"):
            metadata_lines.append("- **エラー一覧**:\n")
            for error in validation_result["validation"]["errors"]:
                metadata_lines.append(f"  - {error}\n")

        # Append metadata to content
        content += "".join(metadata_lines)

        # Write updated content
        with open(csv_md_path, 'w', encoding='utf-8') as f:
            f.write(content)

    except Exception as e:
        print(f"Error updating metadata section: {e}", file=sys.stderr)
        raise


def append_verification_metadata_from_data(
    csv_md_path: str,
    excel_filename: str,
    csv_data_dict: Dict,
    opts: Dict = None
) -> None:
    """Append verification metadata section to CSV markdown file from CSV data.

    This function is called by excel_to_md during CSV markdown generation.

    Args:
        csv_md_path: Path to CSV markdown file
        excel_filename: Original Excel filename
        csv_data_dict: Dictionary of {sheet_name: {"rows": [...], "range": "A1:D10", "area": (...)}}
        opts: Options dictionary (optional)
    """
    from datetime import datetime

    if opts is None:
        opts = {}

    try:
        # Build metadata from csv_data_dict
        excel_meta_sheets = []
        csv_meta_sheets = []

        for sheet_name, sheet_data in csv_data_dict.items():
            rows = sheet_data["rows"] if isinstance(sheet_data, dict) else sheet_data
            excel_range = sheet_data.get("range", "N/A") if isinstance(sheet_data, dict) else "N/A"

            if rows:
                row_count = len(rows)
                col_count = max(len(row) for row in rows) if rows else 0
                sheet_meta = {
                    "name": sheet_name,
                    "rows": row_count,
                    "cols": col_count,
                    "range": excel_range
                }
                excel_meta_sheets.append(sheet_meta)
                csv_meta_sheets.append(sheet_meta)

        excel_meta = {
            "sheet_count": len(excel_meta_sheets),
            "sheets": excel_meta_sheets,
            "filename": excel_filename
        }
        csv_meta = {
            "sheet_count": len(csv_meta_sheets),
            "sheets": csv_meta_sheets
        }

        # Generate validation result
        validation_result = validate_metadata(excel_meta, csv_meta, excel_filename)

        # Append metadata section to file
        update_metadata_section(csv_md_path, validation_result)

    except Exception as e:
        print(f"Warning: Failed to append verification metadata: {e}", file=sys.stderr)


def verify_csv_markdown(
    excel_path: str,
    csv_markdown_path: str,
    config: Optional[Dict] = None
) -> Dict:
    """Verify CSV markdown file against Excel source (programmatic API).

    Args:
        excel_path: Path to source Excel file
        csv_markdown_path: Path to CSV markdown file
        config: Optional configuration dictionary

    Returns:
        Validation result dictionary with structure:
        {
            "excel": {...},
            "csv_output": {...},
            "validation": {
                "status": "OK" | "FAILED"
            }
        }
    """
    if config is None:
        config = {}

    # Parse Excel metadata
    excel_meta = parse_excel_metadata(excel_path, config)

    # Parse CSV markdown metadata
    csv_meta = parse_csv_markdown(csv_markdown_path)

    # Get Excel filename
    from pathlib import Path
    excel_filename = Path(excel_path).name

    # Validate
    validation_result = validate_metadata(excel_meta, csv_meta, excel_filename)

    return validation_result


def main():
    """CLI entry point for standalone mode."""
    parser = argparse.ArgumentParser(
        description="Verify CSV markdown file against Excel source"
    )
    parser.add_argument("excel_file", help="Path to source Excel file")
    parser.add_argument("csv_markdown_file", help="Path to CSV markdown file")
    parser.add_argument(
        "--update-metadata",
        action="store_true",
        help="Write validation results to metadata section in CSV markdown file"
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Show detailed validation output"
    )
    parser.add_argument(
        "--config",
        default=None,
        help="Path to config file (YAML format, optional)"
    )

    args = parser.parse_args()

    # Check if files exist
    if not Path(args.excel_file).exists():
        print(f"Error: Excel file not found: {args.excel_file}", file=sys.stderr)
        sys.exit(2)

    if not Path(args.csv_markdown_file).exists():
        print(f"Error: CSV markdown file not found: {args.csv_markdown_file}", file=sys.stderr)
        sys.exit(2)

    # Load config if provided
    config = {}
    if args.config:
        if yaml is None:
            print("Warning: PyYAML not installed, config file will be ignored", file=sys.stderr)
        else:
            try:
                with open(args.config, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f) or {}
            except Exception as e:
                print(f"Error loading config file: {e}", file=sys.stderr)
                sys.exit(2)

    try:
        # Run verification
        validation_result = verify_csv_markdown(
            args.excel_file,
            args.csv_markdown_file,
            config
        )

        # Display results
        status = validation_result["validation"]["status"]

        if args.verbose:
            print("=" * 60)
            print("CSV Markdown Verification Results")
            print("=" * 60)
            print(f"\nExcel file: {args.excel_file}")
            print(f"CSV markdown file: {args.csv_markdown_file}")
            print(f"\nValidation status: {status}")
            print("\nExcel metadata:")
            print(json.dumps(validation_result["excel"], ensure_ascii=False, indent=2))
            print("\nCSV output metadata:")
            print(json.dumps(validation_result["csv_output"], ensure_ascii=False, indent=2))
            print("=" * 60)
        else:
            print(f"Validation status: {status}")

        # Update metadata if requested
        if args.update_metadata:
            update_metadata_section(args.csv_markdown_file, validation_result)
            print(f"Metadata section updated in: {args.csv_markdown_file}")

        # Exit code
        if status == "OK":
            sys.exit(0)
        elif status == "FAILED":
            sys.exit(1)
        else:
            sys.exit(2)

    except Exception as e:
        print(f"Error during verification: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(2)


if __name__ == "__main__":
    main()
