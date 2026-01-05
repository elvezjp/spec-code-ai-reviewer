"""
pytest fixtures for v1.7 unit tests.
Spec reference: 付録D.4
"""
import sys
from pathlib import Path
from unittest.mock import MagicMock

import pytest
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.cell.cell import Cell

# Add v1.7 to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))


# ============================================================
# Workbook/Worksheet Fixtures
# ============================================================

@pytest.fixture
def empty_workbook():
    """Create an empty Workbook."""
    return openpyxl.Workbook()


@pytest.fixture
def simple_worksheet(empty_workbook):
    """Create a Worksheet with basic data (2x2 table)."""
    ws = empty_workbook.active
    ws.title = "Sheet1"
    ws['A1'] = 'Header1'
    ws['B1'] = 'Header2'
    ws['A2'] = 'Data1'
    ws['B2'] = 'Data2'
    return ws


@pytest.fixture
def worksheet_with_numbers(empty_workbook):
    """Create a Worksheet with numeric data."""
    ws = empty_workbook.active
    ws.title = "Numbers"
    ws['A1'] = 'Name'
    ws['B1'] = 'Value'
    ws['A2'] = 'Item1'
    ws['B2'] = 100
    ws['A3'] = 'Item2'
    ws['B3'] = 200.5
    ws['A4'] = 'Item3'
    ws['B4'] = -50
    return ws


@pytest.fixture
def worksheet_with_merged_cells(empty_workbook):
    """Create a Worksheet with merged cells."""
    ws = empty_workbook.active
    ws.title = "Merged"
    ws['A1'] = 'Merged Title'
    ws.merge_cells('A1:C1')
    ws['A2'] = 'Col1'
    ws['B2'] = 'Col2'
    ws['C2'] = 'Col3'
    ws['A3'] = 'Data1'
    ws['B3'] = 'Data2'
    ws['C3'] = 'Data3'
    return ws


@pytest.fixture
def worksheet_with_empty_rows(empty_workbook):
    """Create a Worksheet with empty rows (for table splitting test)."""
    ws = empty_workbook.active
    ws.title = "EmptyRows"
    # Table 1: rows 1-2
    ws['A1'] = 'Header1'
    ws['B1'] = 'Header2'
    ws['A2'] = 'Data1'
    ws['B2'] = 'Data2'
    # Row 3 is empty
    # Table 2: rows 4-5
    ws['A4'] = 'Header3'
    ws['B4'] = 'Header4'
    ws['A5'] = 'Data3'
    ws['B5'] = 'Data4'
    return ws


@pytest.fixture
def worksheet_with_hyperlinks(empty_workbook):
    """Create a Worksheet with hyperlinks."""
    ws = empty_workbook.active
    ws.title = "Links"
    ws['A1'] = 'External Link'
    ws['A1'].hyperlink = 'https://example.com'
    ws['A2'] = 'Internal Link'
    ws['A2'].hyperlink = '#Sheet2!A1'
    ws['A3'] = 'No Link'
    return ws


@pytest.fixture
def worksheet_with_fills(empty_workbook):
    """Create a Worksheet with various fill colors."""
    ws = empty_workbook.active
    ws.title = "Fills"

    # Cell with yellow fill
    ws['A1'] = 'Yellow'
    ws['A1'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Cell with white fill (should be treated as no fill)
    ws['A2'] = 'White'
    ws['A2'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Cell with no fill
    ws['A3'] = 'NoFill'

    # Empty cell with yellow fill (should NOT be empty)
    ws['A4'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    return ws


# ============================================================
# Options Fixtures
# ============================================================

@pytest.fixture
def default_opts():
    """Create default options dictionary matching excel_to_md.py defaults."""
    return {
        "strip_whitespace": True,
        "escape_pipes": True,
        "merge_policy": "top_left_only",
        "hyperlink_mode": "inline_plain",
        "markdown_escape_level": "safe",
        "readonly_fill_policy": "assume_no_fill",
        "hidden_policy": "ignore",
        "value_mode": "display",
        "header_detection": True,
        "align_detection": True,
        "numbers_right_threshold": 0.8,
        "max_cells_per_table": 200000,
        "csv_include_description": True,
        "csv_include_metadata": True,
        "csv_normalize_values": True,
        "csv_apply_merge_policy": True,
        "mermaid_enabled": False,
        "mermaid_detect_mode": "shapes",
        "detect_dates": True,
        "date_default_format": "YYYY-MM-DD",
        "date_format_override": None,
        "numeric_thousand_sep": "keep",
        "percent_format": "keep",
        "currency_symbol": "keep",
        "percent_divide_100": False,
    }


@pytest.fixture
def opts_no_strip(default_opts):
    """Options with strip_whitespace=False."""
    opts = default_opts.copy()
    opts["strip_whitespace"] = False
    return opts


@pytest.fixture
def opts_csv_no_description(default_opts):
    """Options for CSV markdown without description (v1.7 feature)."""
    opts = default_opts.copy()
    opts["csv_include_description"] = False
    return opts


@pytest.fixture
def opts_csv_no_metadata(default_opts):
    """Options for CSV markdown without metadata."""
    opts = default_opts.copy()
    opts["csv_include_metadata"] = False
    return opts


# ============================================================
# Mock Cell Helpers
# ============================================================

def create_mock_cell(value=None, fill_color=None, hyperlink_target=None,
                     hyperlink_location=None, is_date=False):
    """
    Create a mock cell for testing.

    Args:
        value: Cell value
        fill_color: Fill color as hex string (e.g., 'FFFF00' for yellow)
        hyperlink_target: External hyperlink target URL
        hyperlink_location: Internal hyperlink location (e.g., 'Sheet2!A1')
        is_date: Whether cell should be treated as date

    Returns:
        MagicMock object simulating an openpyxl Cell
    """
    cell = MagicMock(spec=Cell)
    cell.value = value
    cell.is_date = is_date

    # Fill setup
    if fill_color:
        cell.fill = MagicMock()
        cell.fill.patternType = 'solid'
        cell.fill.fgColor = MagicMock()
        cell.fill.fgColor.rgb = fill_color
        cell.fill.fgColor.type = 'rgb'
    else:
        cell.fill = MagicMock()
        cell.fill.patternType = None

    # Hyperlink setup
    if hyperlink_target or hyperlink_location:
        cell.hyperlink = MagicMock()
        cell.hyperlink.target = hyperlink_target
        cell.hyperlink.location = hyperlink_location
        cell.hyperlink.display = str(value) if value else None
    else:
        cell.hyperlink = None

    # Border setup (no borders by default)
    cell.border = MagicMock()
    cell.border.left = MagicMock(style=None)
    cell.border.right = MagicMock(style=None)
    cell.border.top = MagicMock(style=None)
    cell.border.bottom = MagicMock(style=None)

    return cell


# ============================================================
# Test Data Fixtures
# ============================================================

@pytest.fixture
def sample_md_rows_simple():
    """Simple 2x2 markdown rows."""
    return [
        ['Header1', 'Header2'],
        ['Data1', 'Data2'],
    ]


@pytest.fixture
def sample_md_rows_with_numbers():
    """Markdown rows with numeric data."""
    return [
        ['Name', 'Value', 'Percentage'],
        ['Item1', '100', '10%'],
        ['Item2', '200', '20%'],
        ['Item3', '300', '30%'],
    ]


@pytest.fixture
def sample_md_rows_single_column():
    """Single column markdown rows (for text format detection)."""
    return [
        ['Only one column'],
        ['Another row'],
    ]


@pytest.fixture
def sample_code_rows():
    """Rows that look like source code."""
    return [
        ['public class Example {'],
        ['    private int value;'],
        ['    public void setValue(int v) {'],
        ['        this.value = v;'],
        ['    }'],
        ['}'],
    ]


@pytest.fixture
def sample_csv_data_dict():
    """Sample CSV data dictionary for write_csv_markdown tests."""
    return {
        'Sheet1': {
            'rows': [
                ['Header1', 'Header2', 'Header3'],
                ['Data1', 'Data2', 'Data3'],
                ['Data4', 'Data5', 'Data6'],
            ],
            'range': 'A1:C3',
            'area': (1, 1, 3, 3),
            'mermaid': None,
        }
    }


@pytest.fixture
def sample_csv_data_dict_multi_sheet():
    """Multi-sheet CSV data dictionary."""
    return {
        'Sheet1': {
            'rows': [
                ['A', 'B'],
                ['1', '2'],
            ],
            'range': 'A1:B2',
            'area': (1, 1, 2, 2),
            'mermaid': None,
        },
        'Sheet2': {
            'rows': [
                ['X', 'Y', 'Z'],
                ['10', '20', '30'],
            ],
            'range': 'A1:C2',
            'area': (1, 1, 2, 3),
            'mermaid': None,
        },
    }


# ============================================================
# Grid Fixtures for Table Detection
# ============================================================

@pytest.fixture
def grid_3x3_full():
    """3x3 grid with all cells non-empty."""
    return [
        [1, 1, 1],
        [1, 1, 1],
        [1, 1, 1],
    ]


@pytest.fixture
def grid_with_empty_row():
    """Grid with empty row in the middle (should split into 2 tables)."""
    return [
        [1, 1, 1],
        [0, 0, 0],  # Empty row
        [1, 1, 1],
    ]


@pytest.fixture
def grid_with_empty_col():
    """Grid with empty column in the middle."""
    return [
        [1, 0, 1],
        [1, 0, 1],
        [1, 0, 1],
    ]


@pytest.fixture
def grid_l_shape():
    """L-shaped grid (should be single connected component)."""
    return [
        [1, 1, 1],
        [1, 0, 0],
        [1, 0, 0],
    ]


@pytest.fixture
def grid_two_separate():
    """Two separate regions (should be 2 tables)."""
    return [
        [1, 1, 0, 0],
        [1, 1, 0, 0],
        [0, 0, 1, 1],
        [0, 0, 1, 1],
    ]


@pytest.fixture
def grid_empty():
    """Completely empty grid."""
    return [
        [0, 0, 0],
        [0, 0, 0],
        [0, 0, 0],
    ]
