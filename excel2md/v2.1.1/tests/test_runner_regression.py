"""Regression tests for runner-level bugs.

Issue #24: extract_table truncation path must return a 4-tuple (was 3-tuple).
Issue #25: footnote numbering must be unique across tables within the configured scope.
"""
import re
import sys
import tempfile
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))

from excel2md.cli import build_argparser
from excel2md.runner import run


def _parse_args(argv):
    return build_argparser().parse_args(argv)


def _make_two_table_workbook(path: Path, with_links: bool = True) -> None:
    """Two tables on one sheet, separated by an empty row.

    When ``with_links`` is True, each table contains exactly one external
    hyperlink so footnote numbering is exercised.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["A2"] = "Link1"
    if with_links:
        ws["A2"].hyperlink = "https://example.com/a"
    ws["B2"] = "Data1"
    # Row 3 is empty -> tables split
    ws["A4"] = "Header3"
    ws["B4"] = "Header4"
    ws["A5"] = "Link2"
    if with_links:
        ws["A5"].hyperlink = "https://example.com/b"
    ws["B5"] = "Data2"
    wb.save(path)


def _make_multi_sheet_workbook(path: Path) -> None:
    """Two sheets, each with a table containing a hyperlink."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "S1"
    ws1["A1"] = "H1"
    ws1["B1"] = "H2"
    ws1["A2"] = "L1"
    ws1["A2"].hyperlink = "https://example.com/s1"
    ws1["B2"] = "V1"

    ws2 = wb.create_sheet("S2")
    ws2["A1"] = "H1"
    ws2["B1"] = "H2"
    ws2["A2"] = "L2"
    ws2["A2"].hyperlink = "https://example.com/s2"
    ws2["B2"] = "V2"
    wb.save(path)


# ============================================================
# Issue #24: truncation return tuple must be 4 elements
# ============================================================

class TestIssue24Truncation:
    """runner.run() must not crash when max_cells_per_table forces truncation."""

    def test_truncation_does_not_crash_runner(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx = Path(tmpdir) / "trunc.xlsx"
            _make_two_table_workbook(xlsx, with_links=False)
            out_md = Path(tmpdir) / "out.md"

            args = _parse_args([
                str(xlsx),
                "-o", str(out_md),
                "--no-csv-markdown-enabled",
                "--max-cells-per-table", "2",
            ])

            # Pre-fix this raised ValueError: not enough values to unpack
            result = run(str(xlsx), str(out_md), args)
            assert result is not None
            assert Path(result).exists()
            text = Path(result).read_text(encoding="utf-8")
            assert "max_cells_per_table" in text


# ============================================================
# Issue #25: footnote IDs unique across tables
# ============================================================

_REF_RE = re.compile(r"\[\^(\d+)\]")
_DEF_RE = re.compile(r"^\[\^(\d+)\]:\s*(.+)$", re.MULTILINE)


def _footnote_refs_and_defs(text: str):
    refs = [int(n) for n in _REF_RE.findall(text)]
    defs = [(int(n), body.strip()) for n, body in _DEF_RE.findall(text)]
    return refs, defs


class TestIssue25FootnoteNumbering:
    """Footnote numbers must be unique and sequential within the configured scope."""

    def test_book_scope_footnotes_are_unique_and_sequential(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx = Path(tmpdir) / "book.xlsx"
            _make_two_table_workbook(xlsx, with_links=True)
            out_md = Path(tmpdir) / "out.md"

            args = _parse_args([
                str(xlsx),
                "-o", str(out_md),
                "--no-csv-markdown-enabled",
                "--hyperlink-mode", "footnote",
                "--footnote-scope", "book",
            ])
            run(str(xlsx), str(out_md), args)

            text = out_md.read_text(encoding="utf-8")
            refs, defs = _footnote_refs_and_defs(text)

            # Two hyperlinks -> two distinct footnote refs (in body and definitions)
            assert sorted(set(refs)) == [1, 2], f"refs={refs}"
            # Definitions must be unique and contain the two distinct URLs
            assert sorted(n for n, _ in defs) == [1, 2]
            def_urls = {body for _, body in defs}
            assert def_urls == {"https://example.com/a", "https://example.com/b"}

    def test_sheet_scope_resets_per_sheet(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx = Path(tmpdir) / "sheet.xlsx"
            _make_multi_sheet_workbook(xlsx)
            out_md = Path(tmpdir) / "out.md"

            args = _parse_args([
                str(xlsx),
                "-o", str(out_md),
                "--no-csv-markdown-enabled",
                "--hyperlink-mode", "footnote",
                "--footnote-scope", "sheet",
            ])
            run(str(xlsx), str(out_md), args)

            text = out_md.read_text(encoding="utf-8")
            refs, defs = _footnote_refs_and_defs(text)

            # Each sheet has one link; with sheet scope each sheet starts at 1.
            assert refs.count(1) >= 2, f"refs={refs}"
            assert defs == [
                (1, "https://example.com/s1"),
                (1, "https://example.com/s2"),
            ]
